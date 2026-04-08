"""
Fase 3: Importar dados extraidos pelo Claude Code e gerar relatorio Excel.

Le os arquivos _extracted.json de cada processo, compara com o estado atual
do banco, atualiza campos vazios e gera relatorio Excel antes/depois.

Uso:
    # Dry-run (apenas mostra o que seria alterado)
    python scripts/importar_analise_diarias.py

    # Executar de fato
    python scripts/importar_analise_diarias.py --executar

    # Processar apenas um processo
    python scripts/importar_analise_diarias.py --processo 00002.003034/2026-83 --executar

    # Apenas gerar relatorio (sem importar)
    python scripts/importar_analise_diarias.py --relatorio-only
"""

import sys
import os
import json
import time
import argparse
from datetime import datetime, date
from decimal import Decimal, InvalidOperation

import requests
import urllib3

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

if sys.platform == 'win32':
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
    sys.stderr.reconfigure(encoding='utf-8', errors='replace')

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

from app import create_app
from app.extensions import db

# Diretorio dos documentos
DEST_DIR = os.path.normpath(
    os.path.join(os.path.dirname(__file__), '..', '..', 'diarias')
)


# ═══════════════════════════════════════════════════════════════════════
# SNAPSHOT DO ESTADO ATUAL
# ═══════════════════════════════════════════════════════════════════════

def capturar_estado(itinerario):
    """Captura snapshot completo do estado atual de um itinerario no banco."""
    from app.models.diaria import (
        DiariasItemItinerario, DiariasDocumentoSei,
        DiariasQuadroOrcamentario, DiariasJustificativa,
    )

    if not itinerario:
        return {'exists': False}

    # Dados basicos
    estado = {
        'exists': True,
        'id': itinerario.id,
        'sei_protocolo': itinerario.sei_protocolo,
        'tipo_solicitacao_id': itinerario.tipo_solicitacao_id,
        'tipo_itinerario': itinerario.tipo_itinerario,
        'data_viagem': str(itinerario.data_viagem) if itinerario.data_viagem else None,
        'data_retorno': str(itinerario.data_retorno) if itinerario.data_retorno else None,
        'origem': itinerario.origem,
        'estado_origem': itinerario.estado_origem,
        'estado_destino': itinerario.estado_destino,
        'objetivo': itinerario.objetivo[:80] if itinerario.objetivo else None,
        'valor_total': str(itinerario.valor_total) if itinerario.valor_total else None,
        'etapa_atual_id': itinerario.etapa_atual_id,
        'qtd_diarias_solicitadas': itinerario.qtd_diarias_solicitadas,
    }

    # Pessoas
    pessoas = itinerario.itens.all()
    estado['qtd_integrantes'] = len(pessoas)
    estado['integrantes'] = [
        {
            'cpf': p.cpf_pessoa,
            'nome': p.nome_pessoa,
            'cargo_id': p.cargo_id,
            'valor_cargo': str(p.valor_cargo) if p.valor_cargo else None,
            'banco_agencia': p.banco_agencia,
            'banco_conta': p.banco_conta,
        }
        for p in pessoas
    ]

    # Quadro orcamentario
    qo = itinerario.quadro_orcamentario
    if qo and qo.ug:
        estado['quadro_orcamentario'] = {
            'ug': qo.ug,
            'funcao': qo.funcao,
            'subfuncao': qo.subfuncao,
            'programa': qo.programa,
            'fonte_recursos': qo.fonte_recursos,
            'natureza_despesa': qo.natureza_despesa,
        }
    else:
        estado['quadro_orcamentario'] = None

    # Documentos SEI
    docs = {d.tipo_documento: d.sei_id for d in itinerario.documentos_sei}
    estado['documentos_sei'] = docs
    estado['qtd_docs_sei'] = len([v for v in docs.values() if v])

    return estado


# ═══════════════════════════════════════════════════════════════════════
# IMPORTACAO DOS DADOS EXTRAIDOS
# ═══════════════════════════════════════════════════════════════════════

def safe_decimal(valor):
    """Converte valor para Decimal de forma segura."""
    if valor is None:
        return None
    try:
        # Remover R$, pontos de milhar, trocar virgula por ponto
        s = str(valor).replace('R$', '').replace(' ', '').strip()
        s = s.replace('.', '').replace(',', '.')
        return Decimal(s)
    except (InvalidOperation, ValueError):
        return None


def safe_date(valor):
    """Converte string para date/datetime."""
    if not valor:
        return None
    for fmt in ('%Y-%m-%d %H:%M', '%Y-%m-%dT%H:%M', '%Y-%m-%d', '%d/%m/%Y', '%d/%m/%Y %H:%M'):
        try:
            return datetime.strptime(str(valor).strip(), fmt)
        except ValueError:
            continue
    return None


def importar_processo(data, executar=False):
    """Importa dados de um _extracted.json para o banco."""
    from app.models.diaria import (
        DiariasItinerario, DiariasItemItinerario, DiariasDocumentoSei,
        DiariasQuadroOrcamentario, DiariasJustificativa,
    )

    protocolo = data.get('processo')
    if not protocolo:
        return {'protocolo': '?', 'status': 'erro', 'msg': 'Campo processo ausente'}

    # Buscar itinerario existente
    itinerario = DiariasItinerario.query.filter_by(sei_protocolo=protocolo).first()
    if not itinerario:
        # Tentar por n_processo
        itinerario = DiariasItinerario.query.filter_by(n_processo=protocolo).first()

    if not itinerario:
        return {'protocolo': protocolo, 'status': 'nao_encontrado',
                'msg': 'Itinerario nao encontrado no banco'}

    # Snapshot antes
    estado_antes = capturar_estado(itinerario)
    campos_alterados = []
    itin_data = data.get('itinerario', {})

    # ── Atualizar campos basicos (apenas se NULL/vazio no banco) ──
    mapa_campos = {
        'tipo_solicitacao_id': 'tipo_solicitacao_id',
        'tipo_itinerario': 'tipo_itinerario',
        'objetivo': 'objetivo',
        'valor_total': 'valor_total',
        'estado_origem': 'estado_origem',
        'estado_destino': 'estado_destino',
        'origem': 'origem',
        'qtd_diarias_solicitadas': 'qtd_diarias_solicitadas',
    }

    for json_campo, db_campo in mapa_campos.items():
        novo_valor = itin_data.get(json_campo)
        if novo_valor is None:
            continue

        atual = getattr(itinerario, db_campo, None)

        # Verificar se campo esta vazio/nulo
        vazio = atual is None
        if not vazio and db_campo == 'objetivo':
            vazio = not str(atual).strip()
        if not vazio and db_campo == 'valor_total':
            vazio = float(atual or 0) == 0

        if not vazio:
            continue  # Nao sobrescrever dados existentes

        # Converter tipos especiais
        if db_campo == 'valor_total':
            novo_valor = safe_decimal(novo_valor)
        elif db_campo in ('tipo_solicitacao_id', 'tipo_itinerario', 'estado_origem', 'estado_destino'):
            novo_valor = int(novo_valor) if novo_valor else None
        elif db_campo == 'qtd_diarias_solicitadas':
            novo_valor = float(novo_valor) if novo_valor else None

        if novo_valor is not None:
            if executar:
                setattr(itinerario, db_campo, novo_valor)
            campos_alterados.append({
                'campo': db_campo,
                'antes': str(atual),
                'depois': str(novo_valor),
            })

    # Datas
    for campo_json, campo_db in [('data_viagem', 'data_viagem'), ('data_retorno', 'data_retorno')]:
        novo = itin_data.get(campo_json)
        if not novo:
            continue
        atual = getattr(itinerario, campo_db, None)
        if atual:
            continue
        dt = safe_date(novo)
        if dt:
            if executar:
                setattr(itinerario, campo_db, dt)
            campos_alterados.append({
                'campo': campo_db,
                'antes': None,
                'depois': str(dt),
            })

    # ── Documentos SEI ──
    docs_data = data.get('documentos_sei', {})
    for tipo_doc, doc_info in docs_data.items():
        if not doc_info:
            continue

        # Aceitar tanto dict quanto string simples
        if isinstance(doc_info, str):
            sei_id = doc_info
            sei_formatado = None
            codigo = None
        else:
            sei_id = doc_info.get('sei_id')
            sei_formatado = doc_info.get('sei_formatado')
            codigo = doc_info.get('codigo')

        if not sei_id:
            continue

        # Verificar se ja existe
        if itinerario.has_doc(tipo_doc):
            continue

        if executar:
            itinerario.set_doc(tipo_doc, sei_id=sei_id, sei_formatado=sei_formatado, codigo=codigo)

        campos_alterados.append({
            'campo': f'doc_sei.{tipo_doc}',
            'antes': None,
            'depois': sei_formatado or sei_id,
        })

    # ── Quadro Orcamentario ──
    qo_data = data.get('quadro_orcamentario', {})
    if qo_data:
        qo = itinerario.quadro_orcamentario
        qo_existia = qo is not None and qo.ug is not None

        if not qo_existia:
            if not qo:
                if executar:
                    qo = DiariasQuadroOrcamentario(itinerario_id=itinerario.id)
                    db.session.add(qo)

            for campo in ('ug', 'funcao', 'subfuncao', 'programa', 'plano_interno',
                          'fonte_recursos', 'natureza_despesa'):
                novo = qo_data.get(campo)
                if novo:
                    if executar and qo:
                        setattr(qo, campo, str(novo))
                    campos_alterados.append({
                        'campo': f'quadro.{campo}',
                        'antes': None,
                        'depois': str(novo),
                    })

            for campo in ('valor_inicial_nr', 'saldo_nr', 'valor_despesa', 'saldo_atual_nr'):
                novo = qo_data.get(campo)
                if novo:
                    dec = safe_decimal(novo)
                    if dec and executar and qo:
                        setattr(qo, campo, dec)
                    if dec:
                        campos_alterados.append({
                            'campo': f'quadro.{campo}',
                            'antes': None,
                            'depois': str(dec),
                        })

    # ── Integrantes (atualizar campos faltantes) ──
    integrantes_data = data.get('integrantes', [])
    if integrantes_data:
        pessoas_existentes = {p.cpf_pessoa: p for p in itinerario.itens.all()}

        for int_data in integrantes_data:
            cpf = int_data.get('cpf', '').strip()
            if not cpf:
                continue

            pessoa = pessoas_existentes.get(cpf)
            if not pessoa:
                # Pessoa nao existe — nao criamos novos, apenas atualizamos existentes
                continue

            # Atualizar campos vazios
            for json_campo, db_campo in [
                ('nome', 'nome_pessoa'),
                ('matricula', 'matricula_pessoa'),
                ('banco_agencia', 'banco_agencia'),
                ('banco_conta', 'banco_conta'),
                ('vinculo', 'vinculo'),
                ('cargo_folha', 'cargo_folha'),
                ('setor', 'setor'),
                ('orgao', 'orgao'),
            ]:
                novo = int_data.get(json_campo)
                if not novo:
                    continue
                atual = getattr(pessoa, db_campo, None)
                if atual and str(atual).strip():
                    continue
                if executar:
                    setattr(pessoa, db_campo, str(novo).strip())
                campos_alterados.append({
                    'campo': f'integrante[{cpf}].{db_campo}',
                    'antes': atual,
                    'depois': str(novo).strip(),
                })

            # Valor cargo
            valor = int_data.get('valor_unitario') or int_data.get('valor_cargo')
            if valor:
                dec = safe_decimal(valor)
                if dec and (not pessoa.valor_cargo or float(pessoa.valor_cargo) == 0):
                    if executar:
                        pessoa.valor_cargo = dec
                    campos_alterados.append({
                        'campo': f'integrante[{cpf}].valor_cargo',
                        'antes': str(pessoa.valor_cargo),
                        'depois': str(dec),
                    })

    # ── Commit ──
    if executar and campos_alterados:
        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            return {
                'protocolo': protocolo,
                'status': 'erro_db',
                'msg': str(e),
                'campos_alterados': campos_alterados,
            }

    estado_depois = capturar_estado(itinerario) if executar else None

    return {
        'protocolo': protocolo,
        'status': 'ok' if campos_alterados else 'sem_alteracoes',
        'campos_alterados': campos_alterados,
        'qtd_alteracoes': len(campos_alterados),
        'estado_antes': estado_antes,
        'estado_depois': estado_depois,
        'confidence': data.get('confidence', '?'),
        'notes': data.get('notes', ''),
    }


# ═══════════════════════════════════════════════════════════════════════
# RELATORIO EXCEL
# ═══════════════════════════════════════════════════════════════════════

def gerar_relatorio_excel(resultados, output_path):
    """Gera relatorio Excel com 3 planilhas."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()

    # ── Estilos ──
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
    ok_fill = PatternFill(start_color='D5F5E3', end_color='D5F5E3', fill_type='solid')
    warn_fill = PatternFill(start_color='FCF3CF', end_color='FCF3CF', fill_type='solid')
    err_fill = PatternFill(start_color='FADBD8', end_color='FADBD8', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    def style_header(ws, row=1):
        for cell in ws[row]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

    # ═══ PLANILHA 1: RESUMO ═══
    ws1 = wb.active
    ws1.title = 'Resumo'
    ws1.append([
        '#', 'Processo', 'Status', 'Confidence', 'Integrantes',
        'Valor Total', 'Etapa', 'Campos Atualizados', 'Notas',
    ])
    style_header(ws1)

    for i, r in enumerate(resultados, 1):
        antes = r.get('estado_antes', {})
        ws1.append([
            i,
            r['protocolo'],
            r['status'],
            r.get('confidence', '?'),
            antes.get('qtd_integrantes', 0),
            antes.get('valor_total', ''),
            antes.get('etapa_atual_id', ''),
            r.get('qtd_alteracoes', 0),
            r.get('notes', '')[:100],
        ])
        row = ws1.max_row
        status = r['status']
        fill = ok_fill if status == 'ok' else warn_fill if status == 'sem_alteracoes' else err_fill
        for cell in ws1[row]:
            cell.border = thin_border
            if cell.column == 3:  # Status
                cell.fill = fill

    # Larguras
    for col, w in enumerate([5, 25, 16, 12, 12, 15, 8, 18, 40], 1):
        ws1.column_dimensions[ws1.cell(row=1, column=col).column_letter].width = w

    # ═══ PLANILHA 2: ANTES/DEPOIS ═══
    ws2 = wb.create_sheet('Antes_Depois')
    ws2.append(['Processo', 'Campo', 'Antes', 'Depois', 'Mudou'])
    style_header(ws2)

    for r in resultados:
        for alt in r.get('campos_alterados', []):
            ws2.append([
                r['protocolo'],
                alt['campo'],
                str(alt.get('antes', '')),
                str(alt.get('depois', '')),
                'SIM',
            ])
            for cell in ws2[ws2.max_row]:
                cell.border = thin_border

    for col, w in enumerate([25, 30, 30, 30, 8], 1):
        ws2.column_dimensions[ws2.cell(row=1, column=col).column_letter].width = w

    # ═══ PLANILHA 3: ERROS/AVISOS ═══
    ws3 = wb.create_sheet('Erros_Avisos')
    ws3.append(['Processo', 'Problema', 'Severidade'])
    style_header(ws3)

    for r in resultados:
        if r['status'] == 'nao_encontrado':
            ws3.append([r['protocolo'], 'Itinerario nao encontrado no banco', 'ERRO'])
            ws3[ws3.max_row][2].fill = err_fill
        elif r['status'] == 'erro_db':
            ws3.append([r['protocolo'], f"Erro DB: {r.get('msg', '')}", 'ERRO'])
            ws3[ws3.max_row][2].fill = err_fill
        elif r['status'] == 'erro':
            ws3.append([r['protocolo'], r.get('msg', 'Erro desconhecido'), 'ERRO'])
            ws3[ws3.max_row][2].fill = err_fill
        elif r.get('confidence') == 'low':
            ws3.append([r['protocolo'], f"Confianca baixa: {r.get('notes', '')}", 'AVISO'])
            ws3[ws3.max_row][2].fill = warn_fill
        elif r['status'] == 'sem_alteracoes':
            ws3.append([r['protocolo'], 'Nenhum campo novo extraido', 'INFO'])
            ws3[ws3.max_row][2].fill = ok_fill

        for cell in ws3[ws3.max_row]:
            cell.border = thin_border

    for col, w in enumerate([25, 60, 12], 1):
        ws3.column_dimensions[ws3.cell(row=1, column=col).column_letter].width = w

    wb.save(output_path)
    print(f"Relatorio salvo em: {output_path}")
    return output_path


# ═══════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════

def listar_extracted():
    """Lista todos os _extracted.json disponiveis."""
    extraidos = []
    if not os.path.isdir(DEST_DIR):
        return extraidos

    for nome in sorted(os.listdir(DEST_DIR)):
        caminho = os.path.join(DEST_DIR, nome)
        if not os.path.isdir(caminho):
            continue
        extracted = os.path.join(caminho, '_extracted.json')
        if os.path.exists(extracted):
            extraidos.append({
                'pasta': nome,
                'protocolo': nome.replace('_', '/'),
                'caminho': extracted,
            })
    return extraidos


def main():
    parser = argparse.ArgumentParser(description='Importar dados extraidos e gerar relatorio Excel')
    parser.add_argument('--executar', action='store_true', help='Gravar alteracoes no banco')
    parser.add_argument('--processo', help='Processar apenas este protocolo')
    parser.add_argument('--relatorio-only', action='store_true', help='Apenas gerar relatorio sem importar')
    args = parser.parse_args()

    app = create_app()
    with app.app_context():
        # Listar JSONs disponiveis
        extraidos = listar_extracted()
        if not extraidos:
            print(f"Nenhum _extracted.json encontrado em: {DEST_DIR}")
            print("Execute primeiro a analise com o Claude Code (Fase 2).")
            return

        # Filtrar se especificou processo
        if args.processo:
            pasta_alvo = args.processo.replace('/', '_')
            extraidos = [e for e in extraidos if e['pasta'] == pasta_alvo]
            if not extraidos:
                print(f"Processo {args.processo} nao encontrado.")
                return

        print(f"{'='*60}")
        print(f"IMPORTACAO DE DADOS EXTRAIDOS")
        print(f"{'='*60}")
        print(f"  Processos com _extracted.json: {len(extraidos)}")
        print(f"  Modo: {'EXECUTAR' if args.executar else 'DRY-RUN'}")
        print()

        resultados = []
        inicio = time.time()

        for i, ext in enumerate(extraidos, 1):
            print(f"[{i}/{len(extraidos)}] {ext['protocolo']}...", end=' ')

            try:
                with open(ext['caminho'], 'r', encoding='utf-8') as f:
                    data = json.load(f)
            except (json.JSONDecodeError, IOError) as e:
                print(f"ERRO ao ler JSON: {e}")
                resultados.append({
                    'protocolo': ext['protocolo'],
                    'status': 'erro',
                    'msg': f'Erro JSON: {e}',
                    'campos_alterados': [],
                    'qtd_alteracoes': 0,
                })
                continue

            resultado = importar_processo(data, executar=args.executar)
            resultados.append(resultado)

            status = resultado['status']
            qtd = resultado.get('qtd_alteracoes', 0)
            if status == 'ok':
                print(f"OK ({qtd} campo(s))")
            elif status == 'sem_alteracoes':
                print("sem novidades")
            else:
                print(f"{status}: {resultado.get('msg', '')}")

        duracao = time.time() - inicio

        # Estatisticas
        ok = sum(1 for r in resultados if r['status'] == 'ok')
        sem_alt = sum(1 for r in resultados if r['status'] == 'sem_alteracoes')
        erros = sum(1 for r in resultados if r['status'] not in ('ok', 'sem_alteracoes'))
        total_campos = sum(r.get('qtd_alteracoes', 0) for r in resultados)

        print(f"\n{'='*60}")
        print(f"RESUMO ({duracao:.1f}s)")
        print(f"  Atualizados: {ok}")
        print(f"  Sem alteracoes: {sem_alt}")
        print(f"  Erros: {erros}")
        print(f"  Total campos alterados: {total_campos}")

        # Gerar relatorio Excel
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        modo = 'exec' if args.executar else 'dryrun'
        excel_path = os.path.join(DEST_DIR, f'relatorio_diarias_{modo}_{timestamp}.xlsx')
        gerar_relatorio_excel(resultados, excel_path)


if __name__ == '__main__':
    main()
