'''
Relatorio 03 - Vinculacoes (Novas + Sobrescritas)
===================================================
Gera Excel detalhando as vinculacoes que serao inseridas e
as que serao sobrescritas, incluindo valores originais.

Uso: python scripts/relatorio_03_vinculacoes.py
'''
import os
import unicodedata
from datetime import datetime

import pandas as pd
import pymysql
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
load_dotenv(os.path.join(BASE_DIR, '.env'))

VINC_PATH = os.path.join(BASE_DIR, 'itens vinculação.xlsx')
OUTPUT = os.path.join(BASE_DIR, f'relatorio_03_vinculacoes_{datetime.now():%Y%m%d_%H%M}.xlsx')

DB = dict(host=os.getenv('DB_HOST','localhost'), user=os.getenv('DB_USER','root'),
          password=os.getenv('DB_PASS',''), database=os.getenv('DB_NAME','sgc'), charset='utf8mb4')

HDR_FILL = PatternFill('solid', fgColor='1F4E79')
HDR_FONT = Font(bold=True, color='FFFFFF', size=11)
TITLE_FONT = Font(bold=True, size=14, color='1F4E79')
LABEL_FONT = Font(bold=True, size=11)
THIN = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
FILL_GREEN = PatternFill('solid', fgColor='E2EFDA')
FILL_YELLOW = PatternFill('solid', fgColor='FFF9C4')
FILL_GRAY = PatternFill('solid', fgColor='F2F2F2')


def norm(t):
    if not t: return ''
    return unicodedata.normalize('NFKD', str(t)).encode('ascii','ignore').decode().upper().strip()


def norm_tipo(t):
    n = norm(t)
    return 'S' if 'SERVIC' in n else 'M'


def write_hdr(ws, row, headers):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = HDR_FONT; c.fill = HDR_FILL; c.alignment = CENTER; c.border = THIN


def write_row(ws, row, values, fill=None):
    for i, v in enumerate(values, 1):
        c = ws.cell(row=row, column=i, value=v)
        c.border = THIN
        c.alignment = Alignment(vertical='center', wrap_text=isinstance(v, str) and len(str(v)) > 30)
        if fill:
            c.fill = fill


def main():
    print('Relatorio 03 - Vinculacoes')
    print('='*50)

    # Ler Excel
    print('[1/3] Lendo arquivo Excel...')
    df = pd.read_excel(VINC_PATH)
    registros = []
    for _, row in df.iterrows():
        siafe = row.iloc[0]
        item_id = row.iloc[1]
        tipo = row.iloc[2]
        if pd.isna(siafe) or str(siafe).strip() in ('', '-'):
            continue
        if pd.isna(item_id):
            continue
        registros.append({
            'siafe': str(int(float(str(siafe)))),
            'id': int(float(str(item_id))),
            'tipo': norm_tipo(tipo),
            'tipo_original': str(tipo).strip(),
        })

    # Dedup
    dedup = {}
    for r in registros:
        k = (r['siafe'], r['tipo'], r['id'])
        dedup[k] = r
    print(f'  Registros unicos: {len(dedup)}')

    # Consultar BD
    print('[2/3] Consultando banco de dados...')
    conn = pymysql.connect(**DB)
    cur = conn.cursor()

    cur.execute("SELECT codigo, numeroOriginal, nomeContratado, objeto FROM contratos")
    contratos = {}
    for r in cur.fetchall():
        contratos[str(r[0])] = {'numeroOriginal': r[1] or '', 'contratado': r[2] or '', 'objeto': r[3] or ''}

    cur.execute("SELECT codigo_servico, nome FROM catserv_servicos")
    catserv = {r[0]: r[1] for r in cur.fetchall()}

    cur.execute("SELECT codigo, descricao FROM catmat_itens")
    catmat_itens = {r[0]: r[1] for r in cur.fetchall()}

    cur.execute("SELECT codigo, nome FROM catmat_pdms")
    catmat_pdms = {r[0]: r[1] for r in cur.fetchall()}

    cur.execute("""
        SELECT id, codigo_contrato, tipo, catserv_servico_id, catmat_item_id
        FROM itens_vinculados
    """)
    db_vinc = cur.fetchall()
    conn.close()

    # Indexar vinculados existentes por (contrato, tipo, id_catalogo)
    db_idx = {}
    db_by_contrato = {}
    for r in db_vinc:
        cod = str(r[1])
        cat_id = r[3] if r[2] == 'S' else r[4]
        db_idx[(cod, r[2], cat_id)] = r
        db_by_contrato.setdefault((cod, r[2]), []).append(r)

    # Classificar
    novas = []
    sobrescritas = []
    ja_existem = []
    ignoradas = []

    for k, v in dedup.items():
        siafe, tipo, item_id = k
        if siafe not in contratos:
            ignoradas.append({**v, 'motivo': 'Contrato inexistente'})
            continue
        if tipo == 'S' and item_id not in catserv:
            ignoradas.append({**v, 'motivo': 'Servico inexistente no CATSERV'})
            continue
        if tipo == 'M' and item_id not in catmat_itens and item_id not in catmat_pdms:
            ignoradas.append({**v, 'motivo': 'Material inexistente no CATMAT'})
            continue

        db_key = (siafe, tipo, item_id)
        if db_key in db_idx:
            ja_existem.append(v)
        else:
            existing = db_by_contrato.get((siafe, tipo), [])
            encontrou_diferente = False
            for ex in existing:
                old_id = ex[3] if tipo == 'S' else ex[4]
                if old_id != item_id:
                    old_desc = ''
                    new_desc = ''
                    if tipo == 'S':
                        old_desc = catserv.get(old_id, '(nao encontrado)')
                        new_desc = catserv.get(item_id, '(nao encontrado)')
                    else:
                        old_desc = catmat_itens.get(old_id, {}) if isinstance(catmat_itens.get(old_id), str) else catmat_itens.get(old_id, '')
                        if not old_desc and old_id:
                            old_desc = catmat_pdms.get(old_id, '')
                        new_desc = catmat_itens.get(item_id, '')
                        if not new_desc:
                            new_desc = catmat_pdms.get(item_id, '')
                    sobrescritas.append({
                        'siafe': siafe, 'tipo': tipo,
                        'old_id': old_id, 'old_desc': old_desc,
                        'new_id': item_id, 'new_desc': new_desc,
                        'db_row_id': ex[0],
                    })
                    encontrou_diferente = True
                    break
            if not encontrou_diferente:
                novas.append(v)

    print(f'  Novas: {len(novas)}')
    print(f'  Sobrescritas: {len(sobrescritas)}')
    print(f'  Ja existem: {len(ja_existem)}')
    print(f'  Ignoradas: {len(ignoradas)}')

    # Gerar Excel
    print('[3/3] Gerando relatorio...')
    wb = Workbook()

    # --- Aba Resumo ---
    ws0 = wb.active
    ws0.title = 'Resumo'
    ws0.sheet_properties.tabColor = '4472C4'
    ws0.merge_cells('A1:C1')
    ws0['A1'].value = 'RELATORIO DE VINCULACOES'
    ws0['A1'].font = TITLE_FONT
    ws0['A3'].value = f'Data: {datetime.now():%d/%m/%Y %H:%M}'
    ws0['A3'].font = LABEL_FONT
    ws0['A4'].value = f'Arquivo: itens vinculacao.xlsx'

    resumo = [
        ('Total linhas no Excel (dedup)', len(dedup)),
        ('Novas vinculacoes (INSERT)', len(novas)),
        ('Vinculacoes sobrescritas (UPDATE)', len(sobrescritas)),
        ('Ja existem (sem alteracao)', len(ja_existem)),
        ('Ignoradas (contrato/catalogo inexistente)', len(ignoradas)),
    ]
    for i, (label, val) in enumerate(resumo, 6):
        ws0.cell(row=i, column=1, value=label).font = LABEL_FONT
        ws0.cell(row=i, column=2, value=val)
    ws0.column_dimensions['A'].width = 45
    ws0.column_dimensions['B'].width = 12

    # --- Aba Sobrescritas ---
    ws1 = wb.create_sheet('Sobrescritas (UPDATE)')
    ws1.sheet_properties.tabColor = 'ED7D31'
    ws1.merge_cells('A1:H1')
    ws1['A1'].value = 'VINCULACOES QUE SERAO SOBRESCRITAS (valores originais vs novos)'
    ws1['A1'].font = TITLE_FONT
    ws1['A2'].value = f'Total: {len(sobrescritas)} registros'
    ws1['A2'].font = LABEL_FONT

    headers = ['N SIAFE', 'Num. Original', 'Contratado', 'Tipo',
               'ID Anterior (BD)', 'Descricao Anterior (BD)',
               'ID Novo (Excel)', 'Descricao Novo (Excel)']
    write_hdr(ws1, 4, headers)
    for i, s in enumerate(sobrescritas, 5):
        ct = contratos.get(s['siafe'], {})
        write_row(ws1, i, [
            s['siafe'], ct.get('numeroOriginal', ''), ct.get('contratado', ''),
            'Servico' if s['tipo'] == 'S' else 'Material',
            s['old_id'], s['old_desc'] or '',
            s['new_id'], s['new_desc'] or '',
        ], fill=FILL_YELLOW)
    if sobrescritas:
        ws1.auto_filter.ref = f'A4:{get_column_letter(len(headers))}{ws1.max_row}'
    ws1.freeze_panes = 'A5'
    for col, w in zip('ABCDEFGH', [14, 18, 30, 12, 16, 40, 16, 40]):
        ws1.column_dimensions[col].width = w

    # --- Aba Novas ---
    ws2 = wb.create_sheet('Novas (INSERT)')
    ws2.sheet_properties.tabColor = '00B050'
    ws2.merge_cells('A1:F1')
    ws2['A1'].value = 'NOVAS VINCULACOES A INSERIR'
    ws2['A1'].font = TITLE_FONT
    ws2['A2'].value = f'Total: {len(novas)} registros'
    ws2['A2'].font = LABEL_FONT

    headers2 = ['N SIAFE', 'Num. Original', 'Contratado', 'Tipo', 'ID Catalogo', 'Descricao Catalogo']
    write_hdr(ws2, 4, headers2)
    for i, n in enumerate(novas, 5):
        ct = contratos.get(n['siafe'], {})
        desc = ''
        if n['tipo'] == 'S':
            desc = catserv.get(n['id'], '')
        else:
            desc = catmat_itens.get(n['id'], '')
            if not desc:
                desc = catmat_pdms.get(n['id'], '')
        write_row(ws2, i, [
            n['siafe'], ct.get('numeroOriginal', ''), ct.get('contratado', ''),
            'Servico' if n['tipo'] == 'S' else 'Material',
            n['id'], desc,
        ], fill=FILL_GREEN)
    if novas:
        ws2.auto_filter.ref = f'A4:{get_column_letter(len(headers2))}{ws2.max_row}'
    ws2.freeze_panes = 'A5'
    for col, w in zip('ABCDEF', [14, 18, 30, 12, 16, 45]):
        ws2.column_dimensions[col].width = w

    # --- Aba Ignoradas ---
    ws3 = wb.create_sheet('Ignoradas')
    ws3.sheet_properties.tabColor = 'A6A6A6'
    ws3.merge_cells('A1:D1')
    ws3['A1'].value = 'VINCULACOES IGNORADAS'
    ws3['A1'].font = TITLE_FONT
    ws3['A2'].value = f'Total: {len(ignoradas)} registros'
    ws3['A2'].font = LABEL_FONT

    headers3 = ['N SIAFE', 'Tipo', 'ID', 'Motivo']
    write_hdr(ws3, 4, headers3)
    for i, ig in enumerate(ignoradas, 5):
        write_row(ws3, i, [
            ig['siafe'], 'Servico' if ig['tipo'] == 'S' else 'Material',
            ig['id'], ig['motivo'],
        ], fill=FILL_GRAY)
    if ignoradas:
        ws3.auto_filter.ref = f'A4:{get_column_letter(len(headers3))}{ws3.max_row}'
    ws3.freeze_panes = 'A5'
    for col, w in zip('ABCD', [14, 12, 16, 45]):
        ws3.column_dimensions[col].width = w

    # --- Aba Ja existem ---
    ws4 = wb.create_sheet('Ja existem (sem alteracao)')
    ws4.sheet_properties.tabColor = '4472C4'
    ws4.merge_cells('A1:E1')
    ws4['A1'].value = 'VINCULACOES QUE JA EXISTEM NO BANCO (sem alteracao necessaria)'
    ws4['A1'].font = TITLE_FONT
    ws4['A2'].value = f'Total: {len(ja_existem)} registros'
    ws4['A2'].font = LABEL_FONT

    headers4 = ['N SIAFE', 'Num. Original', 'Tipo', 'ID Catalogo', 'Descricao']
    write_hdr(ws4, 4, headers4)
    for i, je in enumerate(ja_existem, 5):
        ct = contratos.get(je['siafe'], {})
        desc = ''
        if je['tipo'] == 'S':
            desc = catserv.get(je['id'], '')
        else:
            desc = catmat_itens.get(je['id'], '')
            if not desc:
                desc = catmat_pdms.get(je['id'], '')
        write_row(ws4, i, [
            je['siafe'], ct.get('numeroOriginal', ''),
            'Servico' if je['tipo'] == 'S' else 'Material',
            je['id'], desc,
        ])
    if ja_existem:
        ws4.auto_filter.ref = f'A4:{get_column_letter(len(headers4))}{ws4.max_row}'
    ws4.freeze_panes = 'A5'
    for col, w in zip('ABCDE', [14, 18, 12, 16, 45]):
        ws4.column_dimensions[col].width = w

    wb.save(OUTPUT)
    print(f'\nSalvo em: {OUTPUT}')
    print('='*50)


if __name__ == '__main__':
    main()
