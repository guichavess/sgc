'''
Relatorio de Importacao - Contratos SGC
========================================
Gera Excel com 8 abas detalhando o resultado da importacao
dos arquivos "itens vinculacao.xlsx" e "itens correcao.xlsx".

Uso: python scripts/relatorio_importacao.py
'''
import os
import sys
import unicodedata
from datetime import datetime

import pymysql
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side, numbers
)
from openpyxl.utils import get_column_letter

# ────────────────────── Configuracao ──────────────────────
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
ENV_PATH = os.path.join(BASE_DIR, '.env')
VINCULACAO_PATH = os.path.join(BASE_DIR, 'itens vinculação.xlsx')
CORRECAO_PATH = os.path.join(BASE_DIR, 'itens correção.xlsx')
TODAY = datetime.now().strftime('%Y%m%d_%H%M')
OUTPUT_PATH = os.path.join(BASE_DIR, f'relatorio_importacao_{TODAY}.xlsx')

load_dotenv(ENV_PATH)

DB_CONFIG = {
    'host': os.getenv('DB_HOST', 'localhost'),
    'user': os.getenv('DB_USER', 'root'),
    'password': os.getenv('DB_PASS', ''),
    'database': os.getenv('DB_NAME', 'sgc'),
    'charset': 'utf8mb4',
}

# ────────────────────── Estilos ──────────────────────
HEADER_FILL = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
TITLE_FONT = Font(bold=True, size=16, color='1F4E79')
SUBTITLE_FONT = Font(bold=True, size=12, color='1F4E79')
SECTION_FONT = Font(bold=True, size=13, color='1F4E79')
LABEL_FONT = Font(bold=True, size=11)
VALUE_FONT = Font(size=11)
WRAP_ALIGN = Alignment(vertical='center', wrap_text=True)
CENTER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)
BRL_FORMAT = '#,##0.00'

FILL_GREEN = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
FILL_RED = PatternFill(start_color='FCE4EC', end_color='FCE4EC', fill_type='solid')
FILL_YELLOW = PatternFill(start_color='FFF9C4', end_color='FFF9C4', fill_type='solid')
FILL_BLUE = PatternFill(start_color='DBEEF4', end_color='DBEEF4', fill_type='solid')
FILL_GRAY = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

TAB_GREEN = '00B050'
TAB_RED = 'FF0000'
TAB_YELLOW = 'FFC000'
TAB_BLUE = '4472C4'
TAB_ORANGE = 'ED7D31'
TAB_GRAY = 'A6A6A6'
TAB_PURPLE = '7030A0'


# ────────────────────── Helpers ──────────────────────
def norm(texto):
    if not texto:
        return ''
    return unicodedata.normalize('NFKD', str(texto)).encode('ascii', 'ignore').decode().upper().strip()


def norm_tipo(tipo_str):
    t = norm(tipo_str)
    return 'S' if 'SERVIC' in t else 'M'


def auto_width(ws, min_w=10, max_w=55):
    for col_cells in ws.columns:
        letter = get_column_letter(col_cells[0].column)
        mx = min_w
        for c in col_cells:
            if c.value is not None:
                mx = max(mx, len(str(c.value)))
        ws.column_dimensions[letter].width = min(mx + 3, max_w)


def hdr(ws, row, headers):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER_ALIGN
        c.border = THIN_BORDER


def wrow(ws, row, values, currency_cols=None, int_cols=None, fill=None):
    currency_cols = currency_cols or set()
    int_cols = int_cols or set()
    for i, v in enumerate(values, 1):
        c = ws.cell(row=row, column=i, value=v)
        c.border = THIN_BORDER
        c.alignment = Alignment(vertical='center', wrap_text=(i == 1 or isinstance(v, str) and len(str(v)) > 30))
        if i in currency_cols and v is not None:
            c.number_format = BRL_FORMAT
        if i in int_cols and v is not None:
            c.number_format = '0'
        if fill:
            c.fill = fill


def autofilter(ws, hdr_row, ncols):
    ws.auto_filter.ref = f'A{hdr_row}:{get_column_letter(ncols)}{ws.max_row}'


def summary_cell(ws, row, col, label, value, is_currency=False):
    cl = ws.cell(row=row, column=col, value=label)
    cl.font = LABEL_FONT
    cv = ws.cell(row=row, column=col + 1, value=value)
    cv.font = VALUE_FONT
    if is_currency:
        cv.number_format = BRL_FORMAT


# ────────────────────── Leitura dos Excel ──────────────────────
def read_vinculacao():
    """Le itens vinculacao.xlsx e retorna lista de dicts."""
    wb = load_workbook(VINCULACAO_PATH, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()
    result = []
    for r in rows:
        siafe = r[0]
        item_id = r[1]
        tipo = r[2]
        if siafe is None or str(siafe).strip() in ('', '-'):
            continue
        if item_id is None:
            continue
        result.append({
            'siafe': str(int(float(str(siafe)))),
            'id': int(float(str(item_id))),
            'tipo': norm_tipo(tipo),
            'tipo_original': str(tipo).strip(),
        })
    return result


def read_correcao():
    """Le itens correcao.xlsx e retorna lista de dicts."""
    wb = load_workbook(CORRECAO_PATH, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()
    result = []
    for r in rows:
        siafe = r[0]
        item = r[1]
        data_valor = r[2]
        valor = r[3]
        qtde = r[4]
        contratado = r[5]
        if siafe is None:
            continue
        # Parse data
        if isinstance(data_valor, datetime):
            dt = data_valor
        else:
            dt = None
        # Parse qtde
        try:
            q = int(float(str(qtde)))
        except (ValueError, TypeError):
            q = None
        # Parse valor
        try:
            v = float(valor) if valor is not None else None
        except (ValueError, TypeError):
            v = None

        result.append({
            'siafe': str(int(float(str(siafe)))),
            'item': str(item).strip() if item else '',
            'data': dt,
            'mes': dt.month if dt else None,
            'ano': dt.year if dt else None,
            'valor': v,
            'qtde': q,
            'contratado': str(contratado).strip() if contratado else '',
        })
    return result


# ────────────────────── Queries BD ──────────────────────
def get_conn():
    return pymysql.connect(**DB_CONFIG, cursorclass=pymysql.cursors.DictCursor)


def fetch_contratos(conn):
    """Retorna dict codigo -> {numero_contrato, contratado, objeto, situacao}"""
    with conn.cursor() as cur:
        cur.execute("""
            SELECT c.codigo, c.numero_contrato, cr.nome AS contratado,
                   c.objeto, s.descricao AS situacao
            FROM contratos c
            LEFT JOIN credor cr ON c.credor_id = cr.id
            LEFT JOIN situacao s ON c.situacao_id = s.id
        """)
        return {str(r['codigo']): r for r in cur.fetchall()}


def fetch_catserv(conn):
    with conn.cursor() as cur:
        cur.execute("SELECT codigo_servico, descricao FROM catserv_servicos")
        return {r['codigo_servico']: r['descricao'] for r in cur.fetchall()}


def fetch_catmat_itens(conn):
    with conn.cursor() as cur:
        cur.execute("SELECT id, codigo, descricao FROM catmat_itens")
        by_codigo = {}
        for r in cur.fetchall():
            by_codigo[r['codigo']] = r
        return by_codigo


def fetch_catmat_pdms(conn):
    with conn.cursor() as cur:
        cur.execute("SELECT id, codigo, descricao FROM catmat_pdms")
        return {r['codigo']: r for r in cur.fetchall()}


def fetch_vinculados(conn):
    """Retorna lista de dicts com vinculacoes existentes."""
    with conn.cursor() as cur:
        cur.execute("""
            SELECT iv.id, iv.codigo_contrato, iv.tipo,
                   iv.catserv_servico_id, iv.catmat_item_id,
                   iv.item_contrato_id,
                   cs.descricao AS desc_catserv,
                   ci.descricao AS desc_catmat
            FROM itens_vinculados iv
            LEFT JOIN catserv_servicos cs ON iv.catserv_servico_id = cs.codigo_servico
            LEFT JOIN catmat_itens ci ON iv.catmat_item_id = ci.id
        """)
        return cur.fetchall()


def fetch_execucoes(conn):
    """Retorna lista de dicts com execucoes existentes."""
    with conn.cursor() as cur:
        cur.execute("""
            SELECT e.id, e.codigo_contrato, e.itens_contrato_id,
                   e.data, e.valor, e.quantidade, e.mes, e.ano, e.tipo,
                   e.catserv_servico_id, e.catmat_item_id,
                   ic.descricao AS desc_item
            FROM execucoes e
            LEFT JOIN itens_contrato ic ON e.itens_contrato_id = ic.id
        """)
        return cur.fetchall()


def fetch_itens_contrato(conn):
    with conn.cursor() as cur:
        cur.execute("""
            SELECT ic.id, ic.descricao, ic.tipo_item, ic.categoria_id,
                   cc.codigo_contrato
            FROM itens_contrato ic
            LEFT JOIN categoria_contrato cc ON ic.categoria_id = cc.id
        """)
        return cur.fetchall()


# ────────────────────── Analise ──────────────────────
def analisar(vinc_data, corr_data, contratos, catserv, catmat_itens, catmat_pdms,
             db_vinculados, db_execucoes, db_itens_contrato):
    resultado = {}

    # --- Contratos nao encontrados ---
    siafes_vinc = set(v['siafe'] for v in vinc_data)
    siafes_corr = set(c['siafe'] for c in corr_data)
    todos_siafes = siafes_vinc | siafes_corr

    contratos_nao_encontrados = []
    for s in sorted(todos_siafes):
        if s not in contratos:
            origem = []
            if s in siafes_vinc:
                origem.append('Vinculacao')
            if s in siafes_corr:
                origem.append('Correcao')
            contratos_nao_encontrados.append({
                'siafe': s,
                'origem': ', '.join(origem),
            })
    resultado['contratos_nao_encontrados'] = contratos_nao_encontrados

    # --- Material nao encontrado ---
    mat_nao_encontrados = []
    mat_ids_checked = set()
    for v in vinc_data:
        if v['tipo'] == 'M' and v['id'] not in mat_ids_checked:
            mat_ids_checked.add(v['id'])
            found_in = None
            desc = ''
            if v['id'] in catmat_itens:
                found_in = 'catmat_itens'
                desc = catmat_itens[v['id']]['descricao']
            elif v['id'] in catmat_pdms:
                found_in = 'catmat_pdms'
                desc = catmat_pdms[v['id']]['descricao']
            if not found_in:
                # Pegar contratos que referenciam esse ID
                contratos_ref = [vv['siafe'] for vv in vinc_data if vv['id'] == v['id']]
                mat_nao_encontrados.append({
                    'id': v['id'],
                    'contratos': ', '.join(sorted(set(contratos_ref))),
                    'qtd_contratos': len(set(contratos_ref)),
                })
    resultado['material_nao_encontrado'] = mat_nao_encontrados

    # --- Servico nao encontrado ---
    serv_nao_encontrados = []
    serv_ids_checked = set()
    for v in vinc_data:
        if v['tipo'] == 'S' and v['id'] not in serv_ids_checked:
            serv_ids_checked.add(v['id'])
            if v['id'] not in catserv:
                contratos_ref = [vv['siafe'] for vv in vinc_data if vv['id'] == v['id']]
                serv_nao_encontrados.append({
                    'id': v['id'],
                    'contratos': ', '.join(sorted(set(contratos_ref))),
                    'qtd_contratos': len(set(contratos_ref)),
                })
    resultado['servico_nao_encontrado'] = serv_nao_encontrados

    # --- Vinculacoes: indice do BD ---
    db_vinc_idx = {}
    for dv in db_vinculados:
        cod = str(dv['codigo_contrato'])
        cat_id = dv['catserv_servico_id'] if dv['tipo'] == 'S' else dv['catmat_item_id']
        key = (cod, dv['tipo'], cat_id)
        db_vinc_idx[key] = dv
    # Indice por (contrato, tipo) para detectar sobreposicoes
    db_vinc_by_ct = {}
    for dv in db_vinculados:
        cod = str(dv['codigo_contrato'])
        k = (cod, dv['tipo'])
        db_vinc_by_ct.setdefault(k, []).append(dv)

    vinc_novas = []
    vinc_sobrescritas = []
    vinc_ja_existem = []
    vinc_ignoradas_contrato = []
    vinc_ignoradas_catalogo = []

    # Dedup vinculacao
    vinc_dedup = {}
    for v in vinc_data:
        k = (v['siafe'], v['tipo'], v['id'])
        vinc_dedup[k] = v

    for k, v in vinc_dedup.items():
        siafe, tipo, item_id = k
        # Contrato existe?
        if siafe not in contratos:
            vinc_ignoradas_contrato.append(v)
            continue
        # Catalogo existe?
        if tipo == 'S' and item_id not in catserv:
            vinc_ignoradas_catalogo.append(v)
            continue
        if tipo == 'M' and item_id not in catmat_itens and item_id not in catmat_pdms:
            vinc_ignoradas_catalogo.append(v)
            continue

        # Verificar se ja existe no BD
        db_key = (siafe, tipo, item_id)
        if db_key in db_vinc_idx:
            vinc_ja_existem.append({**v, 'db': db_vinc_idx[db_key]})
        else:
            # Verificar se existe vinculacao diferente para esse contrato+tipo
            existing = db_vinc_by_ct.get((siafe, tipo), [])
            if existing:
                # Tem vinculacao existente mas com ID diferente - sobrescrever
                for ex in existing:
                    old_id = ex['catserv_servico_id'] if tipo == 'S' else ex['catmat_item_id']
                    if old_id != item_id:
                        old_desc = ''
                        new_desc = ''
                        if tipo == 'S':
                            old_desc = catserv.get(old_id, '(nao encontrado)')
                            new_desc = catserv.get(item_id, '(nao encontrado)')
                        else:
                            if old_id and old_id in catmat_itens:
                                old_desc = catmat_itens[old_id]['descricao']
                            elif old_id and old_id in catmat_pdms:
                                old_desc = catmat_pdms[old_id]['descricao']
                            if item_id in catmat_itens:
                                new_desc = catmat_itens[item_id]['descricao']
                            elif item_id in catmat_pdms:
                                new_desc = catmat_pdms[item_id]['descricao']
                        vinc_sobrescritas.append({
                            'siafe': siafe,
                            'tipo': tipo,
                            'contrato': contratos.get(siafe, {}),
                            'old_id': old_id,
                            'old_desc': old_desc,
                            'new_id': item_id,
                            'new_desc': new_desc,
                            'db_row_id': ex['id'],
                        })
                        break
                else:
                    vinc_novas.append(v)
            else:
                vinc_novas.append(v)

    resultado['vinc_novas'] = vinc_novas
    resultado['vinc_sobrescritas'] = vinc_sobrescritas
    resultado['vinc_ja_existem'] = vinc_ja_existem
    resultado['vinc_ignoradas_contrato'] = vinc_ignoradas_contrato
    resultado['vinc_ignoradas_catalogo'] = vinc_ignoradas_catalogo

    # --- Execucoes ---
    # Indice do BD: (contrato, item_norm, mes, ano) -> row
    ic_by_id = {r['id']: r for r in db_itens_contrato}
    db_exec_idx = {}
    for ex in db_execucoes:
        desc = ''
        if ex['itens_contrato_id'] and ex['itens_contrato_id'] in ic_by_id:
            desc = norm(ic_by_id[ex['itens_contrato_id']]['descricao'])
        elif ex['desc_item']:
            desc = norm(ex['desc_item'])
        key = (str(ex['codigo_contrato']), desc, ex['mes'], ex['ano'])
        db_exec_idx[key] = ex

    # Indice itens_contrato por (contrato, desc_norm)
    ic_by_contrato_desc = {}
    for ic in db_itens_contrato:
        if ic['codigo_contrato']:
            k = (str(ic['codigo_contrato']), norm(ic['descricao']))
            ic_by_contrato_desc[k] = ic

    exec_novas = []
    exec_atualizadas = []
    exec_iguais = []
    exec_sem_valor = []
    exec_ignoradas = []

    for c in corr_data:
        if c['siafe'] not in contratos:
            exec_ignoradas.append(c)
            continue
        if c['valor'] is None and c['qtde'] is None:
            exec_sem_valor.append(c)
            continue

        key = (c['siafe'], norm(c['item']), c['mes'], c['ano'])
        if key in db_exec_idx:
            db_row = db_exec_idx[key]
            db_val = float(db_row['valor']) if db_row['valor'] is not None else None
            db_qtd = int(db_row['quantidade']) if db_row['quantidade'] is not None else None
            xl_val = c['valor']
            xl_qtd = c['qtde']
            val_diff = abs((db_val or 0) - (xl_val or 0)) > 0.005
            qtd_diff = (db_qtd or 0) != (xl_qtd or 0)
            if val_diff or qtd_diff:
                exec_atualizadas.append({
                    **c,
                    'db_id': db_row['id'],
                    'db_valor': db_val,
                    'db_qtde': db_qtd,
                    'diff_valor': val_diff,
                    'diff_qtde': qtd_diff,
                })
            else:
                exec_iguais.append(c)
        else:
            exec_novas.append(c)

    resultado['exec_novas'] = exec_novas
    resultado['exec_atualizadas'] = exec_atualizadas
    resultado['exec_iguais'] = exec_iguais
    resultado['exec_sem_valor'] = exec_sem_valor
    resultado['exec_ignoradas'] = exec_ignoradas

    # Totalizadores
    resultado['totais'] = {
        'vinc_total_excel': len(vinc_dedup),
        'vinc_novas': len(vinc_novas),
        'vinc_sobrescritas': len(vinc_sobrescritas),
        'vinc_ja_existem': len(vinc_ja_existem),
        'vinc_ignoradas': len(vinc_ignoradas_contrato) + len(vinc_ignoradas_catalogo),
        'exec_total_excel': len(corr_data),
        'exec_novas': len(exec_novas),
        'exec_atualizadas': len(exec_atualizadas),
        'exec_iguais': len(exec_iguais),
        'exec_sem_valor': len(exec_sem_valor),
        'exec_ignoradas': len(exec_ignoradas),
        'contratos_nao_encontrados': len(contratos_nao_encontrados),
        'material_nao_encontrado': len(mat_nao_encontrados),
        'servico_nao_encontrado': len(serv_nao_encontrados),
    }
    return resultado


# ────────────────────── Gerar Excel ──────────────────────
def gerar_relatorio(resultado, contratos, catserv, catmat_itens, catmat_pdms):
    wb = Workbook()
    t = resultado['totais']

    # ====== ABA 1: RESUMO ======
    ws = wb.active
    ws.title = '1. Resumo Geral'
    ws.sheet_properties.tabColor = TAB_BLUE

    ws.merge_cells('A1:F1')
    c = ws.cell(row=1, column=1, value='RELATORIO DE IMPORTACAO - SGC Contratos')
    c.font = TITLE_FONT
    ws.cell(row=2, column=1, value=f'Data: {datetime.now().strftime("%d/%m/%Y %H:%M")}').font = VALUE_FONT
    ws.cell(row=3, column=1, value=f'Arquivo vinculacao: itens vinculacao.xlsx').font = VALUE_FONT
    ws.cell(row=4, column=1, value=f'Arquivo execucoes: itens correcao.xlsx').font = VALUE_FONT

    row = 6
    ws.cell(row=row, column=1, value='VINCULACOES (itens vinculacao.xlsx)').font = SECTION_FONT
    row += 1
    summary_cell(ws, row, 1, 'Total de linhas no Excel (dedup):', t['vinc_total_excel'])
    row += 1
    summary_cell(ws, row, 1, 'Novas vinculacoes a inserir:', t['vinc_novas'])
    row += 1
    summary_cell(ws, row, 1, 'Vinculacoes a sobrescrever:', t['vinc_sobrescritas'])
    row += 1
    summary_cell(ws, row, 1, 'Ja existem (sem alteracao):', t['vinc_ja_existem'])
    row += 1
    summary_cell(ws, row, 1, 'Ignoradas (contrato/catalogo inexistente):', t['vinc_ignoradas'])

    row += 2
    ws.cell(row=row, column=1, value='EXECUCOES (itens correcao.xlsx)').font = SECTION_FONT
    row += 1
    summary_cell(ws, row, 1, 'Total de linhas no Excel:', t['exec_total_excel'])
    row += 1
    summary_cell(ws, row, 1, 'Novas execucoes a inserir:', t['exec_novas'])
    row += 1
    summary_cell(ws, row, 1, 'Execucoes a atualizar (valor/qtde diferente):', t['exec_atualizadas'])
    row += 1
    summary_cell(ws, row, 1, 'Ja existem (sem alteracao):', t['exec_iguais'])
    row += 1
    summary_cell(ws, row, 1, 'Sem valor/quantidade (ignoradas):', t['exec_sem_valor'])
    row += 1
    summary_cell(ws, row, 1, 'Contrato inexistente (ignoradas):', t['exec_ignoradas'])

    row += 2
    ws.cell(row=row, column=1, value='PROBLEMAS ENCONTRADOS').font = SECTION_FONT
    row += 1
    summary_cell(ws, row, 1, 'Contratos nao encontrados no banco:', t['contratos_nao_encontrados'])
    row += 1
    summary_cell(ws, row, 1, 'IDs de material nao encontrados (CATMAT):', t['material_nao_encontrado'])
    row += 1
    summary_cell(ws, row, 1, 'IDs de servico nao encontrados (CATSERV):', t['servico_nao_encontrado'])

    ws.column_dimensions['A'].width = 48
    ws.column_dimensions['B'].width = 15

    # ====== ABA 2: CONTRATOS NAO ENCONTRADOS ======
    ws2 = wb.create_sheet('2. Contratos Inexistentes')
    ws2.sheet_properties.tabColor = TAB_RED
    headers = ['N SIAFE', 'Origem (Arquivo)']
    hdr(ws2, 1, headers)
    for i, c_nf in enumerate(resultado['contratos_nao_encontrados'], 2):
        wrow(ws2, i, [c_nf['siafe'], c_nf['origem']])
    autofilter(ws2, 1, len(headers))
    ws2.freeze_panes = 'A2'
    auto_width(ws2)

    # ====== ABA 3: MATERIAL NAO ENCONTRADO ======
    ws3 = wb.create_sheet('3. Material Inexistente')
    ws3.sheet_properties.tabColor = TAB_RED
    headers = ['ID (Codigo)', 'Qtd Contratos Referenciando', 'Contratos (N SIAFE)']
    hdr(ws3, 1, headers)
    for i, m in enumerate(resultado['material_nao_encontrado'], 2):
        wrow(ws3, i, [m['id'], m['qtd_contratos'], m['contratos']])
    autofilter(ws3, 1, len(headers))
    ws3.freeze_panes = 'A2'
    auto_width(ws3)

    # ====== ABA 4: SERVICO NAO ENCONTRADO ======
    ws4 = wb.create_sheet('4. Servico Inexistente')
    ws4.sheet_properties.tabColor = TAB_RED
    headers = ['ID (Codigo)', 'Qtd Contratos Referenciando', 'Contratos (N SIAFE)']
    hdr(ws4, 1, headers)
    for i, s in enumerate(resultado['servico_nao_encontrado'], 2):
        wrow(ws4, i, [s['id'], s['qtd_contratos'], s['contratos']])
    autofilter(ws4, 1, len(headers))
    ws4.freeze_panes = 'A2'
    auto_width(ws4)

    # ====== ABA 5: VINCULACOES SOBRESCRITAS ======
    ws5 = wb.create_sheet('5. Vinculacoes Sobrescritas')
    ws5.sheet_properties.tabColor = TAB_ORANGE
    headers = ['N SIAFE', 'Num Contrato', 'Contratado', 'Tipo',
               'ID Anterior', 'Descricao Anterior',
               'ID Novo', 'Descricao Novo']
    hdr(ws5, 1, headers)
    for i, vs in enumerate(resultado['vinc_sobrescritas'], 2):
        ct = vs.get('contrato', {})
        wrow(ws5, i, [
            vs['siafe'],
            ct.get('numero_contrato', ''),
            ct.get('contratado', ''),
            'Servico' if vs['tipo'] == 'S' else 'Material',
            vs['old_id'],
            vs['old_desc'] or '',
            vs['new_id'],
            vs['new_desc'] or '',
        ], fill=FILL_YELLOW)
    autofilter(ws5, 1, len(headers))
    ws5.freeze_panes = 'A2'
    auto_width(ws5)

    # ====== ABA 6: VINCULACOES NOVAS ======
    ws6 = wb.create_sheet('6. Vinculacoes Novas')
    ws6.sheet_properties.tabColor = TAB_GREEN
    headers = ['N SIAFE', 'Num Contrato', 'Contratado', 'Tipo', 'ID Catalogo', 'Descricao Catalogo']
    hdr(ws6, 1, headers)
    for i, vn in enumerate(resultado['vinc_novas'], 2):
        ct = contratos.get(vn['siafe'], {})
        desc = ''
        if vn['tipo'] == 'S':
            desc = catserv.get(vn['id'], '')
        else:
            if vn['id'] in catmat_itens:
                desc = catmat_itens[vn['id']]['descricao']
            elif vn['id'] in catmat_pdms:
                desc = catmat_pdms[vn['id']]['descricao']
        wrow(ws6, i, [
            vn['siafe'],
            ct.get('numero_contrato', ''),
            ct.get('contratado', ''),
            'Servico' if vn['tipo'] == 'S' else 'Material',
            vn['id'],
            desc,
        ], fill=FILL_GREEN)
    autofilter(ws6, 1, len(headers))
    ws6.freeze_panes = 'A2'
    auto_width(ws6)

    # ====== ABA 7: EXECUCOES ATUALIZADAS ======
    ws7 = wb.create_sheet('7. Execucoes Atualizadas')
    ws7.sheet_properties.tabColor = TAB_ORANGE
    headers = ['N SIAFE', 'Item', 'Mes', 'Ano', 'Contratado',
               'Valor Anterior (BD)', 'Valor Novo (Excel)', 'Dif. Valor',
               'Qtde Anterior (BD)', 'Qtde Novo (Excel)', 'Dif. Qtde']
    hdr(ws7, 1, headers)
    currency = {6, 7, 8}
    intcols = {3, 4, 9, 10, 11}
    for i, ea in enumerate(resultado['exec_atualizadas'], 2):
        dv = (ea['valor'] or 0) - (ea['db_valor'] or 0) if ea['diff_valor'] else 0
        dq = (ea['qtde'] or 0) - (ea['db_qtde'] or 0) if ea['diff_qtde'] else 0
        wrow(ws7, i, [
            ea['siafe'], ea['item'], ea['mes'], ea['ano'], ea['contratado'],
            ea['db_valor'], ea['valor'], round(dv, 2),
            ea['db_qtde'], ea['qtde'], dq,
        ], currency_cols=currency, int_cols=intcols, fill=FILL_YELLOW)
    autofilter(ws7, 1, len(headers))
    ws7.freeze_panes = 'A2'
    auto_width(ws7)

    # ====== ABA 8: EXECUCOES NOVAS ======
    ws8 = wb.create_sheet('8. Execucoes Novas')
    ws8.sheet_properties.tabColor = TAB_GREEN
    headers = ['N SIAFE', 'Num Contrato', 'Item', 'Mes', 'Ano',
               'Valor', 'Qtde', 'Contratado']
    hdr(ws8, 1, headers)
    currency = {6}
    intcols = {4, 5, 7}
    for i, en in enumerate(resultado['exec_novas'], 2):
        ct = contratos.get(en['siafe'], {})
        wrow(ws8, i, [
            en['siafe'], ct.get('numero_contrato', ''), en['item'],
            en['mes'], en['ano'], en['valor'], en['qtde'], en['contratado'],
        ], currency_cols=currency, int_cols=intcols, fill=FILL_GREEN)
    autofilter(ws8, 1, len(headers))
    ws8.freeze_panes = 'A2'
    auto_width(ws8)

    wb.save(OUTPUT_PATH)
    return OUTPUT_PATH


# ────────────────────── Main ──────────────────────
def main():
    print('='*60)
    print('  RELATORIO DE IMPORTACAO - SGC Contratos')
    print('='*60)

    print('\n[1/4] Lendo arquivos Excel...')
    vinc_data = read_vinculacao()
    print(f'  itens vinculacao: {len(vinc_data)} linhas validas')
    corr_data = read_correcao()
    print(f'  itens correcao: {len(corr_data)} linhas')

    print('\n[2/4] Consultando banco de dados...')
    conn = get_conn()
    contratos = fetch_contratos(conn)
    print(f'  contratos: {len(contratos)}')
    catserv = fetch_catserv(conn)
    print(f'  catserv_servicos: {len(catserv)}')
    catmat_itens = fetch_catmat_itens(conn)
    print(f'  catmat_itens: {len(catmat_itens)}')
    catmat_pdms = fetch_catmat_pdms(conn)
    print(f'  catmat_pdms: {len(catmat_pdms)}')
    db_vinculados = fetch_vinculados(conn)
    print(f'  itens_vinculados: {len(db_vinculados)}')
    db_execucoes = fetch_execucoes(conn)
    print(f'  execucoes: {len(db_execucoes)}')
    db_itens_contrato = fetch_itens_contrato(conn)
    print(f'  itens_contrato: {len(db_itens_contrato)}')
    conn.close()

    print('\n[3/4] Analisando dados...')
    resultado = analisar(
        vinc_data, corr_data, contratos, catserv,
        catmat_itens, catmat_pdms, db_vinculados, db_execucoes, db_itens_contrato
    )

    t = resultado['totais']
    print(f'\n  --- VINCULACOES ---')
    print(f'  Novas: {t["vinc_novas"]}')
    print(f'  Sobrescrever: {t["vinc_sobrescritas"]}')
    print(f'  Ja existem: {t["vinc_ja_existem"]}')
    print(f'  Ignoradas: {t["vinc_ignoradas"]}')
    print(f'\n  --- EXECUCOES ---')
    print(f'  Novas: {t["exec_novas"]}')
    print(f'  Atualizar: {t["exec_atualizadas"]}')
    print(f'  Iguais: {t["exec_iguais"]}')
    print(f'  Sem valor: {t["exec_sem_valor"]}')
    print(f'  Ignoradas: {t["exec_ignoradas"]}')
    print(f'\n  --- PROBLEMAS ---')
    print(f'  Contratos inexistentes: {t["contratos_nao_encontrados"]}')
    print(f'  Material inexistente: {t["material_nao_encontrado"]}')
    print(f'  Servico inexistente: {t["servico_nao_encontrado"]}')

    print('\n[4/4] Gerando relatorio Excel...')
    path = gerar_relatorio(resultado, contratos, catserv, catmat_itens, catmat_pdms)
    print(f'\n  Relatorio salvo em: {path}')
    print('='*60)


if __name__ == '__main__':
    main()
