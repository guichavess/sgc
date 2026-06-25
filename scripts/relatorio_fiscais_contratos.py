"""
Relatório de Fiscais de Contratos
==================================
Exporta todos os contratos do módulo "Execução de Contratos" com a lista
detalhada de fiscais associados (uma linha por fiscal):

  - Código do contrato
  - Nome e código do contratado
  - Tipo do fiscal (Fiscal, Gestor, Suplente, etc.)
  - Nome, CPF, telefone, e-mail, registro profissional
  - Data da última atualização do fiscal

Contratos sem fiscal aparecem uma vez com colunas de fiscal em branco.

Uso:
  python scripts/relatorio_fiscais_contratos.py
  python scripts/relatorio_fiscais_contratos.py --situacao=VIGENTE
  python scripts/relatorio_fiscais_contratos.py --com-fiscais   # exclui contratos sem fiscal
"""
import os
import sys
from datetime import date
import pymysql
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv

# ── Configuração ───────────────────────────────────────────────────────────────
BASE_DIR    = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
RELATORIOS  = r'C:\Users\guilh\OneDrive\Documentos\SEAD\Projetos\Nova pasta\Relatórios'
os.makedirs(RELATORIOS, exist_ok=True)
load_dotenv(os.path.join(BASE_DIR, '.env'))

DB = dict(
    host=os.getenv('DB_HOST', 'localhost'),
    user=os.getenv('DB_USER', 'root'),
    password=os.getenv('DB_PASS', ''),
    database=os.getenv('DB_NAME', 'sgc'),
    charset='utf8mb4',
)

SITUACAO     = next((a.split('=', 1)[1] for a in sys.argv if a.startswith('--situacao=')), None)
COM_FISCAIS  = '--com-fiscais' in sys.argv

# ── Estilo Visual ──────────────────────────────────────────────────────────────
COR_CONTRATO = '2E5090'   # azul — dados do contrato
COR_FISCAL   = 'C55A11'   # laranja — dados do fiscal
COR_RESUMO   = '375623'   # verde-escuro — aba de resumo

H_FONT    = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
D_FONT    = Font(name='Calibri', size=10)
BOLD_FONT = Font(name='Calibri', bold=True, size=10)

H_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
L_ALIGN = Alignment(vertical='top', wrap_text=True)
C_ALIGN = Alignment(horizontal='center', vertical='top')

THIN_BORDER = Border(
    left=Side(style='thin', color='D0D0D0'),
    right=Side(style='thin', color='D0D0D0'),
    top=Side(style='thin', color='D0D0D0'),
    bottom=Side(style='thin', color='D0D0D0'),
)

FILL_EVEN = PatternFill(start_color='F5F7FA', end_color='F5F7FA', fill_type='solid')
FILL_ODD  = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')


def make_fill(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')


# ── Colunas: (grupo, campo, cor, largura, chave) ───────────────────────────────
COLUNAS = [
    ('CONTRATO', 'Código',              COR_CONTRATO, 16, 'codigo'),
    ('CONTRATO', 'Número Original',     COR_CONTRATO, 20, 'numeroOriginal'),
    ('CONTRATO', 'Situação',            COR_CONTRATO, 14, 'situacao'),
    ('CONTRATO', 'Processo SEI',        COR_CONTRATO, 24, 'numProcesso'),
    ('CONTRATO', 'Vigência Início',     COR_CONTRATO, 14, 'vig_inicio'),
    ('CONTRATO', 'Vigência Fim',        COR_CONTRATO, 14, 'vig_final'),
    ('CONTRATADO', 'Código',            COR_CONTRATO, 18, 'codigoContratado'),
    ('CONTRATADO', 'Nome',              COR_CONTRATO, 42, 'nomeContratado'),
    ('CONTRATADO', 'Tipo',              COR_CONTRATO, 12, 'tipoContratado'),
    ('FISCAL',  'Tipo',                 COR_FISCAL,   16, 'fiscal_tipo'),
    ('FISCAL',  'Nome',                 COR_FISCAL,   36, 'fiscal_nome'),
    ('FISCAL',  'CPF',                  COR_FISCAL,   16, 'fiscal_cpf'),
    ('FISCAL',  'Telefone',             COR_FISCAL,   18, 'fiscal_telefone'),
    ('FISCAL',  'E-mail',               COR_FISCAL,   36, 'fiscal_email'),
    ('FISCAL',  'Registro Profissional', COR_FISCAL,  22, 'fiscal_registro'),
    ('FISCAL',  'Atualizado em',        COR_FISCAL,   18, 'fiscal_atualizacao'),
]


# ── Queries ────────────────────────────────────────────────────────────────────

SQL_CONTRATOS_E_FISCAIS = """
    SELECT
        c.codigo,
        c.numeroOriginal,
        c.situacao,
        c.numProcesso,
        c.dataInicioVigencia,
        c.dataFimVigencia,
        c.codigoContratado,
        c.nomeContratado,
        c.tipoContratado,
        f.id              AS fiscal_id,
        f.tipo            AS fiscal_tipo,
        f.nome            AS fiscal_nome,
        f.cpf             AS fiscal_cpf,
        f.telefone        AS fiscal_telefone,
        f.email           AS fiscal_email,
        f.registroProfissional AS fiscal_registro,
        f.data_atualizacao AS fiscal_atualizacao
    FROM contratos c
    LEFT JOIN fiscais_contrato f ON f.codigo_contrato = c.codigo
    {where}
    ORDER BY c.codigo, f.tipo, f.nome
"""


# ── Helpers ────────────────────────────────────────────────────────────────────

def fmt_data(d):
    if d is None:
        return ''
    if hasattr(d, 'strftime'):
        return d.strftime('%d/%m/%Y')
    return str(d)


def fmt_datetime(d):
    if d is None:
        return ''
    if hasattr(d, 'strftime'):
        return d.strftime('%d/%m/%Y %H:%M')
    return str(d)


def fmt_cpf(cpf):
    if not cpf:
        return ''
    s = ''.join(ch for ch in str(cpf) if ch.isdigit())
    if len(s) != 11:
        return str(cpf)
    return f'{s[0:3]}.{s[3:6]}.{s[6:9]}-{s[9:11]}'


# ── Carregar dados ─────────────────────────────────────────────────────────────

def carregar_dados():
    print('[1/3] Conectando ao banco...')
    conn = pymysql.connect(**DB)
    cur = conn.cursor(pymysql.cursors.DictCursor)

    where_parts = []
    if SITUACAO:
        where_parts.append(f"c.situacao = '{SITUACAO}'")
    if COM_FISCAIS:
        where_parts.append("f.id IS NOT NULL")
    where = ('WHERE ' + ' AND '.join(where_parts)) if where_parts else ''

    print('[2/3] Carregando contratos e fiscais...')
    cur.execute(SQL_CONTRATOS_E_FISCAIS.format(where=where))
    linhas = cur.fetchall()
    conn.close()

    return linhas


# ── Montar linha ──────────────────────────────────────────────────────────────

def montar_linha(row):
    return {
        'codigo':           str(row.get('codigo') or ''),
        'numeroOriginal':   str(row.get('numeroOriginal') or ''),
        'situacao':         str(row.get('situacao') or ''),
        'numProcesso':      str(row.get('numProcesso') or ''),
        'vig_inicio':       fmt_data(row.get('dataInicioVigencia')),
        'vig_final':        fmt_data(row.get('dataFimVigencia')),
        'codigoContratado': str(row.get('codigoContratado') or ''),
        'nomeContratado':   str(row.get('nomeContratado') or ''),
        'tipoContratado':   str(row.get('tipoContratado') or ''),
        'fiscal_tipo':         str(row.get('fiscal_tipo') or ''),
        'fiscal_nome':         str(row.get('fiscal_nome') or ''),
        'fiscal_cpf':          fmt_cpf(row.get('fiscal_cpf')),
        'fiscal_telefone':     str(row.get('fiscal_telefone') or ''),
        'fiscal_email':        str(row.get('fiscal_email') or ''),
        'fiscal_registro':     str(row.get('fiscal_registro') or ''),
        'fiscal_atualizacao':  fmt_datetime(row.get('fiscal_atualizacao')),
        '_tem_fiscal':         row.get('fiscal_id') is not None,
    }


# ── Aba principal ──────────────────────────────────────────────────────────────

def criar_aba_fiscais(wb, linhas):
    ws = wb.create_sheet('Fiscais por Contrato')
    ws.sheet_properties.tabColor = COR_FISCAL

    # Cabeçalho (linhas 1 e 2)
    for col_idx, (grupo, campo, cor, _, _chave) in enumerate(COLUNAS, start=1):
        for row_idx, valor in ((1, grupo), (2, campo)):
            cell = ws.cell(row=row_idx, column=col_idx, value=valor)
            cell.font      = H_FONT
            cell.fill      = make_fill(cor)
            cell.alignment = H_ALIGN
            cell.border    = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = COLUNAS[col_idx - 1][3]

    # Mesclar células de cabeçalho do grupo (linha 1) quando colunas adjacentes têm o mesmo grupo
    inicio = 1
    for col_idx in range(2, len(COLUNAS) + 2):
        grupo_atual = COLUNAS[col_idx - 2][0] if col_idx - 2 < len(COLUNAS) else None
        grupo_anterior = COLUNAS[inicio - 1][0]
        if grupo_atual != grupo_anterior:
            if col_idx - 1 > inicio:
                ws.merge_cells(start_row=1, start_column=inicio, end_row=1, end_column=col_idx - 1)
            inicio = col_idx

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 22
    ws.freeze_panes = 'A3'

    # Dados
    for data_row_idx, linha in enumerate(linhas, start=3):
        fill_bg = FILL_EVEN if data_row_idx % 2 == 0 else FILL_ODD

        for col_idx, (grupo, campo, cor, _, chave) in enumerate(COLUNAS, start=1):
            valor = linha.get(chave, '')
            cell = ws.cell(row=data_row_idx, column=col_idx, value=valor)
            cell.font      = D_FONT
            cell.border    = THIN_BORDER
            cell.alignment = L_ALIGN
            cell.fill      = fill_bg

        ws.row_dimensions[data_row_idx].height = 15

    # Auto-filtro
    ws.auto_filter.ref = f'A2:{get_column_letter(len(COLUNAS))}{len(linhas) + 2}'

    return ws


# ── Aba de resumo ──────────────────────────────────────────────────────────────

def criar_aba_resumo(wb, linhas):
    """Resumo por contrato: 1 linha por contrato com contagem de fiscais por tipo."""
    ws = wb.create_sheet('Resumo por Contrato')
    ws.sheet_properties.tabColor = COR_RESUMO

    # Agrupa por contrato
    contratos = {}
    for l in linhas:
        cod = l['codigo']
        if cod not in contratos:
            contratos[cod] = {
                'codigo':           cod,
                'numeroOriginal':   l['numeroOriginal'],
                'situacao':         l['situacao'],
                'codigoContratado': l['codigoContratado'],
                'nomeContratado':   l['nomeContratado'],
                'tipos':            {},
                'total':            0,
            }
        if l['_tem_fiscal']:
            tipo = l['fiscal_tipo'] or '(sem tipo)'
            contratos[cod]['tipos'][tipo] = contratos[cod]['tipos'].get(tipo, 0) + 1
            contratos[cod]['total'] += 1

    headers = [
        ('Código',           16),
        ('Número Original',  20),
        ('Situação',         14),
        ('Cód. Contratado',  18),
        ('Nome Contratado',  44),
        ('Total Fiscais',    14),
        ('Detalhamento por Tipo', 60),
    ]
    for col_idx, (titulo, larg) in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=titulo)
        cell.font      = H_FONT
        cell.fill      = make_fill(COR_RESUMO)
        cell.alignment = H_ALIGN
        cell.border    = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = larg

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = 'A2'

    for idx, (_, c) in enumerate(sorted(contratos.items()), start=2):
        fill_bg = FILL_EVEN if idx % 2 == 0 else FILL_ODD
        detalhe = ' | '.join(f'{t}: {n}' for t, n in sorted(c['tipos'].items())) or '—'
        values = [
            c['codigo'], c['numeroOriginal'], c['situacao'],
            c['codigoContratado'], c['nomeContratado'],
            c['total'], detalhe,
        ]
        for col_idx, v in enumerate(values, start=1):
            cell = ws.cell(row=idx, column=col_idx, value=v)
            cell.font      = D_FONT
            cell.border    = THIN_BORDER
            cell.alignment = C_ALIGN if col_idx == 6 else L_ALIGN
            cell.fill      = fill_bg

    ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}{len(contratos) + 1}'
    return ws


# ── Ponto de entrada ───────────────────────────────────────────────────────────

def main():
    print('=' * 60)
    print('Relatório de Fiscais de Contratos — SGC')
    print('=' * 60)
    if SITUACAO:
        print(f'  Filtro: situacao = {SITUACAO}')
    if COM_FISCAIS:
        print('  Filtro: apenas contratos com pelo menos 1 fiscal')
    print()

    rows = carregar_dados()
    print(f'     {len(rows)} linhas brutas (contrato x fiscal).')

    print('[3/3] Montando linhas...')
    linhas = [montar_linha(r) for r in rows]

    if not linhas:
        print('\nNenhum dado encontrado com os filtros aplicados.')
        return

    wb = Workbook()
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    criar_aba_resumo(wb, linhas)
    criar_aba_fiscais(wb, linhas)

    nome_arquivo = f'relatorio_fiscais_contratos_{date.today().strftime("%Y%m%d")}.xlsx'
    caminho = os.path.join(RELATORIOS, nome_arquivo)
    wb.save(caminho)

    contratos_unicos = len({l['codigo'] for l in linhas})
    com_fiscal       = sum(1 for l in linhas if l['_tem_fiscal'])
    sem_fiscal       = contratos_unicos - len({l['codigo'] for l in linhas if l['_tem_fiscal']})
    print()
    print(f'Arquivo gerado: {caminho}')
    print(f'Total: {contratos_unicos} contratos | {com_fiscal} registros de fiscal | {sem_fiscal} contratos sem fiscal')


if __name__ == '__main__':
    main()
