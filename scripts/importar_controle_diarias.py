"""
Script de importação: Planilha histórica de controle de diárias.

Importa dados de 'Planilha sem título.xlsx' para as tabelas:
  - diarias_controle_viagens
  - diarias_controle_servidores
  - diarias_controle_prestacao

A planilha tem 1 linha por SERVIDOR por viagem, agrupadas por PROCESSO.
O script agrupa por processo → cria 1 viagem → N servidores.

Coluna TRECHO é parseada em origem + destino (removida do modelo final).
Coluna SUPERINTENDÊNCIA é mapeada para setor_id via sigla.

Uso:
  python scripts/importar_controle_diarias.py               # DRY-RUN
  python scripts/importar_controle_diarias.py --executar     # APLICA
"""
import os
import sys
import re
from datetime import datetime, date
from collections import defaultdict
import pymysql
import openpyxl
from dotenv import load_dotenv

load_dotenv(os.path.join(os.path.dirname(__file__), '..', '.env'))

DB_HOST = os.getenv('DB_HOST', 'localhost')
DB_USER = os.getenv('DB_USER', 'root')
DB_PASS = os.getenv('DB_PASS', 'root')
DB_NAME = os.getenv('DB_NAME', 'sgc')

XLSX_PATH = r'C:\Users\guilh\Downloads\Planilha sem título.xlsx'

DRY_RUN = '--executar' not in sys.argv


def get_connection():
    return pymysql.connect(
        host=DB_HOST, user=DB_USER, password=DB_PASS, database=DB_NAME,
        charset='utf8mb4', cursorclass=pymysql.cursors.DictCursor,
    )


def parse_trecho(trecho):
    """
    Extrai origem e destino de um trecho.
    Ex: 'TERESINA / MILTON BRANDÃO / TERESINA' → ('TERESINA', 'MILTON BRANDÃO')
    Ex: 'TERESINA-PI/SANTA FILOMENA-PI/TERESINA-PI' → ('TERESINA-PI', 'SANTA FILOMENA-PI')
    """
    if not trecho:
        return None, None
    # Normaliza separadores
    parts = re.split(r'\s*/\s*', trecho.strip())
    parts = [p.strip() for p in parts if p.strip()]
    if len(parts) >= 2:
        return parts[0], parts[1]
    elif len(parts) == 1:
        return parts[0], None
    return None, None


def parse_status_viagem(status_str):
    """1=Realizada, 2=Cancelada."""
    if not status_str:
        return 1
    s = str(status_str).upper().strip()
    if 'CANCELAD' in s:
        return 2
    return 1


def parse_prestacao(prestacao_str):
    """Retorna (status, ano_referencia).
    'ENTREGUE' → (1, None)
    'PENDENTE 2025' → (2, 2025)
    """
    if not prestacao_str:
        return None, None
    s = str(prestacao_str).upper().strip()
    if s.startswith('ENTREGUE'):
        return 1, None
    if s.startswith('PENDENTE'):
        match = re.search(r'(\d{4})', s)
        ano = int(match.group(1)) if match else None
        return 2, ano
    return None, None


def parse_relatorio(relatorio_str):
    """1=Aprovado, 2=Reprovado, 3=Pendente."""
    if not relatorio_str:
        return None
    s = str(relatorio_str).upper().strip()
    if 'APROVAD' in s:
        return 1
    if 'REPROVAD' in s:
        return 2
    if 'PENDENT' in s:
        return 3
    return None


def safe_date(val):
    """Converte datetime/date/string para date."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    try:
        return datetime.strptime(str(val).strip()[:10], '%Y-%m-%d').date()
    except Exception:
        return None


def safe_float(val):
    """Converte para float, retorna None se impossível."""
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def safe_str(val, max_len=None):
    """Converte para string stripped, trunca se necessário."""
    if val is None:
        return None
    s = str(val).strip()
    if max_len and len(s) > max_len:
        s = s[:max_len]
    return s or None


def run():
    conn = get_connection()
    cursor = conn.cursor()

    modo = "DRY-RUN (use --executar para aplicar)" if DRY_RUN else "EXECUTANDO"
    print("=" * 70)
    print(f"Importação planilha histórica de diárias — {modo}")
    print("=" * 70)

    # 1. Carregar mapa de siglas → setor_id (apenas SEAD, idorgao=1)
    cursor.execute("SELECT identidade, sigla FROM setor WHERE idorgao = 1 AND sigla IS NOT NULL")
    sigla_map = {r['sigla'].upper(): r['identidade'] for r in cursor.fetchall()}
    print(f"\n[1] Mapa de siglas SEAD: {len(sigla_map)} setores")

    # 2. Verificar processos já importados
    cursor.execute("SELECT processo FROM diarias_controle_viagens")
    existing = set(r['processo'] for r in cursor.fetchall())
    print(f"    Processos já no banco: {len(existing)}")

    # 3. Ler Excel
    print(f"\n[2] Lendo planilha: {XLSX_PATH}")
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb[wb.sheetnames[0]]

    # Agrupar por processo (para ao encontrar linha sem processo E sem nome)
    processos = defaultdict(list)
    skipped_no_processo = 0
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True), start=2):
        processo = safe_str(row[0], 50)
        nome = safe_str(row[1])
        if not processo and not nome:
            break  # Fim dos dados reais
        if not processo:
            skipped_no_processo += 1
            continue
        processos[processo].append(row)

    print(f"    Linhas lidas: {ws.max_row - 1}")
    print(f"    Processos distintos: {len(processos)}")
    print(f"    Linhas sem processo: {skipped_no_processo}")

    # 4. Filtrar já importados
    novos = {p: rows for p, rows in processos.items() if p not in existing}
    print(f"    Novos a importar: {len(novos)} processos")

    if not novos:
        print("\n    Nada a importar.")
        conn.close()
        return

    # 5. Importar
    print(f"\n[3] Importando...")
    total_viagens = 0
    total_servidores = 0
    total_prestacoes = 0

    for processo, rows in novos.items():
        first = rows[0]

        # Dados da viagem (do primeiro registro do grupo)
        trecho = safe_str(first[3])
        origem, destino = parse_trecho(trecho)
        data_inicio = safe_date(first[4])
        data_termino = safe_date(first[5])
        status_viagem = parse_status_viagem(first[13])
        observacao = safe_str(first[19])

        # Setor: primeiro tenta SUPERINTENDÊNCIA (col 20), depois SETOR (col 21)
        superintendencia = safe_str(first[20])
        setor_txt = safe_str(first[21])
        setor_id = None
        if superintendencia:
            setor_id = sigla_map.get(superintendencia.upper())
        if not setor_id and setor_txt:
            setor_id = sigla_map.get(setor_txt.upper())

        if not data_inicio or not data_termino:
            continue

        if DRY_RUN:
            total_viagens += 1
            total_servidores += len(rows)
            for r in rows:
                prest_status, _ = parse_prestacao(r[14])
                if prest_status is not None:
                    total_prestacoes += 1
            continue

        # INSERT viagem
        cursor.execute("""
            INSERT INTO diarias_controle_viagens
            (processo, setor_id, origem, destino, data_inicio, data_termino, status_viagem, observacao)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
        """, (processo, setor_id, origem, destino, data_inicio, data_termino, status_viagem, observacao))
        viagem_id = cursor.lastrowid
        total_viagens += 1

        # INSERT servidores
        for r in rows:
            cpf = safe_str(r[2], 14)
            if not cpf:
                continue

            nome = safe_str(r[1], 255)
            vinculo = safe_str(r[6], 50)
            qtd = safe_float(r[10]) or 0
            valor_unit = safe_float(r[11])
            valor_total = safe_float(r[12]) if not isinstance(r[12], str) else None
            natureza = safe_str(r[7], 10)
            sub_item = safe_str(r[8], 10)
            fonte = safe_str(r[9], 20)
            baixa_np = safe_str(r[17], 50)
            scdp = safe_str(r[18], 20)

            # Calcula valor_total se não disponível (fórmula #NAME?)
            if valor_total is None and qtd and valor_unit:
                valor_total = qtd * valor_unit

            cursor.execute("""
                INSERT INTO diarias_controle_servidores
                (viagem_id, cpf, nome, vinculo, qtd_diarias, valor_unitario, valor_total,
                 natureza_despesa, sub_item, fonte_recursos, baixa_np, sistema_scdp)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (viagem_id, cpf, nome, vinculo, qtd, valor_unit, valor_total,
                  natureza, sub_item, fonte, baixa_np, scdp))
            servidor_id = cursor.lastrowid
            total_servidores += 1

            # INSERT prestação (se houver dados)
            prest_status, prest_ano = parse_prestacao(r[14])
            data_entrega = safe_date(r[15])
            relatorio = parse_relatorio(r[16])

            if prest_status is not None:
                cursor.execute("""
                    INSERT INTO diarias_controle_prestacao
                    (servidor_id, status, data_entrega, relatorio, ano_referencia)
                    VALUES (%s, %s, %s, %s, %s)
                """, (servidor_id, prest_status, data_entrega, relatorio, prest_ano))
                total_prestacoes += 1

    if not DRY_RUN:
        conn.commit()

    print(f"\n    Viagens inseridas:    {total_viagens}")
    print(f"    Servidores inseridos: {total_servidores}")
    print(f"    Prestações inseridas: {total_prestacoes}")

    # Verificação final
    if not DRY_RUN:
        cursor.execute("SELECT COUNT(*) as cnt FROM diarias_controle_viagens")
        print(f"\n    Total viagens no banco: {cursor.fetchone()['cnt']}")
        cursor.execute("SELECT COUNT(*) as cnt FROM diarias_controle_servidores")
        print(f"    Total servidores no banco: {cursor.fetchone()['cnt']}")
        cursor.execute("SELECT COUNT(*) as cnt FROM diarias_controle_prestacao")
        print(f"    Total prestações no banco: {cursor.fetchone()['cnt']}")

    cursor.close()
    conn.close()

    print("\n" + "=" * 70)
    print("Concluído!" + (" (DRY-RUN)" if DRY_RUN else ""))
    print("=" * 70)


if __name__ == '__main__':
    run()
