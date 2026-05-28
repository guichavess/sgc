"""
Relatório de Tipificação de Contratos
======================================
Exporta todos os contratos do módulo "Execução de Contratos" com:
  - Dados do cabeçalho (código, contratado, objeto, processo, vigência,
    natureza, fonte, tipo de despesa, sub item)
  - Colunas da aba "Tipo-Contrato" para preenchimento pelo funcionário:
      Categorização (CATSERV / CATMAT) e Gestão (Centro de Custo, Tipo de Execução)
  - Status atual de cada campo

O Excel preenchido pode ser reimportado com: scripts/importar_tipificacao_excel.py

Uso:
  python scripts/relatorio_tipificacao_contratos.py
  python scripts/relatorio_tipificacao_contratos.py --situacao=VIGENTE
  python scripts/relatorio_tipificacao_contratos.py --apenas-pendentes
"""
import os
import sys
from datetime import date
import pymysql
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv

# ── Configuração ───────────────────────────────────────────────────────────────
BASE_DIR    = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
RELATORIOS  = r'C:\Users\guilh\OneDrive\Documentos\SEAD\Projetos\Nova pasta\Relatórios'
os.makedirs(RELATORIOS, exist_ok=True)
load_dotenv(os.path.join(BASE_DIR, '.env'))

DB = dict(
    host=os.getenv('DB_HOST', 'localhost'),
    user=os.getenv('DB_USER', 'root'),
    password=os.getenv('DB_PASS', ''),
    database=os.getenv('DB_NAME', 'sgc'),
    charset='utf8mb4',
)

# Flags CLI
SITUACAO       = next((a.split('=', 1)[1] for a in sys.argv if a.startswith('--situacao=')), None)
APENAS_PENDENTES = '--apenas-pendentes' in sys.argv

# ── Estilo Visual ──────────────────────────────────────────────────────────────
COR_DADOS   = '2E5090'   # azul — dados informativos (somente leitura)
COR_CATEG   = 'C55A11'   # laranja — categorização CATSERV/CATMAT (preencher)
COR_GESTAO  = '375623'   # verde-escuro — centro de custo e tipo exec (preencher)
COR_STATUS  = '595959'   # cinza — status calculado (somente leitura)
COR_INSTR   = 'FFC000'   # amarelo — aba de instruções

H_FONT    = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
D_FONT    = Font(name='Calibri', size=10)
BOLD_FONT = Font(name='Calibri', bold=True, size=10)

H_ALIGN   = Alignment(horizontal='center', vertical='center', wrap_text=True)
L_ALIGN   = Alignment(vertical='top', wrap_text=True)
C_ALIGN   = Alignment(horizontal='center', vertical='top')

THIN_BORDER = Border(
    left=Side(style='thin', color='D0D0D0'),
    right=Side(style='thin', color='D0D0D0'),
    top=Side(style='thin', color='D0D0D0'),
    bottom=Side(style='thin', color='D0D0D0'),
)

# Faixas de cores para linhas alternadas (leve)
FILL_EVEN = PatternFill(start_color='F5F7FA', end_color='F5F7FA', fill_type='solid')
FILL_ODD  = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')


def make_fill(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')


# ── Definição das colunas ──────────────────────────────────────────────────────
#  (grupo, campo, cor_header, largura, chave_no_dict_da_linha)
COLUNAS = [
    # Grupo A — Dados do Contrato
    ('DADOS DO CONTRATO', 'Código SIAFE',    COR_DADOS,  18, 'codigo'),
    ('DADOS DO CONTRATO', 'Contratado',      COR_DADOS,  48, 'contratado'),
    ('DADOS DO CONTRATO', 'Objeto',          COR_DADOS,  60, 'objeto'),
    ('DADOS DO CONTRATO', 'Processo SEI',    COR_DADOS,  26, 'num_processo'),
    ('DADOS DO CONTRATO', 'Vigência Início', COR_DADOS,  16, 'vig_inicio'),
    ('DADOS DO CONTRATO', 'Vigência Final',  COR_DADOS,  16, 'vig_final'),
    ('DADOS DO CONTRATO', 'Natureza',        COR_DADOS,  40, 'natureza'),
    ('DADOS DO CONTRATO', 'Fonte(s)',        COR_DADOS,  18, 'fontes'),
    ('DADOS DO CONTRATO', 'Tipo de Despesa', COR_DADOS,  42, 'tipo_despesa'),
    ('DADOS DO CONTRATO', 'Sub Item',        COR_DADOS,  42, 'sub_item'),
    # Grupo B — Tipificação: Categorização (CATSERV)
    ('TIPIFICAÇÃO — CATSERV', 'Tipo de Contrato (S/M/SM)',  COR_CATEG, 14, 'tipo_contrato'),
    ('TIPIFICAÇÃO — CATSERV', 'Seção',                      COR_CATEG, 32, 'catserv_secao'),
    ('TIPIFICAÇÃO — CATSERV', 'Divisão',                    COR_CATEG, 32, 'catserv_divisao'),
    ('TIPIFICAÇÃO — CATSERV', 'Grupo',                      COR_CATEG, 32, 'catserv_grupo'),
    ('TIPIFICAÇÃO — CATSERV', 'Classe (ID) ⚑',             COR_CATEG, 20, 'catserv_classe_id'),
    # Grupo C — Tipificação: Categorização (CATMAT)
    ('TIPIFICAÇÃO — CATMAT',  'Grupo',                      COR_CATEG, 32, 'catmat_grupo'),
    ('TIPIFICAÇÃO — CATMAT',  'Classe',                     COR_CATEG, 32, 'catmat_classe'),
    ('TIPIFICAÇÃO — CATMAT',  'PDM (ID) ⚑',               COR_CATEG, 20, 'catmat_pdm_id'),
    # Grupo D — Gestão
    ('GESTÃO',               'Centro de Custo',             COR_GESTAO, 32, 'centro_custo'),
    ('GESTÃO',               'Tipo de Execução',            COR_GESTAO, 22, 'tipo_execucao'),
    # Grupo E — Status
    ('STATUS',               'Tipificação',                 COR_STATUS, 20, 'status_tipificacao'),
    ('STATUS',               'Centro de Custo',             COR_STATUS, 18, 'status_cc'),
    ('STATUS',               'Tipo de Execução',            COR_STATUS, 20, 'status_te'),
]


# ── Queries ────────────────────────────────────────────────────────────────────

SQL_CONTRATOS = """
    SELECT
        c.codigo,
        CONCAT(COALESCE(c.codigoContratado, ''), ' - ', COALESCE(c.nomeContratado, '')) AS contratado,
        c.objeto,
        c.numProcesso,
        c.dataInicioVigencia,
        c.dataFimVigencia,
        c.tipo_contrato,
        c.catserv_classe_id,
        c.catserv_grupo_id,
        c.catmat_classe_id,
        c.catmat_pdm_id,
        nd.codigo   AS nat_codigo,
        nd.titulo   AS nat_titulo,
        cc.descricao AS centro_custo,
        te.descricao AS tipo_execucao
    FROM contratos c
    LEFT JOIN natdespesas  nd ON c.natureza_id       = nd.id
    LEFT JOIN centrodecusto cc ON c.centro_de_custo_id = cc.id
    LEFT JOIN tipoexecucao  te ON c.tipo_execucao_id   = te.id
    {where}
    ORDER BY c.codigo
"""

SQL_ADITIVOS = """
    SELECT codigo_contrato, MAX(dtVigenciaFim) AS vigencia_aditivo
    FROM contratos_aditivo
    WHERE dtVigenciaFim IS NOT NULL
    GROUP BY codigo_contrato
"""

SQL_CATSERV_CLASSES = """
    SELECT
        cl.codigo_classe,
        cl.nome         AS classe_nome,
        g.codigo_grupo,
        g.nome          AS grupo_nome,
        d.codigo_divisao,
        d.nome          AS divisao_nome,
        s.codigo_secao,
        s.nome          AS secao_nome
    FROM catserv_classes  cl
    JOIN catserv_grupos   g  ON cl.codigo_grupo   = g.codigo_grupo
    JOIN catserv_divisoes d  ON g.codigo_divisao  = d.codigo_divisao
    JOIN catserv_secoes   s  ON d.codigo_secao    = s.codigo_secao
"""

SQL_CATSERV_GRUPOS = """
    SELECT
        g.codigo_grupo,
        g.nome          AS grupo_nome,
        d.codigo_divisao,
        d.nome          AS divisao_nome,
        s.codigo_secao,
        s.nome          AS secao_nome
    FROM catserv_grupos   g
    JOIN catserv_divisoes d ON g.codigo_divisao = d.codigo_divisao
    JOIN catserv_secoes   s ON d.codigo_secao   = s.codigo_secao
"""

SQL_CATMAT = """
    SELECT
        p.id            AS pdm_id,
        p.codigo        AS pdm_codigo,
        p.nome          AS pdm_nome,
        cl.id           AS classe_id,
        cl.codigo       AS classe_codigo,
        cl.nome         AS classe_nome,
        g.id            AS grupo_id,
        g.codigo        AS grupo_codigo,
        g.nome          AS grupo_nome
    FROM catmat_pdms   p
    JOIN catmat_classes cl ON p.codigo_classe  = cl.codigo
    JOIN catmat_grupos  g  ON cl.codigo_grupo  = g.codigo
"""

SQL_CLASSIFICADORES = """
    SELECT
        ei.CodContrato,
        GROUP_CONCAT(DISTINCT ei.Fonte           SEPARATOR ' | ') AS fontes,
        GROUP_CONCAT(DISTINCT ei.TipoPatrimonial SEPARATOR ' | ') AS tipos_patrimoniais,
        GROUP_CONCAT(DISTINCT ei.SubItemDespesa  SEPARATOR ' | ') AS subitens
    FROM empenho_itens ei
    WHERE ei.Natureza NOT IN ('339092', '449092')
    GROUP BY ei.CodContrato
"""

# Fallback: mesmos classificadores mas INCLUINDO estornos (339092, 449092)
# Para contratos que SÓ possuem empenhos de estorno
SQL_CLASSIFICADORES_FALLBACK = """
    SELECT
        ei.CodContrato,
        GROUP_CONCAT(DISTINCT ei.Fonte           SEPARATOR ' | ') AS fontes,
        GROUP_CONCAT(DISTINCT ei.TipoPatrimonial SEPARATOR ' | ') AS tipos_patrimoniais,
        GROUP_CONCAT(DISTINCT ei.SubItemDespesa  SEPARATOR ' | ') AS subitens
    FROM empenho_itens ei
    GROUP BY ei.CodContrato
"""

SQL_CLASS_SUBITEM = """
    SELECT
        CONCAT(
            COALESCE(valoresClassificador1, ''), '.',
            COALESCE(valoresClassificador2, '')
        ) AS codigo_completo,
        nomeClassificador
    FROM class_subitemdespesa
"""

SQL_CLASS_TIPOPATRIMONIAL = """
    SELECT
        CAST(valoresClassificador1 AS CHAR) AS codigo,
        nomeClassificador
    FROM class_tipopatrimonial
"""

SQL_NATDESPESAS_FULL = """
    SELECT codigo, titulo FROM natdespesas
"""

SQL_NATUREZA_FALLBACK = """
    SELECT ei.CodContrato,
           MIN(CASE WHEN ei.Natureza NOT IN ('339092','449092') THEN ei.Natureza END) AS nat_principal,
           MIN(ei.Natureza) AS nat_qualquer
    FROM empenho_itens ei
    GROUP BY ei.CodContrato
"""


# ── Helpers ────────────────────────────────────────────────────────────────────

def fmt_data(d):
    """Formata date/datetime para dd/mm/yyyy ou retorna ''."""
    if d is None:
        return ''
    if hasattr(d, 'strftime'):
        return d.strftime('%d/%m/%Y')
    return str(d)


def truncar(texto, maxlen=120):
    """Trunca string longa."""
    if not texto:
        return ''
    return texto[:maxlen] + ('…' if len(texto) > maxlen else '')


def resolver_nome_subitem(codigo_raw, mapa_subitem):
    """Dado '2736.51', retorna '2736.51 - SERVICOS TECNICOS PROFISSIONAIS' ou só o código."""
    if not codigo_raw:
        return ''
    partes = []
    for c in str(codigo_raw).split(' | '):
        c = c.strip()
        nome = mapa_subitem.get(c, '')
        partes.append(f'{c} - {nome}' if nome else c)
    return ' | '.join(partes)


def resolver_nome_tp(codigo_raw, mapa_tp):
    """Dado '43', retorna '43 - Serviços de Terceiros - PJ' ou só o código."""
    if not codigo_raw:
        return ''
    partes = []
    for c in str(codigo_raw).split(' | '):
        c = c.strip()
        nome = mapa_tp.get(c, '')
        partes.append(f'{c} - {nome}' if nome else c)
    return ' | '.join(partes)


def esta_tipificado(row):
    """Replica lógica de Contrato.esta_tipificado (Python, sem ORM)."""
    tipo = row.get('tipo_contrato')
    catserv_ok = bool(row.get('catserv_classe_id') or row.get('catserv_grupo_id'))
    catmat_ok  = bool(row.get('catmat_classe_id') and row.get('catmat_pdm_id'))
    if tipo == 'S':
        return catserv_ok
    if tipo == 'M':
        return catmat_ok
    if tipo == 'SM':
        return catserv_ok and catmat_ok
    return False


# ── Carregar dados do banco ────────────────────────────────────────────────────

def carregar_dados():
    print('[1/6] Conectando ao banco...')
    conn = pymysql.connect(**DB)
    cur = conn.cursor(pymysql.cursors.DictCursor)

    # Contratos
    print('[2/6] Carregando contratos...')
    where_clause = ''
    if SITUACAO:
        where_clause = f"WHERE c.situacao = '{SITUACAO}'"
    cur.execute(SQL_CONTRATOS.format(where=where_clause))
    contratos = cur.fetchall()
    print(f'     {len(contratos)} contratos encontrados.')

    # Aditivos → vigência final efetiva
    print('[3/6] Carregando aditivos...')
    cur.execute(SQL_ADITIVOS)
    mapa_aditivos = {r['codigo_contrato']: r['vigencia_aditivo'] for r in cur.fetchall()}

    # Hierarquias CATSERV
    print('[4/6] Carregando hierarquias CATSERV...')
    cur.execute(SQL_CATSERV_CLASSES)
    mapa_classe = {r['codigo_classe']: r for r in cur.fetchall()}

    cur.execute(SQL_CATSERV_GRUPOS)
    mapa_grupo_catserv = {r['codigo_grupo']: r for r in cur.fetchall()}

    # Hierarquias CATMAT
    cur.execute(SQL_CATMAT)
    mapa_catmat_pdm = {r['pdm_id']: r for r in cur.fetchall()}

    # Classificadores (Fonte, SubItem, TipoPatrimonial) por contrato
    print('[5/6] Carregando classificadores...')
    cur.execute(SQL_CLASSIFICADORES)
    mapa_classif = {}
    for r in cur.fetchall():
        try:
            cod = int(r['CodContrato'])
        except (TypeError, ValueError):
            continue
        mapa_classif[cod] = r

    # Fallback: classificadores incluindo estornos (para contratos que SÓ têm estornos)
    cur.execute(SQL_CLASSIFICADORES_FALLBACK)
    cnt_fallback = 0
    for r in cur.fetchall():
        try:
            cod = int(r['CodContrato'])
        except (TypeError, ValueError):
            continue
        if cod not in mapa_classif:
            mapa_classif[cod] = r
            cnt_fallback += 1
    if cnt_fallback:
        print(f'       +{cnt_fallback} contratos via fallback (apenas estornos).')

    # Nomes de SubItem e TipoPatrimonial
    cur.execute(SQL_CLASS_SUBITEM)
    mapa_subitem = {r['codigo_completo']: r['nomeClassificador'] for r in cur.fetchall()}

    cur.execute(SQL_CLASS_TIPOPATRIMONIAL)
    mapa_tp = {r['codigo']: r['nomeClassificador'] for r in cur.fetchall()}

    # Mapa completo natdespesas: codigo → titulo  (para fallback)
    cur.execute(SQL_NATDESPESAS_FULL)
    mapa_natdespesas = {str(r['codigo']).strip(): r['titulo'] for r in cur.fetchall()}

    # Fallback de natureza via empenho_itens (para contratos sem natureza_id)
    print('       Carregando fallback de natureza via empenho_itens...')
    cur.execute(SQL_NATUREZA_FALLBACK)
    mapa_nat_fallback = {}
    for r in cur.fetchall():
        try:
            cod = int(r['CodContrato'])
        except (TypeError, ValueError):
            continue
        # Prefere natureza principal (excl. estornos); senão pega qualquer uma
        nat = r['nat_principal'] or r['nat_qualquer']
        if nat:
            mapa_nat_fallback[cod] = str(nat).strip()
    print(f'       {len(mapa_nat_fallback)} contratos com natureza via fallback.')

    conn.close()

    return dict(
        contratos=contratos,
        mapa_aditivos=mapa_aditivos,
        mapa_classe=mapa_classe,
        mapa_grupo_catserv=mapa_grupo_catserv,
        mapa_catmat_pdm=mapa_catmat_pdm,
        mapa_classif=mapa_classif,
        mapa_subitem=mapa_subitem,
        mapa_tp=mapa_tp,
        mapa_natdespesas=mapa_natdespesas,
        mapa_nat_fallback=mapa_nat_fallback,
    )


# ── Montar linha do Excel ──────────────────────────────────────────────────────

def montar_linha(row, dados):
    mapa_aditivos      = dados['mapa_aditivos']
    mapa_classe        = dados['mapa_classe']
    mapa_grupo_catserv = dados['mapa_grupo_catserv']
    mapa_catmat_pdm    = dados['mapa_catmat_pdm']
    mapa_classif       = dados['mapa_classif']
    mapa_subitem       = dados['mapa_subitem']
    mapa_tp            = dados['mapa_tp']
    mapa_natdespesas   = dados['mapa_natdespesas']
    mapa_nat_fallback  = dados['mapa_nat_fallback']

    codigo = row['codigo']

    # Vigência final: prefere aditivo
    vig_final_raw = mapa_aditivos.get(str(codigo)) or row.get('dataFimVigencia')
    vig_final = fmt_data(vig_final_raw)

    # Natureza — via natureza_id (JOIN), senão fallback via empenho_itens
    nat_cod   = row.get('nat_codigo') or ''
    nat_titulo = row.get('nat_titulo') or ''
    natureza  = f'{nat_cod} - {nat_titulo}' if nat_cod else ''

    if not natureza:
        # Fallback: buscar via empenho_itens → natdespesas
        try:
            cod_int = int(str(codigo).replace('.', '').replace('/', ''))
        except (ValueError, TypeError):
            cod_int = None
        if cod_int and cod_int in mapa_nat_fallback:
            fb_codigo = mapa_nat_fallback[cod_int]
            fb_titulo = mapa_natdespesas.get(fb_codigo, '')
            natureza = f'{fb_codigo} - {fb_titulo}' if fb_titulo else fb_codigo

    # Classificadores — CodContrato é BIGINT, tentar conversão
    classif = {}
    try:
        classif = mapa_classif.get(int(str(codigo).replace('.', '').replace('/', '')), {})
    except (ValueError, TypeError):
        pass

    fontes     = str(classif.get('fontes') or '').strip()
    tipo_desp  = resolver_nome_tp(classif.get('tipos_patrimoniais'), mapa_tp)
    sub_item   = resolver_nome_subitem(classif.get('subitens'), mapa_subitem)

    # Hierarquia CATSERV
    catserv_classe_id = row.get('catserv_classe_id')
    catserv_grupo_id  = row.get('catserv_grupo_id')
    catserv_secao = catserv_divisao = catserv_grupo_nome = ''

    if catserv_classe_id and catserv_classe_id in mapa_classe:
        h = mapa_classe[catserv_classe_id]
        catserv_secao   = f"{h['codigo_secao']} - {h['secao_nome']}"
        catserv_divisao = f"{h['codigo_divisao']} - {h['divisao_nome']}"
        catserv_grupo_nome = f"{h['codigo_grupo']} - {h['grupo_nome']}"
    elif catserv_grupo_id and catserv_grupo_id in mapa_grupo_catserv:
        h = mapa_grupo_catserv[catserv_grupo_id]
        catserv_secao   = f"{h['codigo_secao']} - {h['secao_nome']}"
        catserv_divisao = f"{h['codigo_divisao']} - {h['divisao_nome']}"
        catserv_grupo_nome = f"{h['codigo_grupo']} - {h['grupo_nome']}"

    # Hierarquia CATMAT
    catmat_pdm_id = row.get('catmat_pdm_id')
    catmat_grupo_nome = catmat_classe_nome = ''

    if catmat_pdm_id and catmat_pdm_id in mapa_catmat_pdm:
        h = mapa_catmat_pdm[catmat_pdm_id]
        catmat_grupo_nome  = f"{h['grupo_codigo']} - {h['grupo_nome']}"
        catmat_classe_nome = f"{h['classe_codigo']} - {h['classe_nome']}"

    # Status
    tipificado = esta_tipificado(row)
    status_tip = '✓ Tipificado' if tipificado else '⚠ Pendente'
    status_cc  = '✓ Definido'   if row.get('centro_custo') else '—'
    status_te  = '✓ Definido'   if row.get('tipo_execucao') else '—'

    return {
        'codigo':           str(codigo),
        'contratado':       str(row.get('contratado') or ''),
        'objeto':           truncar(row.get('objeto') or '', 120),
        'num_processo':     str(row.get('numProcesso') or ''),
        'vig_inicio':       fmt_data(row.get('dataInicioVigencia')),
        'vig_final':        vig_final,
        'natureza':         natureza,
        'fontes':           fontes,
        'tipo_despesa':     tipo_desp,
        'sub_item':         sub_item,
        'tipo_contrato':    str(row.get('tipo_contrato') or ''),
        'catserv_secao':    catserv_secao,
        'catserv_divisao':  catserv_divisao,
        'catserv_grupo':    catserv_grupo_nome,
        'catserv_classe_id': str(catserv_classe_id) if catserv_classe_id else '',
        'catmat_grupo':     catmat_grupo_nome,
        'catmat_classe':    catmat_classe_nome,
        'catmat_pdm_id':    str(catmat_pdm_id) if catmat_pdm_id else '',
        'centro_custo':     str(row.get('centro_custo') or ''),
        'tipo_execucao':    str(row.get('tipo_execucao') or ''),
        'status_tipificacao': status_tip,
        'status_cc':          status_cc,
        'status_te':          status_te,
        '_tipificado':        tipificado,
    }


# ── Construir aba de instruções ────────────────────────────────────────────────

def criar_aba_instrucoes(wb):
    ws = wb.create_sheet('INSTRUÇÕES', 0)
    ws.sheet_properties.tabColor = COR_INSTR
    ws.column_dimensions['A'].width = 100
    ws.row_dimensions[1].height = 30

    instrucoes = [
        ('INSTRUÇÕES DE PREENCHIMENTO — Relatório de Tipificação de Contratos', True, 14),
        ('', False, 11),
        ('COLUNAS AZUIS (DADOS DO CONTRATO): não altere — são apenas informativas.', False, 11),
        ('COLUNAS LARANJA (TIPIFICAÇÃO): preencha para contratos não tipificados.', False, 11),
        ('COLUNAS VERDE (GESTÃO): preencha Centro de Custo e Tipo de Execução conforme necessário.', False, 11),
        ('COLUNAS CINZAS (STATUS): calculadas automaticamente — não altere.', False, 11),
        ('', False, 11),
        ('CAMPOS IMPORTÁVEIS (marcados com ⚑):', True, 11),
        ('  • Col O — CATSERV Classe (ID): informe o código_classe da tabela catserv_classes.', False, 11),
        ('  • Col R — CATMAT PDM (ID): informe o id (chave primária) da tabela catmat_pdms.', False, 11),
        ('  • Col S — Centro de Custo: escreva a descrição exata de centrodecusto.descricao.', False, 11),
        ('  • Col T — Tipo de Execução: escreva a descrição exata de tipoexecucao.descricao.', False, 11),
        ('', False, 11),
        ('COMO REIMPORTAR APÓS PREENCHIMENTO:', True, 11),
        ('  1. Salve o arquivo (mantenha o nome original ou ajuste --arquivo=...).', False, 11),
        ('  2. Execute: python scripts/importar_tipificacao_excel.py --dry-run', False, 11),
        ('  3. Verifique o relatório de validação (erros serão listados).', False, 11),
        ('  4. Execute sem --dry-run para aplicar: python scripts/importar_tipificacao_excel.py', False, 11),
        ('', False, 11),
        ('TIPOS DE CONTRATO:', True, 11),
        ('  S  = Serviço (preencher apenas CATSERV)', False, 11),
        ('  M  = Material (preencher apenas CATMAT)', False, 11),
        ('  SM = Misto / Serviço+Material (preencher ambos)', False, 11),
    ]

    for texto, negrito, size in instrucoes:
        row_idx = ws.max_row + 1
        c = ws.cell(row=row_idx, column=1, value=texto)
        c.font = Font(name='Calibri', bold=negrito, size=size)
        c.alignment = Alignment(wrap_text=True, vertical='top')
        if negrito and size > 11:
            c.fill = make_fill('FFF2CC')
        ws.row_dimensions[row_idx].height = 18 if texto else 8

    return ws


# ── Construir aba principal ────────────────────────────────────────────────────

def criar_aba_contratos(wb, linhas):
    ws = wb.create_sheet('Contratos')
    ws.sheet_properties.tabColor = COR_DADOS

    # ── Cabeçalhos (linhas 1 e 2) ──
    for col_idx, (grupo, campo, cor, _, _chave) in enumerate(COLUNAS, start=1):
        for row_idx, valor in ((1, grupo), (2, campo)):
            cell = ws.cell(row=row_idx, column=col_idx, value=valor)
            cell.font      = H_FONT
            cell.fill      = make_fill(cor)
            cell.alignment = H_ALIGN
            cell.border    = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = COLUNAS[col_idx - 1][3]

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 22
    ws.freeze_panes = 'A3'

    # ── Dados ──
    for data_row_idx, linha in enumerate(linhas, start=3):
        fill_bg = FILL_EVEN if data_row_idx % 2 == 0 else FILL_ODD

        for col_idx, (grupo, campo, cor, _, chave) in enumerate(COLUNAS, start=1):
            valor = linha.get(chave, '')

            cell = ws.cell(row=data_row_idx, column=col_idx, value=valor)
            cell.font      = D_FONT
            cell.border    = THIN_BORDER

            # Status: cor verde/amarelo/vermelho no texto
            if grupo == 'STATUS':
                cell.alignment = C_ALIGN
                if valor and valor.startswith('✓'):
                    cell.font = Font(name='Calibri', size=10, color='1F6B35', bold=True)
                elif valor and valor.startswith('⚠'):
                    cell.font = Font(name='Calibri', size=10, color='C55A11', bold=True)
                else:
                    cell.fill = fill_bg
            else:
                cell.alignment = L_ALIGN
                cell.fill      = fill_bg

        ws.row_dimensions[data_row_idx].height = 15

    # Totais no rodapé
    ultima_linha = len(linhas) + 3
    tipificados  = sum(1 for l in linhas if l.get('_tipificado'))
    pendentes    = len(linhas) - tipificados

    ws.cell(row=ultima_linha, column=1, value='TOTAL').font = BOLD_FONT
    ws.cell(row=ultima_linha, column=2, value=f'{len(linhas)} contratos — {tipificados} tipificados / {pendentes} pendentes').font = D_FONT

    return ws


# ── Ponto de entrada ───────────────────────────────────────────────────────────

def main():
    print('=' * 60)
    print('Relatório de Tipificação de Contratos — SGC')
    print('=' * 60)
    if SITUACAO:
        print(f'  Filtro: situacao = {SITUACAO}')
    if APENAS_PENDENTES:
        print('  Filtro: apenas contratos com tipificação pendente')
    print()

    dados = carregar_dados()
    contratos = dados['contratos']

    print('[6/6] Montando linhas...')
    linhas = [montar_linha(r, dados) for r in contratos]

    if APENAS_PENDENTES:
        linhas = [l for l in linhas if not l['_tipificado']]
        print(f'     Após filtro --apenas-pendentes: {len(linhas)} linhas.')

    if not linhas:
        print('\nNenhum contrato encontrado com os filtros aplicados.')
        return

    # Workbook
    wb = Workbook()
    # Remove a aba padrão "Sheet"
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    criar_aba_instrucoes(wb)
    criar_aba_contratos(wb, linhas)

    # Salvar
    nome_arquivo = f'relatorio_tipificacao_contratos_{date.today().strftime("%Y%m%d")}.xlsx'
    caminho = os.path.join(RELATORIOS, nome_arquivo)
    wb.save(caminho)

    tipificados = sum(1 for l in linhas if l['_tipificado'])
    pendentes   = len(linhas) - tipificados
    print()
    print(f'Arquivo gerado: {caminho}')
    print(f'Total: {len(linhas)} contratos | Tipificados: {tipificados} | Pendentes: {pendentes}')
    print()
    print('Próximo passo: após preencher as colunas laranja/verde,')
    print('execute: python scripts/importar_tipificacao_excel.py --dry-run')


if __name__ == '__main__':
    main()
