"""
Relatório de pendências: contratos sem itens vinculados e sem fiscal.
Gera um Excel com 2 abas.

Uso: python scripts/relatorio_pendencias.py
"""
import pymysql
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date

conn = pymysql.connect(host='localhost', user='root', password='root', database='sgc', charset='utf8mb4')
cur = conn.cursor()

# ── Contratos sem itens vinculados ──
cur.execute('''
    SELECT c.codigo, c.numeroOriginal, c.nomeContratado, c.objeto,
           c.situacao, c.modalidade, c.valor
    FROM contratos c
    WHERE NOT EXISTS (
        SELECT 1 FROM itens_vinculados iv WHERE iv.codigo_contrato = c.codigo
    )
    ORDER BY c.codigo DESC
''')
sem_itens = cur.fetchall()

# ── Contratos sem fiscal ──
cur.execute('''
    SELECT c.codigo, c.numeroOriginal, c.nomeContratado, c.objeto,
           c.situacao, c.modalidade, c.valor
    FROM contratos c
    WHERE NOT EXISTS (
        SELECT 1 FROM fiscais_contrato fc WHERE fc.codigo_contrato = c.codigo
    )
    ORDER BY c.codigo DESC
''')
sem_fiscal = cur.fetchall()

conn.close()

# ── Estilos ──
header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
header_fill = PatternFill(start_color='343990', end_color='343990', fill_type='solid')
header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
thin_border = Border(
    left=Side(style='thin', color='D0D0D0'),
    right=Side(style='thin', color='D0D0D0'),
    top=Side(style='thin', color='D0D0D0'),
    bottom=Side(style='thin', color='D0D0D0')
)
wrap_align = Alignment(vertical='top', wrap_text=True)
money_fmt = '#,##0.00'
headers = ['Codigo', 'Numero Original', 'Contratado', 'Objeto', 'Situacao', 'Modalidade', 'Valor']
col_widths = [12, 14, 35, 50, 14, 22, 16]


def preencher_aba(ws, dados):
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for row_idx, row in enumerate(dados, 2):
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val or '')
            cell.border = thin_border
            if col_idx == 4:
                cell.alignment = wrap_align
            if col_idx == 7:
                cell.value = val or 0
                cell.number_format = money_fmt

    for col_idx, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    ws.auto_filter.ref = f"A1:G{len(dados) + 1}"


wb = Workbook()

ws1 = wb.active
ws1.title = f'Sem Itens ({len(sem_itens)})'
preencher_aba(ws1, sem_itens)

ws2 = wb.create_sheet(f'Sem Fiscal ({len(sem_fiscal)})')
preencher_aba(ws2, sem_fiscal)

filename = f'pendencias_contratos_{date.today().strftime("%Y%m%d")}.xlsx'
wb.save(filename)

print(f'Contratos sem itens vinculados: {len(sem_itens)}')
print(f'Contratos sem fiscal: {len(sem_fiscal)}')
print(f'Arquivo gerado: {filename}')
