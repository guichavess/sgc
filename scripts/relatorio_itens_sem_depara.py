"""
Relatório de itens_contrato que ainda NÃO possuem o de-para
(catserv_servico_id IS NULL AND catmat_item_id IS NULL).

Gera um Excel com 2 abas:
  1. "Vinculados sem de-para" — itens que estão vinculados a contratos, mas sem CATSERV/CATMAT
  2. "Todos sem de-para"     — todos os itens sem de-para (independente de vínculo)

Uso:
    python scripts/relatorio_itens_sem_depara.py
"""
import os
import sys
from datetime import datetime

import pymysql
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from dotenv import load_dotenv

# ── Config ──
base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
load_dotenv(os.path.join(base_dir, '.env'))

DB_CONFIG = {
    'host': os.getenv('DB_HOST', 'localhost'),
    'user': os.getenv('DB_USER', 'root'),
    'password': os.getenv('DB_PASS', 'root'),
    'database': os.getenv('DB_NAME', 'sgc'),
    'charset': 'utf8mb4',
}

OUTPUT_DIR = os.path.join(base_dir, 'scripts')
OUTPUT_FILE = os.path.join(OUTPUT_DIR, 'relatorio_itens_sem_depara.xlsx')

# ── Estilo ──
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
HEADER_FILL_RED = PatternFill(start_color='C0392B', end_color='C0392B', fill_type='solid')
HEADER_FILL_ORANGE = PatternFill(start_color='E67E22', end_color='E67E22', fill_type='solid')
HEADER_ALIGNMENT = Alignment(horizontal='center', vertical='center', wrap_text=True)
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)


def apply_header(ws, headers, fill):
    """Aplica cabeçalho estilizado na primeira linha."""
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = fill
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions


def auto_width(ws):
    """Ajusta largura das colunas automaticamente."""
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 60)


def main():
    conn = pymysql.connect(**DB_CONFIG)
    cur = conn.cursor()

    # ── 1. Itens vinculados a contratos MAS sem de-para ──
    cur.execute('''
        SELECT
            iv.codigo_contrato,
            c.objeto,
            ic.id AS item_id,
            ic.descricao,
            ic.tipo_item
        FROM itens_contrato ic
        JOIN itens_vinculados iv ON iv.item_contrato_id = ic.id
        JOIN contratos c ON c.codigo = iv.codigo_contrato
        WHERE ic.catserv_servico_id IS NULL
          AND ic.catmat_item_id IS NULL
        ORDER BY iv.codigo_contrato, ic.descricao
    ''')
    vinculados_sem = cur.fetchall()

    # ── 2. Todos os itens sem de-para ──
    cur.execute('''
        SELECT
            ic.id,
            ic.descricao,
            ic.tipo_item,
            GROUP_CONCAT(DISTINCT iv.codigo_contrato ORDER BY iv.codigo_contrato SEPARATOR ', ') AS contratos
        FROM itens_contrato ic
        LEFT JOIN itens_vinculados iv ON iv.item_contrato_id = ic.id
        WHERE ic.catserv_servico_id IS NULL
          AND ic.catmat_item_id IS NULL
        GROUP BY ic.id, ic.descricao, ic.tipo_item
        ORDER BY ic.descricao
    ''')
    todos_sem = cur.fetchall()

    # ── Totais para console ──
    cur.execute('SELECT COUNT(*) FROM itens_contrato')
    total_itens = cur.fetchone()[0]

    cur.execute('''
        SELECT COUNT(*) FROM itens_contrato
        WHERE catserv_servico_id IS NOT NULL OR catmat_item_id IS NOT NULL
    ''')
    total_com = cur.fetchone()[0]

    conn.close()

    # ── Gerar Excel ──
    wb = Workbook()

    # --- Aba 1: Vinculados sem de-para ---
    ws1 = wb.active
    ws1.title = 'Vinculados sem de-para'
    headers1 = ['Codigo Contrato', 'Objeto do Contrato', 'Item ID', 'Descricao do Item', 'Tipo Item']
    apply_header(ws1, headers1, HEADER_FILL_RED)

    for row_data in vinculados_sem:
        ws1.append(list(row_data))

    auto_width(ws1)

    # --- Aba 2: Todos sem de-para ---
    ws2 = wb.create_sheet('Todos sem de-para')
    headers2 = ['Item ID', 'Descricao do Item', 'Tipo Item', 'Contratos Vinculados']
    apply_header(ws2, headers2, HEADER_FILL_ORANGE)

    for row_data in todos_sem:
        ws2.append(list(row_data))

    auto_width(ws2)

    # ── Salvar ──
    wb.save(OUTPUT_FILE)

    # ── Console ──
    print('=' * 60)
    print('  RELATORIO: ITENS SEM DE-PARA (CATSERV/CATMAT)')
    print('=' * 60)
    print(f'\n  Total itens_contrato:              {total_itens}')
    print(f'  COM de-para:                       {total_com}')
    print(f'  SEM de-para (total):               {len(todos_sem)}')
    print(f'  SEM de-para (vinculados a contrato): {len(vinculados_sem)}')
    print(f'\n  Arquivo gerado: {OUTPUT_FILE}')
    print(f'  Data: {datetime.now().strftime("%d/%m/%Y %H:%M:%S")}')
    print('=' * 60)


if __name__ == '__main__':
    main()
