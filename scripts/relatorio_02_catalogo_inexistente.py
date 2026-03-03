'''
Relatorio 02 - IDs de Catalogo Inexistentes
=============================================
Gera Excel listando IDs de Material (CATMAT) e Servico (CATSERV)
presentes no arquivo de vinculacao que NAO existem nas tabelas de catalogo.

Uso: python scripts/relatorio_02_catalogo_inexistente.py
'''
import os
import unicodedata
from datetime import datetime

import pandas as pd
import pymysql
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
load_dotenv(os.path.join(BASE_DIR, '.env'))

VINC_PATH = os.path.join(BASE_DIR, 'itens vinculação.xlsx')
OUTPUT = os.path.join(BASE_DIR, f'relatorio_02_catalogo_inexistente_{datetime.now():%Y%m%d_%H%M}.xlsx')

DB = dict(host=os.getenv('DB_HOST','localhost'), user=os.getenv('DB_USER','root'),
          password=os.getenv('DB_PASS',''), database=os.getenv('DB_NAME','sgc'), charset='utf8mb4')

HDR_FILL = PatternFill('solid', fgColor='1F4E79')
HDR_FONT = Font(bold=True, color='FFFFFF', size=11)
TITLE_FONT = Font(bold=True, size=14, color='1F4E79')
SUBTITLE_FONT = Font(bold=True, size=12, color='1F4E79')
LABEL_FONT = Font(bold=True, size=11)
THIN = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)


def norm_tipo(t):
    n = unicodedata.normalize('NFKD', str(t)).encode('ascii','ignore').decode().upper().strip()
    return 'S' if 'SERVIC' in n else 'M'


def main():
    print('Relatorio 02 - Catalogo Inexistente')
    print('='*50)

    # Ler vinculacao
    print('[1/3] Lendo arquivo Excel...')
    df = pd.read_excel(VINC_PATH)
    registros = []
    for _, row in df.iterrows():
        siafe = row.iloc[0]
        item_id = row.iloc[1]
        tipo = row.iloc[2]
        if pd.isna(siafe) or str(siafe).strip() in ('', '-'):
            continue
        if pd.isna(item_id):
            continue
        registros.append({
            'siafe': str(int(float(str(siafe)))),
            'id': int(float(str(item_id))),
            'tipo': norm_tipo(tipo),
        })
    print(f'  Registros validos: {len(registros)}')

    # Consultar catalogo
    print('[2/3] Consultando banco de dados...')
    conn = pymysql.connect(**DB)
    cur = conn.cursor()

    cur.execute("SELECT codigo_servico, nome FROM catserv_servicos")
    catserv = {r[0]: r[1] for r in cur.fetchall()}

    cur.execute("SELECT codigo, descricao FROM catmat_itens")
    catmat_itens = {r[0]: r[1] for r in cur.fetchall()}

    cur.execute("SELECT codigo, nome FROM catmat_pdms")
    catmat_pdms = {r[0]: r[1] for r in cur.fetchall()}

    # Contratos para referencia
    cur.execute("SELECT codigo, numeroOriginal FROM contratos")
    contratos = {str(r[0]): (r[1] or '') for r in cur.fetchall()}

    conn.close()

    # Analisar material
    mat_ids = set()
    mat_nao_encontrados = {}
    serv_ids = set()
    serv_nao_encontrados = {}

    for r in registros:
        if r['tipo'] == 'M':
            if r['id'] not in mat_ids:
                mat_ids.add(r['id'])
                if r['id'] not in catmat_itens and r['id'] not in catmat_pdms:
                    mat_nao_encontrados.setdefault(r['id'], []).append(r['siafe'])
                elif r['id'] not in mat_nao_encontrados:
                    pass
            else:
                if r['id'] in mat_nao_encontrados:
                    mat_nao_encontrados[r['id']].append(r['siafe'])
        elif r['tipo'] == 'S':
            if r['id'] not in serv_ids:
                serv_ids.add(r['id'])
                if r['id'] not in catserv:
                    serv_nao_encontrados.setdefault(r['id'], []).append(r['siafe'])
                elif r['id'] not in serv_nao_encontrados:
                    pass
            else:
                if r['id'] in serv_nao_encontrados:
                    serv_nao_encontrados[r['id']].append(r['siafe'])

    print(f'  IDs material verificados: {len(mat_ids)} | Nao encontrados: {len(mat_nao_encontrados)}')
    print(f'  IDs servico verificados: {len(serv_ids)} | Nao encontrados: {len(serv_nao_encontrados)}')

    # Gerar Excel
    print('[3/3] Gerando relatorio...')
    wb = Workbook()

    # --- Aba 1: Material ---
    ws1 = wb.active
    ws1.title = 'Material Inexistente'
    ws1.sheet_properties.tabColor = 'FF0000'

    ws1.merge_cells('A1:E1')
    ws1['A1'].value = 'IDs DE MATERIAL NAO ENCONTRADOS NO CATMAT'
    ws1['A1'].font = TITLE_FONT

    ws1['A3'].value = f'Data: {datetime.now():%d/%m/%Y %H:%M}'
    ws1['A3'].font = LABEL_FONT
    ws1['A4'].value = f'Total IDs material verificados: {len(mat_ids)}'
    ws1['A4'].font = LABEL_FONT
    ws1['A5'].value = f'Total IDs NAO encontrados: {len(mat_nao_encontrados)}'
    ws1['A5'].font = LABEL_FONT
    ws1['A6'].value = 'Verificado em: catmat_itens e catmat_pdms'
    ws1['A6'].font = Font(italic=True, size=10, color='666666')

    headers = ['ID (Codigo)', 'Qtd Contratos', 'Contratos (N SIAFE)', 'Num. Contratos', 'Observacao']
    row = 8
    for i, h in enumerate(headers, 1):
        c = ws1.cell(row=row, column=i, value=h)
        c.font = HDR_FONT; c.fill = HDR_FILL; c.alignment = CENTER; c.border = THIN

    for idx, (mid, siafes) in enumerate(sorted(mat_nao_encontrados.items()), row + 1):
        uniq = sorted(set(siafes))
        nums = [contratos.get(s, '?') for s in uniq]
        for ci, val in enumerate([mid, len(uniq), ', '.join(uniq), ', '.join(nums),
                                   'ID nao existe em catmat_itens nem catmat_pdms'], 1):
            cell = ws1.cell(row=idx, column=ci, value=val)
            cell.border = THIN

    ws1.auto_filter.ref = f'A{row}:{get_column_letter(len(headers))}{ws1.max_row}'
    ws1.freeze_panes = f'A{row+1}'
    for col, w in {'A':15, 'B':15, 'C':45, 'D':45, 'E':50}.items():
        ws1.column_dimensions[col].width = w

    # --- Aba 2: Servico ---
    ws2 = wb.create_sheet('Servico Inexistente')
    ws2.sheet_properties.tabColor = 'FF0000'

    ws2.merge_cells('A1:E1')
    ws2['A1'].value = 'IDs DE SERVICO NAO ENCONTRADOS NO CATSERV'
    ws2['A1'].font = TITLE_FONT

    ws2['A3'].value = f'Data: {datetime.now():%d/%m/%Y %H:%M}'
    ws2['A3'].font = LABEL_FONT
    ws2['A4'].value = f'Total IDs servico verificados: {len(serv_ids)}'
    ws2['A4'].font = LABEL_FONT
    ws2['A5'].value = f'Total IDs NAO encontrados: {len(serv_nao_encontrados)}'
    ws2['A5'].font = LABEL_FONT

    headers2 = ['ID (Codigo Servico)', 'Qtd Contratos', 'Contratos (N SIAFE)', 'Num. Contratos', 'Observacao']
    row = 7
    for i, h in enumerate(headers2, 1):
        c = ws2.cell(row=row, column=i, value=h)
        c.font = HDR_FONT; c.fill = HDR_FILL; c.alignment = CENTER; c.border = THIN

    for idx, (sid, siafes) in enumerate(sorted(serv_nao_encontrados.items()), row + 1):
        uniq = sorted(set(siafes))
        nums = [contratos.get(s, '?') for s in uniq]
        for ci, val in enumerate([sid, len(uniq), ', '.join(uniq), ', '.join(nums),
                                   'ID nao existe em catserv_servicos'], 1):
            cell = ws2.cell(row=idx, column=ci, value=val)
            cell.border = THIN

    ws2.auto_filter.ref = f'A{row}:{get_column_letter(len(headers2))}{ws2.max_row}'
    ws2.freeze_panes = f'A{row+1}'
    for col, w in {'A':20, 'B':15, 'C':45, 'D':45, 'E':45}.items():
        ws2.column_dimensions[col].width = w

    # --- Aba 3: Resumo de validacao ---
    ws3 = wb.create_sheet('Resumo Validacao')
    ws3.sheet_properties.tabColor = '4472C4'

    ws3.merge_cells('A1:C1')
    ws3['A1'].value = 'RESUMO DA VALIDACAO DE CATALOGO'
    ws3['A1'].font = TITLE_FONT

    data_resumo = [
        ('MATERIAL', '', ''),
        ('IDs unicos verificados', len(mat_ids), ''),
        ('Encontrados em catmat_itens', len([m for m in mat_ids if m in catmat_itens]), ''),
        ('Encontrados em catmat_pdms', len([m for m in mat_ids if m not in catmat_itens and m in catmat_pdms]), ''),
        ('NAO encontrados', len(mat_nao_encontrados), 'Precisam ser importados ou corrigidos'),
        ('', '', ''),
        ('SERVICO', '', ''),
        ('IDs unicos verificados', len(serv_ids), ''),
        ('Encontrados em catserv_servicos', len([s for s in serv_ids if s in catserv]), ''),
        ('NAO encontrados', len(serv_nao_encontrados), 'Precisam ser importados ou corrigidos'),
    ]

    for i, (label, val, obs) in enumerate(data_resumo, 3):
        ws3.cell(row=i, column=1, value=label).font = LABEL_FONT if val == '' else Font(size=11)
        ws3.cell(row=i, column=2, value=val if val != '' else None)
        ws3.cell(row=i, column=3, value=obs).font = Font(italic=True, color='CC0000') if obs else Font()

    ws3.column_dimensions['A'].width = 35
    ws3.column_dimensions['B'].width = 15
    ws3.column_dimensions['C'].width = 45

    wb.save(OUTPUT)
    print(f'\nSalvo em: {OUTPUT}')
    print('='*50)


if __name__ == '__main__':
    main()
