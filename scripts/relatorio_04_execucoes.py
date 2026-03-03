'''
Relatorio 04 - Execucoes (Novas + Atualizadas)
================================================
Gera Excel detalhando as execucoes que serao inseridas e
as que serao atualizadas (com valores originais vs novos).

Uso: python scripts/relatorio_04_execucoes.py
'''
import os
import unicodedata
from datetime import datetime

import pandas as pd
import pymysql
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
load_dotenv(os.path.join(BASE_DIR, '.env'))

CORR_PATH = os.path.join(BASE_DIR, 'itens correção.xlsx')
OUTPUT = os.path.join(BASE_DIR, f'relatorio_04_execucoes_{datetime.now():%Y%m%d_%H%M}.xlsx')

DB = dict(host=os.getenv('DB_HOST','localhost'), user=os.getenv('DB_USER','root'),
          password=os.getenv('DB_PASS',''), database=os.getenv('DB_NAME','sgc'), charset='utf8mb4')

HDR_FILL = PatternFill('solid', fgColor='1F4E79')
HDR_FONT = Font(bold=True, color='FFFFFF', size=11)
TITLE_FONT = Font(bold=True, size=14, color='1F4E79')
LABEL_FONT = Font(bold=True, size=11)
THIN = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
FILL_GREEN = PatternFill('solid', fgColor='E2EFDA')
FILL_YELLOW = PatternFill('solid', fgColor='FFF9C4')
FILL_GRAY = PatternFill('solid', fgColor='F2F2F2')
BRL = '#,##0.00'


def norm(t):
    if not t: return ''
    return unicodedata.normalize('NFKD', str(t)).encode('ascii','ignore').decode().upper().strip()


def write_hdr(ws, row, headers):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = HDR_FONT; c.fill = HDR_FILL; c.alignment = CENTER; c.border = THIN


def write_row(ws, row, values, currency_cols=None, fill=None):
    currency_cols = currency_cols or set()
    for i, v in enumerate(values, 1):
        c = ws.cell(row=row, column=i, value=v)
        c.border = THIN
        c.alignment = Alignment(vertical='center', wrap_text=isinstance(v, str) and len(str(v)) > 25)
        if i in currency_cols and v is not None:
            c.number_format = BRL
        if fill:
            c.fill = fill


def main():
    print('Relatorio 04 - Execucoes')
    print('='*50)

    # Ler Excel
    print('[1/3] Lendo arquivo Excel...')
    df = pd.read_excel(CORR_PATH)
    registros = []
    for _, row in df.iterrows():
        siafe = row.iloc[0]
        item = row.iloc[1]
        data_valor = row.iloc[2]
        valor = row.iloc[3]
        qtde = row.iloc[4]
        contratado = row.iloc[5]

        if pd.isna(siafe):
            continue

        dt = data_valor if isinstance(data_valor, datetime) else None
        try:
            v = float(valor) if not pd.isna(valor) else None
        except (ValueError, TypeError):
            v = None
        try:
            q = int(float(str(qtde))) if not pd.isna(qtde) and str(qtde).strip() != '-' else None
        except (ValueError, TypeError):
            q = None

        registros.append({
            'siafe': str(int(float(str(siafe)))),
            'item': str(item).strip() if not pd.isna(item) else '',
            'data': dt,
            'mes': dt.month if dt else None,
            'ano': dt.year if dt else None,
            'valor': v,
            'qtde': q,
            'contratado': str(contratado).strip() if not pd.isna(contratado) else '',
        })
    print(f'  Registros: {len(registros)}')

    # Consultar BD
    print('[2/3] Consultando banco de dados...')
    conn = pymysql.connect(**DB)
    cur = conn.cursor()

    cur.execute("SELECT codigo, numeroOriginal, nomeContratado FROM contratos")
    contratos = {}
    for r in cur.fetchall():
        contratos[str(r[0])] = {'numeroOriginal': r[1] or '', 'contratado': r[2] or ''}

    cur.execute("""
        SELECT e.id, e.codigo_contrato, e.itens_contrato_id,
               e.mes, e.ano, e.valor, e.quantidade,
               ic.descricao AS desc_item
        FROM execucoes e
        LEFT JOIN itens_contrato ic ON e.itens_contrato_id = ic.id
    """)
    db_exec = cur.fetchall()
    conn.close()

    # Indexar execucoes BD por (contrato, item_norm, mes, ano)
    db_idx = {}
    for r in db_exec:
        desc = norm(r[7]) if r[7] else ''
        key = (str(r[1]), desc, r[3], r[4])
        db_idx[key] = {
            'id': r[0], 'valor': float(r[5]) if r[5] is not None else None,
            'qtde': int(r[6]) if r[6] is not None else None,
        }

    # Classificar
    novas = []
    atualizadas = []
    iguais = []
    sem_valor = []
    ignoradas = []

    for c in registros:
        if c['siafe'] not in contratos:
            ignoradas.append({**c, 'motivo': 'Contrato inexistente'})
            continue
        if c['valor'] is None and c['qtde'] is None:
            sem_valor.append(c)
            continue

        key = (c['siafe'], norm(c['item']), c['mes'], c['ano'])
        if key in db_idx:
            db = db_idx[key]
            db_val = db['valor']
            db_qtd = db['qtde']
            val_diff = abs((db_val or 0) - (c['valor'] or 0)) > 0.005
            qtd_diff = (db_qtd or 0) != (c['qtde'] or 0)
            if val_diff or qtd_diff:
                atualizadas.append({
                    **c, 'db_id': db['id'],
                    'db_valor': db_val, 'db_qtde': db_qtd,
                    'diff_valor': val_diff, 'diff_qtde': qtd_diff,
                })
            else:
                iguais.append(c)
        else:
            novas.append(c)

    print(f'  Novas: {len(novas)}')
    print(f'  Atualizadas: {len(atualizadas)}')
    print(f'  Iguais: {len(iguais)}')
    print(f'  Sem valor: {len(sem_valor)}')
    print(f'  Ignoradas: {len(ignoradas)}')

    # Totalizar valores
    total_novas_valor = sum(n['valor'] or 0 for n in novas)
    total_atualizadas_antes = sum(a['db_valor'] or 0 for a in atualizadas)
    total_atualizadas_depois = sum(a['valor'] or 0 for a in atualizadas)

    # Gerar Excel
    print('[3/3] Gerando relatorio...')
    wb = Workbook()

    # --- Aba Resumo ---
    ws0 = wb.active
    ws0.title = 'Resumo'
    ws0.sheet_properties.tabColor = '4472C4'
    ws0.merge_cells('A1:C1')
    ws0['A1'].value = 'RELATORIO DE EXECUCOES'
    ws0['A1'].font = TITLE_FONT
    ws0['A3'].value = f'Data: {datetime.now():%d/%m/%Y %H:%M}'
    ws0['A3'].font = LABEL_FONT
    ws0['A4'].value = 'Arquivo: itens correcao.xlsx'

    resumo = [
        ('', '', ''),
        ('QUANTITATIVOS', '', ''),
        ('Total linhas no Excel', len(registros), ''),
        ('Novas execucoes (INSERT)', len(novas), ''),
        ('Execucoes atualizadas (UPDATE)', len(atualizadas), ''),
        ('Ja existem (sem alteracao)', len(iguais), ''),
        ('Sem valor/quantidade', len(sem_valor), ''),
        ('Contrato inexistente (ignoradas)', len(ignoradas), ''),
        ('', '', ''),
        ('VALORES', '', ''),
        ('Valor total das novas execucoes', total_novas_valor, 'R$'),
        ('Valor total ANTES das atualizacoes', total_atualizadas_antes, 'R$'),
        ('Valor total DEPOIS das atualizacoes', total_atualizadas_depois, 'R$'),
        ('Diferenca liquida', total_atualizadas_depois - total_atualizadas_antes, 'R$'),
    ]
    for i, (label, val, fmt) in enumerate(resumo, 6):
        c1 = ws0.cell(row=i, column=1, value=label)
        c2 = ws0.cell(row=i, column=2, value=val if val != '' else None)
        if label in ('QUANTITATIVOS', 'VALORES'):
            c1.font = Font(bold=True, size=12, color='1F4E79')
        else:
            c1.font = LABEL_FONT
        if fmt == 'R$' and val is not None:
            c2.number_format = BRL

    ws0.column_dimensions['A'].width = 45
    ws0.column_dimensions['B'].width = 20

    # --- Aba Atualizadas ---
    ws1 = wb.create_sheet('Atualizadas (UPDATE)')
    ws1.sheet_properties.tabColor = 'ED7D31'
    ws1.merge_cells('A1:K1')
    ws1['A1'].value = 'EXECUCOES QUE SERAO ATUALIZADAS (valores originais vs novos)'
    ws1['A1'].font = TITLE_FONT
    ws1['A2'].value = f'Total: {len(atualizadas)} registros'
    ws1['A2'].font = LABEL_FONT

    headers = ['N SIAFE', 'Num. Original', 'Item', 'Mes', 'Ano', 'Contratado',
               'Valor Anterior (BD)', 'Valor Novo (Excel)', 'Diferenca Valor',
               'Qtde Anterior (BD)', 'Qtde Novo (Excel)']
    write_hdr(ws1, 4, headers)
    curr = {7, 8, 9}
    for i, a in enumerate(atualizadas, 5):
        ct = contratos.get(a['siafe'], {})
        dv = (a['valor'] or 0) - (a['db_valor'] or 0) if a['diff_valor'] else 0
        write_row(ws1, i, [
            a['siafe'], ct.get('numeroOriginal', ''), a['item'],
            a['mes'], a['ano'], a['contratado'],
            a['db_valor'], a['valor'], round(dv, 2),
            a['db_qtde'], a['qtde'],
        ], currency_cols=curr, fill=FILL_YELLOW)
    if atualizadas:
        ws1.auto_filter.ref = f'A4:{get_column_letter(len(headers))}{ws1.max_row}'
    ws1.freeze_panes = 'A5'
    for col, w in zip('ABCDEFGHIJK', [14, 18, 35, 6, 6, 25, 18, 18, 16, 16, 16]):
        ws1.column_dimensions[col].width = w

    # --- Aba Novas ---
    ws2 = wb.create_sheet('Novas (INSERT)')
    ws2.sheet_properties.tabColor = '00B050'
    ws2.merge_cells('A1:H1')
    ws2['A1'].value = 'NOVAS EXECUCOES A INSERIR'
    ws2['A1'].font = TITLE_FONT
    ws2['A2'].value = f'Total: {len(novas)} registros | Valor total: R$ {total_novas_valor:,.2f}'
    ws2['A2'].font = LABEL_FONT

    headers2 = ['N SIAFE', 'Num. Original', 'Item', 'Mes', 'Ano', 'Valor', 'Qtde', 'Contratado']
    write_hdr(ws2, 4, headers2)
    for i, n in enumerate(novas, 5):
        ct = contratos.get(n['siafe'], {})
        write_row(ws2, i, [
            n['siafe'], ct.get('numeroOriginal', ''), n['item'],
            n['mes'], n['ano'], n['valor'], n['qtde'], n['contratado'],
        ], currency_cols={6}, fill=FILL_GREEN)
    if novas:
        ws2.auto_filter.ref = f'A4:{get_column_letter(len(headers2))}{ws2.max_row}'
    ws2.freeze_panes = 'A5'
    for col, w in zip('ABCDEFGH', [14, 18, 35, 6, 6, 16, 10, 25]):
        ws2.column_dimensions[col].width = w

    # --- Aba Sem valor ---
    ws3 = wb.create_sheet('Sem valor')
    ws3.sheet_properties.tabColor = 'A6A6A6'
    ws3.merge_cells('A1:F1')
    ws3['A1'].value = 'EXECUCOES SEM VALOR/QUANTIDADE (nao serao importadas)'
    ws3['A1'].font = TITLE_FONT
    ws3['A2'].value = f'Total: {len(sem_valor)} registros'
    ws3['A2'].font = LABEL_FONT

    headers3 = ['N SIAFE', 'Item', 'Mes', 'Ano', 'Contratado', 'Observacao']
    write_hdr(ws3, 4, headers3)
    for i, sv in enumerate(sem_valor, 5):
        write_row(ws3, i, [
            sv['siafe'], sv['item'], sv['mes'], sv['ano'], sv['contratado'],
            'Valor e quantidade ausentes no Excel',
        ], fill=FILL_GRAY)
    if sem_valor:
        ws3.auto_filter.ref = f'A4:{get_column_letter(len(headers3))}{ws3.max_row}'
    ws3.freeze_panes = 'A5'
    for col, w in zip('ABCDEF', [14, 35, 6, 6, 25, 40]):
        ws3.column_dimensions[col].width = w

    # --- Aba Ignoradas ---
    ws4 = wb.create_sheet('Ignoradas')
    ws4.sheet_properties.tabColor = 'FF0000'
    ws4.merge_cells('A1:F1')
    ws4['A1'].value = 'EXECUCOES IGNORADAS (contrato inexistente)'
    ws4['A1'].font = TITLE_FONT
    ws4['A2'].value = f'Total: {len(ignoradas)} registros'
    ws4['A2'].font = LABEL_FONT

    headers4 = ['N SIAFE', 'Item', 'Mes', 'Ano', 'Contratado', 'Motivo']
    write_hdr(ws4, 4, headers4)
    for i, ig in enumerate(ignoradas, 5):
        write_row(ws4, i, [
            ig['siafe'], ig['item'], ig['mes'], ig['ano'], ig['contratado'], ig['motivo'],
        ], fill=PatternFill('solid', fgColor='FCE4EC'))
    if ignoradas:
        ws4.auto_filter.ref = f'A4:{get_column_letter(len(headers4))}{ws4.max_row}'
    ws4.freeze_panes = 'A5'
    for col, w in zip('ABCDEF', [14, 35, 6, 6, 25, 30]):
        ws4.column_dimensions[col].width = w

    wb.save(OUTPUT)
    print(f'\nSalvo em: {OUTPUT}')
    print('='*50)


if __name__ == '__main__':
    main()
