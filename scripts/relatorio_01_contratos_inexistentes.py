'''
Relatorio 01 - Contratos Inexistentes
======================================
Gera Excel listando todos os N SIAFE presentes nos arquivos
de importacao que NAO existem na tabela contratos do banco.

Uso: python scripts/relatorio_01_contratos_inexistentes.py
'''
import os
from datetime import datetime

import pandas as pd
import pymysql
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
load_dotenv(os.path.join(BASE_DIR, '.env'))

VINC_PATH = os.path.join(BASE_DIR, 'itens vinculação.xlsx')
CORR_PATH = os.path.join(BASE_DIR, 'itens correção.xlsx')
OUTPUT = os.path.join(BASE_DIR, f'relatorio_01_contratos_inexistentes_{datetime.now():%Y%m%d_%H%M}.xlsx')

DB = dict(host=os.getenv('DB_HOST','localhost'), user=os.getenv('DB_USER','root'),
          password=os.getenv('DB_PASS',''), database=os.getenv('DB_NAME','sgc'), charset='utf8mb4')

# Estilos
HDR_FILL = PatternFill('solid', fgColor='1F4E79')
HDR_FONT = Font(bold=True, color='FFFFFF', size=11)
TITLE_FONT = Font(bold=True, size=14, color='1F4E79')
LABEL_FONT = Font(bold=True, size=11)
THIN = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)


def main():
    print('Relatorio 01 - Contratos Inexistentes')
    print('='*50)

    # Ler Excel vinculacao
    print('[1/3] Lendo arquivos Excel...')
    df_vinc = pd.read_excel(VINC_PATH)
    df_corr = pd.read_excel(CORR_PATH)

    # Extrair SIAFEs unicos
    siafes_vinc = set()
    for v in df_vinc.iloc[:, 0].dropna():
        s = str(v).strip().replace('-', '')
        if s:
            try:
                siafes_vinc.add(str(int(float(s))))
            except ValueError:
                pass

    siafes_corr = set()
    for v in df_corr.iloc[:, 0].dropna():
        try:
            siafes_corr.add(str(int(float(str(v)))))
        except ValueError:
            pass

    todos = siafes_vinc | siafes_corr
    print(f'  SIAFEs unicos encontrados: {len(todos)}')

    # Consultar banco
    print('[2/3] Consultando banco de dados...')
    conn = pymysql.connect(**DB)
    with conn.cursor() as cur:
        cur.execute("SELECT codigo FROM contratos")
        db_codigos = {str(r[0]) for r in cur.fetchall()}
    conn.close()

    # Cruzar
    nao_encontrados = sorted(todos - db_codigos)
    print(f'  Contratos inexistentes: {len(nao_encontrados)}')

    # Gerar Excel
    print('[3/3] Gerando relatorio...')
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'Contratos Inexistentes'
    ws.sheet_properties.tabColor = 'FF0000'

    # Titulo
    ws.merge_cells('A1:D1')
    ws['A1'].value = 'CONTRATOS NAO ENCONTRADOS NO BANCO DE DADOS'
    ws['A1'].font = TITLE_FONT

    ws['A3'].value = f'Data do relatorio: {datetime.now():%d/%m/%Y %H:%M}'
    ws['A3'].font = LABEL_FONT
    ws['A4'].value = f'Total de contratos inexistentes: {len(nao_encontrados)}'
    ws['A4'].font = LABEL_FONT
    ws['A5'].value = f'Total de SIAFEs verificados: {len(todos)}'
    ws['A5'].font = LABEL_FONT

    # Header
    headers = ['N SIAFE', 'Presente em Vinculacao', 'Presente em Correcao', 'Observacao']
    row = 7
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = HDR_FONT
        c.fill = HDR_FILL
        c.alignment = CENTER
        c.border = THIN

    # Dados
    for idx, siafe in enumerate(nao_encontrados, row + 1):
        em_vinc = 'Sim' if siafe in siafes_vinc else 'Nao'
        em_corr = 'Sim' if siafe in siafes_corr else 'Nao'
        obs = 'Contrato precisa ser importado antes da vinculacao/execucao'
        for ci, val in enumerate([siafe, em_vinc, em_corr, obs], 1):
            cell = ws.cell(row=idx, column=ci, value=val)
            cell.border = THIN
            if ci in (2, 3):
                cell.alignment = Alignment(horizontal='center')
                if val == 'Sim':
                    cell.fill = PatternFill('solid', fgColor='FCE4EC')

    # Autofilter + largura
    ws.auto_filter.ref = f'A{row}:{get_column_letter(len(headers))}{ws.max_row}'
    ws.freeze_panes = f'A{row+1}'
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 55

    wb.save(OUTPUT)
    print(f'\nSalvo em: {OUTPUT}')
    print('='*50)


if __name__ == '__main__':
    main()
