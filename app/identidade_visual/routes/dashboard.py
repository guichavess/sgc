import io
import os
import uuid
from datetime import datetime
from flask import render_template, request, jsonify, send_file, current_app, url_for
from flask_login import current_user
from app.utils.permissions import requires_admin, requires_permission
from werkzeug.utils import secure_filename
from sqlalchemy import inspect as sa_inspect, func
from app.extensions import db
from app.models.identidade_visual import (
    IdentidadeVisualLocal, IdentidadeVisualArquivo, IdentidadeVisualLog,
    MunicipioPiaui, TIPOS_LOCAL, TIPO_LOCAL_VEICULO,
)
from app.services.detran_service import consultar_veiculo, normalizar_placa, DetranError
from app.identidade_visual.routes import identidade_visual_bp

UPLOAD_FOLDER = os.path.join('app', 'static', 'uploads', 'identidade_visual')
ALLOWED_EXTENSIONS = {'pdf', 'jpg', 'jpeg', 'heic', 'png', 'webp'}
MAX_FILE_SIZE = 25 * 1024 * 1024  # 25 MB por arquivo (mantenha em sync com MAX_FOTO_MB no template)
PER_PAGE = 15


def _allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def _registrar_log(acao, descricao, local_id=None):
    """Adiciona um registro de auditoria à sessão (commit fica a cargo do chamador)."""
    db.session.add(IdentidadeVisualLog(
        local_id=local_id,
        acao=acao,
        descricao=descricao,
        usuario_id=getattr(current_user, 'id', None),
        usuario_nome=getattr(current_user, 'nome', None),
    ))


def _rotulo(local):
    """Rótulo curto de um registro para logs/descrições.

    Veículo → 'Veículo PLACA'; local físico → nome da cidade.
    """
    if local.is_veiculo:
        return f'Veículo {local.placa or ""}'.strip()
    return local.cidade or ''


def _marca_modelo(local):
    """Concatena marca+modelo com "/" para logs/exibições legadas."""
    partes = [p for p in (local.marca, local.modelo) if p]
    return '/'.join(partes)


@identidade_visual_bp.route('/')
@requires_permission('identidade_visual.visualizar')
def dashboard():
    # Filtros multi-seleção (estilo Pagamentos): cada um aceita vários valores.
    cidades_sel = [c for c in request.args.getlist('cidade') if c]
    tipos_sel = [t for t in request.args.getlist('tipo_local') if t]
    bairros_sel = [b for b in request.args.getlist('bairro') if b]
    status_sel = [s for s in request.args.getlist('status') if s]
    page = max(1, int(request.args.get('page', 1) or 1))

    # Quantidade de arquivos por local em UMA query (group by) — base para o
    # cálculo de status sem N+1 (antes cada l.status disparava arquivos.count()).
    contagem_arquivos = dict(
        db.session.query(
            IdentidadeVisualArquivo.local_id,
            func.count(IdentidadeVisualArquivo.id),
        ).group_by(IdentidadeVisualArquivo.local_id).all()
    )

    def _status(lid, data_acao):
        return 'REALIZADO' if (data_acao is not None and contagem_arquivos.get(lid, 0) > 0) else 'PENDENTE'

    # Conjunto FILTRADO, consulta leve (só colunas necessárias p/ ordenar/filtrar/KPI).
    # A aba "Fachadas" mostra apenas locais físicos — veículos têm aba própria.
    base = db.session.query(
        IdentidadeVisualLocal.id,
        IdentidadeVisualLocal.cidade,
        IdentidadeVisualLocal.custo,
        IdentidadeVisualLocal.data_acao,
    ).filter(IdentidadeVisualLocal.tipo_local != TIPO_LOCAL_VEICULO)
    if cidades_sel:
        base = base.filter(IdentidadeVisualLocal.cidade.in_(cidades_sel))
    if tipos_sel:
        base = base.filter(IdentidadeVisualLocal.tipo_local.in_(tipos_sel))
    if bairros_sel:
        base = base.filter(IdentidadeVisualLocal.bairro.in_(bairros_sel))

    registros = []
    for lid, cidade, custo, data_acao in base.all():
        st = _status(lid, data_acao)
        if status_sel and st not in status_sel:
            continue
        registros.append({'id': lid, 'cidade': cidade, 'custo': custo, 'status': st})

    # Prioridade de exibição: PENDENTES primeiro, depois por cidade.
    registros.sort(key=lambda r: (0 if r['status'] == 'PENDENTE' else 1, (r['cidade'] or '').lower()))

    total = len(registros)
    total_pages = max(1, (total + PER_PAGE - 1) // PER_PAGE)
    page = min(page, total_pages)
    page_slice = registros[(page - 1) * PER_PAGE: page * PER_PAGE]
    page_ids = [r['id'] for r in page_slice]

    realizados = sum(1 for r in registros if r['status'] == 'REALIZADO')
    pendentes = total - realizados
    custo_total = sum(float(r['custo']) for r in registros if r['custo'])

    # Carrega APENAS os objetos completos da página, preservando a ordem.
    objs = {}
    if page_ids:
        objs = {l.id: l for l in IdentidadeVisualLocal.query.filter(
            IdentidadeVisualLocal.id.in_(page_ids)).all()}
    locais = [objs[i] for i in page_ids if i in objs]
    status_por_local = {r['id']: r['status'] for r in page_slice}
    qtd_arquivos = {i: contagem_arquivos.get(i, 0) for i in page_ids}

    # ── Contagens globais para os badges dos filtros (só locais físicos) ──
    rows = db.session.query(
        IdentidadeVisualLocal.id,
        IdentidadeVisualLocal.cidade,
        IdentidadeVisualLocal.tipo_local,
        IdentidadeVisualLocal.bairro,
        IdentidadeVisualLocal.data_acao,
    ).filter(IdentidadeVisualLocal.tipo_local != TIPO_LOCAL_VEICULO).all()

    contagem_cidades, contagem_tipos, contagem_bairros = {}, {}, {}
    contagem_status = {'PENDENTE': 0, 'REALIZADO': 0}
    for rid, rcidade, rtipo, rbairro, rdata in rows:
        if rcidade:
            contagem_cidades[rcidade] = contagem_cidades.get(rcidade, 0) + 1
        if rtipo:
            contagem_tipos[rtipo] = contagem_tipos.get(rtipo, 0) + 1
        if rbairro:
            contagem_bairros[rbairro] = contagem_bairros.get(rbairro, 0) + 1
        contagem_status[_status(rid, rdata)] += 1

    cidades = sorted(contagem_cidades.keys())
    tipos_local_existentes = sorted(contagem_tipos.keys())
    bairros = sorted(contagem_bairros.keys())

    # ── KPIs GERAIS + detalhamento por Tipo Local (TODOS os registros) ──
    # Cards gerais acima das abas; clique expande a quebra por tipo. Query leve
    # (colunas não-deferred) → segura antes do ALTER (sem veículos, só físicos).
    todos_kpi = db.session.query(
        IdentidadeVisualLocal.id,
        IdentidadeVisualLocal.tipo_local,
        IdentidadeVisualLocal.custo,
        IdentidadeVisualLocal.data_acao,
    ).all()

    tipos_presentes = []
    for _kid, ktipo, _kc, _kd in todos_kpi:
        if ktipo and ktipo not in tipos_presentes:
            tipos_presentes.append(ktipo)
    # Ordem: os tipos conhecidos (Veículo, Espaço, Sala) e depois eventuais legados.
    kpi_tipos = list(TIPOS_LOCAL) + [t for t in tipos_presentes if t not in TIPOS_LOCAL]

    total_por_tipo = {t: 0 for t in kpi_tipos}
    realizados_por_tipo = {t: 0 for t in kpi_tipos}
    pendentes_por_tipo = {t: 0 for t in kpi_tipos}
    custo_por_tipo = {t: 0.0 for t in kpi_tipos}
    geral_total = 0
    geral_realizados = 0
    geral_custo = 0.0
    for kid, ktipo, kcusto, kdata in todos_kpi:
        if not ktipo:
            continue
        geral_total += 1
        total_por_tipo[ktipo] += 1
        if _status(kid, kdata) == 'REALIZADO':
            geral_realizados += 1
            realizados_por_tipo[ktipo] += 1
        else:
            pendentes_por_tipo[ktipo] += 1
        if kcusto:
            geral_custo += float(kcusto)
            custo_por_tipo[ktipo] += float(kcusto)
    geral_pendentes = geral_total - geral_realizados

    if 'municipios_pi' in sa_inspect(db.engine).get_table_names():
        municipios = MunicipioPiaui.query.order_by(MunicipioPiaui.nome).all()
    else:
        municipios = []

    # ── Aba "Veículos": conjunto separado, estilo dashboard de Diárias ──
    # Filtros próprios da aba (prefixo v_ p/ não colidir com os de Fachadas):
    # Tipo do Veículo, Cor, Status + busca livre (placa/marca).
    v_tipos_sel = [t for t in request.args.getlist('v_tipo') if t]
    v_cores_sel = [c for c in request.args.getlist('v_cor') if c]
    v_status_sel = [s for s in request.args.getlist('v_status') if s]
    v_busca = (request.args.get('v_q') or '').strip().lower()

    # Sem `undefer`: o SELECT padrão não inclui as colunas de veículo (deferred),
    # então a página continua abrindo antes do ALTER (não há veículos pré-migração,
    # o loop fica vazio e as colunas nunca são acessadas). Frota pequena → o
    # lazy-load por linha aqui é irrelevante.
    todos_veiculos = (IdentidadeVisualLocal.query
                      .filter(IdentidadeVisualLocal.tipo_local == TIPO_LOCAL_VEICULO)
                      .all())
    status_por_veiculo = {v.id: _status(v.id, v.data_acao) for v in todos_veiculos}

    # Opções + contagens dos filtros calculadas sobre TODOS os veículos.
    contagem_tipo_veic, contagem_cor_veic = {}, {}
    contagem_status_veic = {'PENDENTE': 0, 'REALIZADO': 0}
    for v in todos_veiculos:
        if v.tipo_veiculo:
            contagem_tipo_veic[v.tipo_veiculo] = contagem_tipo_veic.get(v.tipo_veiculo, 0) + 1
        if v.cor:
            contagem_cor_veic[v.cor] = contagem_cor_veic.get(v.cor, 0) + 1
        contagem_status_veic[status_por_veiculo[v.id]] += 1

    def _veiculo_passa(v):
        if v_tipos_sel and (v.tipo_veiculo or '') not in v_tipos_sel:
            return False
        if v_cores_sel and (v.cor or '') not in v_cores_sel:
            return False
        if v_status_sel and status_por_veiculo[v.id] not in v_status_sel:
            return False
        if v_busca:
            alvo = ' '.join(x for x in (v.placa, v.marca, v.modelo, v.tipo_veiculo, v.cor) if x).lower()
            if v_busca not in alvo:
                return False
        return True

    # PENDENTES primeiro, depois por placa.
    veiculos = [v for v in todos_veiculos if _veiculo_passa(v)]
    veiculos.sort(key=lambda v: (0 if status_por_veiculo[v.id] == 'PENDENTE' else 1,
                                 (v.placa or '').lower()))
    qtd_arquivos_veiculo = {v.id: contagem_arquivos.get(v.id, 0) for v in veiculos}
    total_veiculos = len(veiculos)
    veic_realizados = sum(1 for v in veiculos if status_por_veiculo[v.id] == 'REALIZADO')
    veic_pendentes = total_veiculos - veic_realizados
    veic_custo_total = sum(float(v.custo) for v in veiculos if v.custo)
    tipos_veic = sorted(contagem_tipo_veic.keys())
    cores_veic = sorted(contagem_cor_veic.keys())

    return render_template(
        'identidade_visual/dashboard.html',
        locais=locais,
        status_por_local=status_por_local,
        qtd_arquivos=qtd_arquivos,
        veiculos=veiculos,
        status_por_veiculo=status_por_veiculo,
        qtd_arquivos_veiculo=qtd_arquivos_veiculo,
        total_veiculos=total_veiculos,
        veic_realizados=veic_realizados,
        veic_pendentes=veic_pendentes,
        veic_custo_total=veic_custo_total,
        tipos_veic=tipos_veic,
        cores_veic=cores_veic,
        contagem_tipo_veic=contagem_tipo_veic,
        contagem_cor_veic=contagem_cor_veic,
        contagem_status_veic=contagem_status_veic,
        filtro_v_tipo=v_tipos_sel,
        filtro_v_cor=v_cores_sel,
        filtro_v_status=v_status_sel,
        filtro_v_q=v_busca,
        aba_ativa=request.args.get('aba', 'fachadas'),
        geral_total=geral_total,
        geral_realizados=geral_realizados,
        geral_pendentes=geral_pendentes,
        geral_custo=geral_custo,
        kpi_tipos=kpi_tipos,
        total_por_tipo=total_por_tipo,
        realizados_por_tipo=realizados_por_tipo,
        pendentes_por_tipo=pendentes_por_tipo,
        custo_por_tipo=custo_por_tipo,
        cidades=cidades,
        tipos_local=tipos_local_existentes,
        tipos_local_opcoes=TIPOS_LOCAL,
        bairros=bairros,
        tipo_veiculo_nome=TIPO_LOCAL_VEICULO,
        contagem_cidades=contagem_cidades,
        contagem_tipos=contagem_tipos,
        contagem_bairros=contagem_bairros,
        contagem_status=contagem_status,
        filtro_cidades=cidades_sel,
        filtro_tipos=tipos_sel,
        filtro_bairros=bairros_sel,
        filtro_status_sel=status_sel,
        page=page,
        total_pages=total_pages,
        total=total,
        realizados=realizados,
        pendentes=pendentes,
        custo_total=custo_total,
        municipios=municipios,
        pode_criar=current_user.tem_permissao('identidade_visual', 'criar'),
        pode_editar=current_user.tem_permissao('identidade_visual', 'editar'),
        pode_excluir=current_user.tem_permissao('identidade_visual', 'excluir'),
        pode_gerar_relatorio_fotografico=current_user.is_admin,
    )


@identidade_visual_bp.route('/relatorio-fotografico')
@requires_admin
def relatorio_fotografico():
    """Exibe a versão imprimível dos registros com status REALIZADO.

    As imagens são servidas diretamente da pasta de uploads da aplicação no
    servidor. Dessa forma, não há extração manual de arquivos para montar o
    relatório. PDFs e HEIC permanecem disponíveis no cadastro, mas não entram
    no layout fotográfico por não serem renderizáveis no navegador.
    """
    locais = (IdentidadeVisualLocal.query
              .join(IdentidadeVisualArquivo)
              .filter(IdentidadeVisualLocal.data_acao.isnot(None))
              .distinct()
              .all())

    local_ids = [local.id for local in locais]
    arquivos_por_local = {local_id: [] for local_id in local_ids}
    if local_ids:
        arquivos = (IdentidadeVisualArquivo.query
                    .filter(IdentidadeVisualArquivo.local_id.in_(local_ids))
                    .order_by(
                        IdentidadeVisualArquivo.local_id,
                        IdentidadeVisualArquivo.created_at,
                        IdentidadeVisualArquivo.id,
                    )
                    .all())
        extensoes_imagem = {'jpg', 'jpeg', 'png', 'webp'}
        for arquivo in arquivos:
            if arquivo.tipo.lower() not in extensoes_imagem:
                continue
            arquivos_por_local[arquivo.local_id].append({
                'nome': arquivo.nome_original,
                'url': url_for(
                    'static',
                    filename=f'uploads/identidade_visual/{arquivo.nome_servidor}',
                ),
            })

    registros = []
    for local in locais:
        if local.is_veiculo:
            identificacao = {
                'categoria': 'Veículo',
                'titulo': local.placa or 'Veículo sem placa',
                'detalhes': [
                    ('Tipo', local.tipo_veiculo),
                    ('Marca', local.marca),
                    ('Modelo', local.modelo),
                    ('Cor', local.cor),
                ],
            }
        else:
            identificacao = {
                'categoria': local.tipo_local,
                'titulo': local.cidade or 'Local sem cidade',
                'detalhes': [
                    ('Endereço', local.endereco),
                    ('Bairro', local.bairro),
                    ('CEP', local.cep),
                ],
            }

        registros.append({
            **identificacao,
            'data_acao': local.data_acao,
            'fotos': arquivos_por_local[local.id],
        })

    registros.sort(key=lambda registro: (
        registro['categoria'] == TIPO_LOCAL_VEICULO,
        registro['titulo'].lower(),
    ))
    return render_template(
        'identidade_visual/relatorio_fotografico.html',
        registros=registros,
        data_geracao=datetime.now(),
    )


@identidade_visual_bp.route('/relatorio-fotografico/pdf')
@requires_admin
def relatorio_fotografico_pdf():
    """Gera o Relatório Fotográfico em PDF para download direto.

    Usa fpdf2 para renderizar o relatório server-side, sem depender do
    Ctrl+P do navegador. As fotos são lidas diretamente do disco do
    servidor (UPLOAD_FOLDER).
    """
    from fpdf import FPDF
    from PIL import Image as PILImage

    # ── Cores (mesmas do template HTML) ──
    AZUL = (7, 90, 155)
    CINZA = (93, 102, 112)
    PRETO = (32, 37, 42)
    COR_BORDA = (220, 226, 232)

    # ── Dimensões A4 ──
    PG_W = 210
    MARGEM = 15
    LARG = PG_W - 2 * MARGEM       # 180 mm úteis
    GAP_COL = 8                     # espaço entre colunas
    COL_W = (LARG - GAP_COL) / 2   # ~86 mm por foto
    MAX_IMG_H = 105                 # altura máxima por foto (mm)
    Y_MAX = 297 - 20                # limite inferior antes do footer

    brasao_path = os.path.join(current_app.root_path, 'static', 'img',
                               'brasao_piaui.png')

    # Fonte core do fpdf2 escreve em latin-1: qualquer caractere fora dessa
    # tabela (travessao, aspas curvas, emoji colados no cadastro) levantaria
    # UnicodeEncodeError e derrubaria o relatorio inteiro.
    SUBSTITUTOS = {
        '–': '-', '—': '-', '‐': '-', '‑': '-',
        '‘': "'", '’': "'", '“': '"', '”': '"',
        '…': '...', ' ': ' ', '•': '-',
    }

    def _txt(valor):
        """Devolve o texto seguro para a fonte core do PDF."""
        if valor is None:
            return ''
        texto = str(valor)
        for origem, destino in SUBSTITUTOS.items():
            texto = texto.replace(origem, destino)
        return texto.encode('latin-1', 'replace').decode('latin-1')

    # ── Subclasse com cabeçalho/rodapé ──
    class _PDF(FPDF):
        def header(self):
            if self.page_no() == 1:
                if os.path.isfile(brasao_path):
                    self.image(brasao_path, x=(PG_W - 22) / 2, y=12, w=22)
                self.set_y(37)
                self.set_font('Helvetica', '', 8)
                self.set_text_color(*CINZA)
                self.cell(LARG, 4,
                          'ESTADO DO PIAU\u00cd - SECRETARIA DE ADMINISTRA\u00c7\u00c3O (SEAD)',
                          align='C', new_x='LMARGIN', new_y='NEXT')
                self.set_font('Helvetica', 'B', 20)
                self.set_text_color(*AZUL)
                self.cell(LARG, 12, 'Relat\u00f3rio Fotogr\u00e1fico',
                          align='C', new_x='LMARGIN', new_y='NEXT')
                self.set_font('Helvetica', '', 9)
                self.set_text_color(*CINZA)
                self.cell(LARG, 5,
                          'Identidade Visual - registros com status realizado',
                          align='C', new_x='LMARGIN', new_y='NEXT')
                y = self.get_y() + 3
                self.set_draw_color(*AZUL)
                self.set_line_width(0.8)
                self.line(MARGEM, y, PG_W - MARGEM, y)
                self.set_y(y + 6)
            else:
                self.set_font('Helvetica', 'B', 9)
                self.set_text_color(*AZUL)
                self.cell(LARG, 6,
                          'Relat\u00f3rio Fotogr\u00e1fico - Identidade Visual',
                          new_x='LMARGIN', new_y='NEXT')
                self.set_draw_color(*AZUL)
                self.set_line_width(0.3)
                self.line(MARGEM, self.get_y(), PG_W - MARGEM, self.get_y())
                self.ln(4)

        def footer(self):
            self.set_y(-12)
            self.set_font('Helvetica', '', 8)
            self.set_text_color(*CINZA)
            self.cell(0, 10, f'P\u00e1gina {self.page_no()}/{{nb}}', align='C')

    # ── Buscar registros realizados (mesma query do HTML) ──
    locais = (IdentidadeVisualLocal.query
              .join(IdentidadeVisualArquivo)
              .filter(IdentidadeVisualLocal.data_acao.isnot(None))
              .distinct()
              .all())

    local_ids = [local.id for local in locais]
    arquivos_por_local = {lid: [] for lid in local_ids}
    if local_ids:
        arquivos = (IdentidadeVisualArquivo.query
                    .filter(IdentidadeVisualArquivo.local_id.in_(local_ids))
                    .order_by(IdentidadeVisualArquivo.local_id,
                              IdentidadeVisualArquivo.created_at,
                              IdentidadeVisualArquivo.id)
                    .all())
        ext_imagem = {'jpg', 'jpeg', 'png', 'webp'}
        for arq in arquivos:
            if arq.tipo.lower() not in ext_imagem:
                continue
            caminho = os.path.join(UPLOAD_FOLDER, arq.nome_servidor)
            if os.path.isfile(caminho):
                arquivos_por_local[arq.local_id].append({
                    'nome': arq.nome_original,
                    'caminho': caminho,
                })

    registros = []
    for local in locais:
        if local.is_veiculo:
            ident = {
                'categoria': 'Ve\u00edculo',
                'titulo': local.placa or 'Ve\u00edculo sem placa',
                'detalhes': [
                    ('Tipo', local.tipo_veiculo), ('Marca', local.marca),
                    ('Modelo', local.modelo), ('Cor', local.cor),
                ],
            }
        else:
            ident = {
                'categoria': local.tipo_local,
                'titulo': local.cidade or 'Local sem cidade',
                'detalhes': [
                    ('Endere\u00e7o', local.endereco), ('Bairro', local.bairro),
                    ('CEP', local.cep),
                ],
            }
        registros.append({
            **ident,
            'data_acao': local.data_acao,
            'fotos': arquivos_por_local[local.id],
        })

    registros.sort(key=lambda r: (
        r['categoria'] == TIPO_LOCAL_VEICULO,
        r['titulo'].lower(),
    ))

    # ── Montar PDF ──
    pdf = _PDF('P', 'mm', 'A4')
    pdf.set_margins(MARGEM, MARGEM, MARGEM)
    pdf.set_auto_page_break(auto=False)
    pdf.alias_nb_pages()
    pdf.add_page()

    agora = datetime.now()

    # Resumo
    pdf.set_fill_color(237, 245, 250)
    pdf.set_font('Helvetica', '', 10)
    pdf.set_text_color(*PRETO)
    y0 = pdf.get_y()
    pdf.rect(MARGEM, y0, LARG, 8, style='F')
    pdf.set_draw_color(*AZUL)
    pdf.set_line_width(0.8)
    pdf.line(MARGEM, y0, MARGEM, y0 + 8)
    pdf.set_xy(MARGEM + 4, y0 + 1)
    # Formato fora da f-string: producao roda Python 3.10, que nao
    # aceita backslash dentro da expressao de uma f-string.
    fmt_data = '%d/%m/%Y \u00e0s %H:%M'
    resumo = (f'{len(registros)} registro(s) realizado(s), gerado em '
              f'{agora.strftime(fmt_data)}.')
    pdf.cell(LARG - 4, 6, resumo)
    pdf.set_y(y0 + 14)

    if not registros:
        pdf.set_font('Helvetica', '', 12)
        pdf.set_text_color(*CINZA)
        pdf.cell(LARG, 30,
                 'N\u00e3o h\u00e1 registros realizados para exibir.', align='C')

    def _img_size(caminho_img):
        """Retorna (w, h) em mm que cabem em COL_W x MAX_IMG_H."""
        try:
            with PILImage.open(caminho_img) as img:
                w_px, h_px = img.size
        except Exception:
            return 0, 0
        aspect = h_px / w_px
        w, h = COL_W, COL_W * aspect
        if h > MAX_IMG_H:
            h = MAX_IMG_H
            w = h / aspect
        return w, h

    for idx, reg in enumerate(registros):
        if idx > 0:
            pdf.add_page()

        # Categoria
        pdf.set_font('Helvetica', 'B', 9)
        pdf.set_text_color(*AZUL)
        pdf.cell(LARG, 5, _txt(reg['categoria']).upper(),
                 new_x='LMARGIN', new_y='NEXT')

        # Título + Data
        pdf.set_font('Helvetica', 'B', 16)
        pdf.set_text_color(*PRETO)
        pdf.cell(LARG * 0.6, 9, _txt(reg['titulo']))

        data_str = ''
        if reg['data_acao']:
            data_str = ('Realizado em '
                        + reg['data_acao'].strftime(fmt_data))
        pdf.set_font('Helvetica', '', 9)
        pdf.set_text_color(*CINZA)
        pdf.cell(LARG * 0.4, 9, data_str, align='R',
                 new_x='LMARGIN', new_y='NEXT')

        # Detalhes
        detalhes = [f'{_txt(r)}: {_txt(v)}' for r, v in reg['detalhes'] if v]
        if detalhes:
            pdf.set_font('Helvetica', '', 10)
            pdf.set_text_color(*CINZA)
            pdf.cell(LARG, 6, '  |  '.join(detalhes),
                     new_x='LMARGIN', new_y='NEXT')

        # Separador
        pdf.ln(3)
        pdf.set_draw_color(*COR_BORDA)
        pdf.set_line_width(0.3)
        pdf.line(MARGEM, pdf.get_y(), PG_W - MARGEM, pdf.get_y())
        pdf.ln(5)

        # ── Fotos em grid 2 colunas ──
        fotos = reg.get('fotos', [])
        if not fotos:
            pdf.set_font('Helvetica', 'I', 10)
            pdf.set_text_color(109, 77, 0)
            pdf.cell(LARG, 8,
                     'Sem fotos em formato compat\u00edvel para o relat\u00f3rio.',
                     new_x='LMARGIN', new_y='NEXT')
            continue

        i = 0
        while i < len(fotos):
            par = fotos[i:i + 2]
            tamanhos = [_img_size(f['caminho']) for f in par]
            alturas = [h for _, h in tamanhos if h > 0]
            row_h = max(alturas) if alturas else 0

            # Bloco = foto + legenda (~7mm extra)
            if pdf.get_y() + row_h + 7 > Y_MAX:
                pdf.add_page()

            y_row = pdf.get_y()

            for j, foto in enumerate(par):
                w_img, h_img = tamanhos[j]
                if w_img == 0:
                    continue
                x = MARGEM + j * (COL_W + GAP_COL)
                x_offset = (COL_W - w_img) / 2 if w_img < COL_W else 0

                # Borda fina ao redor da foto
                pdf.set_draw_color(*COR_BORDA)
                pdf.set_line_width(0.2)
                pdf.rect(x + x_offset - 0.5, y_row - 0.5,
                         w_img + 1, h_img + 1)

                try:
                    pdf.image(foto['caminho'],
                              x=x + x_offset, y=y_row,
                              w=w_img, h=h_img)
                except Exception:
                    current_app.logger.warning(
                        '[IV-PDF] Erro ao inserir imagem: %s',
                        foto['caminho'])

                # Legenda (nome do arquivo)
                pdf.set_xy(x, y_row + h_img + 1)
                pdf.set_font('Helvetica', '', 7)
                pdf.set_text_color(*CINZA)
                nome_leg = _txt(foto['nome'])
                if len(nome_leg) > 40:
                    nome_leg = nome_leg[:37] + '...'
                pdf.cell(COL_W, 4, nome_leg)

            pdf.set_y(y_row + row_h + 8)
            i += 2

    # ── Retornar PDF como download ──
    buf = io.BytesIO()
    pdf.output(buf)
    buf.seek(0)

    nome_arquivo = f'Relatorio_Fotografico_{agora.strftime("%Y%m%d_%H%M")}.pdf'
    return send_file(
        buf,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=nome_arquivo,
    )

@identidade_visual_bp.route('/api/salvar-acao/<int:local_id>', methods=['POST'])
@requires_permission('identidade_visual.editar')
def salvar_acao(local_id):
    local = IdentidadeVisualLocal.query.get(local_id)
    if not local:
        return jsonify({'erro': 'Local não encontrado'}), 404

    data_acao = request.form.get('data_acao', '').strip()
    if not data_acao:
        return jsonify({'erro': 'Data e hora são obrigatórios'}), 400
    try:
        local.data_acao = datetime.strptime(data_acao, '%Y-%m-%dT%H:%M')
    except ValueError:
        return jsonify({'erro': 'Formato de data inválido'}), 400

    arquivos = request.files.getlist('arquivos')
    tem_arquivo_novo = any(a and a.filename for a in arquivos)
    tem_arquivo_existente = local.arquivos.count() > 0

    if not tem_arquivo_novo and not tem_arquivo_existente:
        return jsonify({'erro': 'É necessário anexar pelo menos um arquivo'}), 400

    custo_raw = request.form.get('custo', '').strip()
    if custo_raw:
        custo_raw = custo_raw.replace('R$', '').replace('.', '').replace(',', '.').strip()
        try:
            local.custo = float(custo_raw)
        except ValueError:
            return jsonify({'erro': 'Valor de custo inválido'}), 400
    else:
        local.custo = None

    arquivos_novos = [a for a in arquivos if a and a.filename]

    # Valida TODOS os arquivos antes de gravar qualquer um — evita deixar
    # arquivos órfãos no disco se um arquivo posterior for rejeitado.
    for arquivo in arquivos_novos:
        if not _allowed_file(arquivo.filename):
            return jsonify({'erro': f'Tipo de arquivo não permitido: {arquivo.filename}'}), 400
        # Mede o tamanho sem carregar o arquivo inteiro na memória
        arquivo.seek(0, os.SEEK_END)
        tamanho = arquivo.tell()
        arquivo.seek(0)
        if tamanho > MAX_FILE_SIZE:
            limite_mb = MAX_FILE_SIZE // (1024 * 1024)
            return jsonify({'erro': f'O arquivo "{arquivo.filename}" tem '
                                    f'{tamanho / (1024 * 1024):.1f} MB e excede o '
                                    f'limite de {limite_mb} MB por arquivo.'}), 400

    os.makedirs(UPLOAD_FOLDER, exist_ok=True)
    for arquivo in arquivos_novos:
        # Extensão vem do nome ORIGINAL (já validado em _allowed_file).
        # secure_filename pode descartar nomes não-ASCII e devolver algo sem
        # extensão (ou vazio) — por isso não derivamos a extensão dele.
        extensao = arquivo.filename.rsplit('.', 1)[1].lower()
        base_segura = secure_filename(arquivo.filename)
        if not base_segura or '.' not in base_segura:
            base_segura = f'foto.{extensao}'
        nome_unico = f"{uuid.uuid4().hex[:8]}_{base_segura}"
        caminho = os.path.join(UPLOAD_FOLDER, nome_unico)
        arquivo.save(caminho)

        novo_arquivo = IdentidadeVisualArquivo(
            local_id=local.id,
            nome_original=arquivo.filename,
            nome_servidor=nome_unico,
            tipo=extensao,
        )
        db.session.add(novo_arquivo)

    local.atualizado_por_id = getattr(current_user, 'id', None)
    _registrar_log('EDITAR', f'Registrou ação/anexos em {_rotulo(local)} — {local.tipo_local}', local_id=local.id)

    db.session.commit()
    return jsonify({'ok': True, 'status': local.status})


@identidade_visual_bp.route('/api/remover-arquivo/<int:arquivo_id>', methods=['POST'])
@requires_permission('identidade_visual.editar')
def remover_arquivo(arquivo_id):
    arquivo = IdentidadeVisualArquivo.query.get(arquivo_id)
    if not arquivo:
        return jsonify({'erro': 'Arquivo não encontrado'}), 404

    caminho = os.path.join(UPLOAD_FOLDER, arquivo.nome_servidor)
    if os.path.exists(caminho):
        os.remove(caminho)

    local = arquivo.local_ref
    db.session.delete(arquivo)
    db.session.commit()
    return jsonify({'ok': True, 'status': local.status})


@identidade_visual_bp.route('/api/remover-custo/<int:local_id>', methods=['POST'])
@requires_permission('identidade_visual.editar')
def remover_custo(local_id):
    local = IdentidadeVisualLocal.query.get(local_id)
    if not local:
        return jsonify({'erro': 'Local não encontrado'}), 404

    local.custo = None
    db.session.commit()
    return jsonify({'ok': True, 'status': local.status})


@identidade_visual_bp.route('/api/arquivos/<int:local_id>')
@requires_permission('identidade_visual.visualizar')
def listar_arquivos(local_id):
    local = IdentidadeVisualLocal.query.get(local_id)
    if not local:
        return jsonify({'erro': 'Local não encontrado'}), 404

    arquivos = [{
        'id': a.id,
        'nome': a.nome_original,
        'tipo': a.tipo,
        'url': f'/static/uploads/identidade_visual/{a.nome_servidor}',
        'data': a.created_at.strftime('%d/%m/%Y %H:%M') if a.created_at else '',
    } for a in local.arquivos.all()]

    try:
        mid = local.municipio_id
    except Exception:
        mid = None

    return jsonify({
        'arquivos': arquivos,
        'custo': str(local.custo) if local.custo else None,
        'data_acao': local.data_acao.strftime('%Y-%m-%dT%H:%M') if local.data_acao else None,
        'cidade': local.cidade,
        'municipio_id': mid,
        'tipo_local': local.tipo_local,
        'endereco': local.endereco or '',
        'bairro': local.bairro or '',
        'cep': local.cep or '',
        'is_veiculo': local.is_veiculo,
        'placa': local.placa or '',
        'tipo_veiculo': local.tipo_veiculo or '',
        'marca': local.marca or '',
        'modelo': local.modelo or '',
        'cor': local.cor or '',
    })


def _validar_local(data):
    """Valida campos de criação/edição. Retorna (erro_msg, status_code) ou None.

    A validação diverge por tipo: 'Veículo' exige só a placa; os demais tipos
    (locais físicos) exigem município, endereço, bairro e CEP.
    """
    tipo_local = (data.get('tipo_local') or '').strip()
    if not tipo_local:
        return 'Tipo Local é obrigatório', 400
    if tipo_local not in TIPOS_LOCAL:
        return 'Tipo Local inválido', 400

    if tipo_local == TIPO_LOCAL_VEICULO:
        if len(normalizar_placa(data.get('placa'))) != 7:
            return 'Placa é obrigatória (7 caracteres)', 400
        return None

    # Local físico
    if not data.get('municipio_id'):
        return 'Cidade é obrigatória', 400
    if not (data.get('endereco') or '').strip():
        return 'Endereço é obrigatório', 400
    if not (data.get('bairro') or '').strip():
        return 'Bairro é obrigatório', 400
    if not (data.get('cep') or '').strip():
        return 'CEP é obrigatório', 400
    return None


@identidade_visual_bp.route('/api/consultar-placa')
@requires_permission('identidade_visual.criar')
def consultar_placa():
    """Proxy server-side da consulta DETRAN por placa.

    Mantém a `X-Api-Key` no servidor e devolve só os campos públicos
    (tipo/marca-modelo/cor) — nunca os dados do proprietário retornados pela API.
    """
    try:
        info = consultar_veiculo(request.args.get('placa', ''))
    except DetranError as e:
        return jsonify({'erro': str(e)}), 400
    except Exception:
        current_app.logger.exception('[IDENTIDADE_VISUAL] Falha ao consultar placa no DETRAN')
        return jsonify({'erro': 'Falha ao consultar a placa. Tente novamente.'}), 502

    if not info.get('marca') and not info.get('modelo') and not info.get('tipo_veiculo'):
        return jsonify({'erro': 'Veículo não encontrado para essa placa'}), 404
    return jsonify({'ok': True, **info})


@identidade_visual_bp.route('/api/criar-local', methods=['POST'])
@requires_permission('identidade_visual.criar')
def criar_local():
    data = request.get_json() or {}

    erro = _validar_local(data)
    if erro:
        return jsonify({'erro': erro[0]}), erro[1]

    uid = getattr(current_user, 'id', None)
    tipo_local = data['tipo_local'].strip()

    if tipo_local == TIPO_LOCAL_VEICULO:
        # Dados do veículo vêm do DETRAN (fonte confiável) — não confiamos no
        # que o cliente enviou como marca/cor, só na placa.
        try:
            info = consultar_veiculo(data.get('placa'))
        except DetranError as e:
            return jsonify({'erro': str(e)}), 400
        except Exception:
            current_app.logger.exception('[IDENTIDADE_VISUAL] Falha ao consultar placa no DETRAN')
            return jsonify({'erro': 'Falha ao consultar a placa. Tente novamente.'}), 502

        local = IdentidadeVisualLocal(
            tipo_local=tipo_local,
            placa=info['placa'],
            tipo_veiculo=info.get('tipo_veiculo'),
            marca=info.get('marca'),
            modelo=info.get('modelo'),
            cor=info.get('cor'),
            criado_por_id=uid,
            atualizado_por_id=uid,
        )
        db.session.add(local)
        db.session.flush()
        _registrar_log('CRIAR',
                       f'Criou o veículo {local.placa} — {_marca_modelo(local)}'.strip(' —'),
                       local_id=local.id)
        db.session.commit()
        return jsonify({'ok': True, 'id': local.id})

    municipio = db.session.get(MunicipioPiaui, data['municipio_id'])
    if not municipio:
        return jsonify({'erro': 'Município não encontrado'}), 400

    local = IdentidadeVisualLocal(
        cidade=municipio.nome,
        municipio_id=municipio.id,
        tipo_local=tipo_local,
        endereco=data['endereco'].strip(),
        bairro=data['bairro'].strip(),
        cep=data['cep'].strip(),
        criado_por_id=uid,
        atualizado_por_id=uid,
    )
    db.session.add(local)
    db.session.flush()
    _registrar_log('CRIAR', f'Criou o local {local.cidade} — {local.tipo_local}', local_id=local.id)
    db.session.commit()

    return jsonify({'ok': True, 'id': local.id})


@identidade_visual_bp.route('/api/editar-local/<int:local_id>', methods=['POST'])
@requires_permission('identidade_visual.editar')
def editar_local(local_id):
    local = IdentidadeVisualLocal.query.get(local_id)
    if not local:
        return jsonify({'erro': 'Local não encontrado'}), 404

    data = request.get_json() or {}

    erro = _validar_local(data)
    if erro:
        return jsonify({'erro': erro[0]}), erro[1]

    tipo_local = data['tipo_local'].strip()

    if tipo_local == TIPO_LOCAL_VEICULO:
        try:
            info = consultar_veiculo(data.get('placa'))
        except DetranError as e:
            return jsonify({'erro': str(e)}), 400
        except Exception:
            current_app.logger.exception('[IDENTIDADE_VISUAL] Falha ao consultar placa no DETRAN')
            return jsonify({'erro': 'Falha ao consultar a placa. Tente novamente.'}), 502

        local.tipo_local = tipo_local
        local.placa = info['placa']
        local.tipo_veiculo = info.get('tipo_veiculo')
        local.marca = info.get('marca')
        local.modelo = info.get('modelo')
        local.cor = info.get('cor')
        # Zera campos de local físico — o registro deixou de ser um endereço.
        local.cidade = None
        local.municipio_id = None
        local.endereco = None
        local.bairro = None
        local.cep = None
        local.atualizado_por_id = getattr(current_user, 'id', None)
        _registrar_log('EDITAR',
                       f'Editou o veículo {local.placa} — {_marca_modelo(local)}'.strip(' —'),
                       local_id=local.id)
        db.session.commit()
        return jsonify({'ok': True, 'id': local.id})

    municipio = db.session.get(MunicipioPiaui, data['municipio_id'])
    if not municipio:
        return jsonify({'erro': 'Município não encontrado'}), 400

    local.cidade = municipio.nome
    local.municipio_id = municipio.id
    local.tipo_local = tipo_local
    local.endereco = data['endereco'].strip()
    local.bairro = data['bairro'].strip()
    local.cep = data['cep'].strip()
    # Limpa campos de veículo caso o registro tenha sido convertido de volta.
    local.placa = None
    local.tipo_veiculo = None
    local.marca = None
    local.modelo = None
    local.cor = None
    local.atualizado_por_id = getattr(current_user, 'id', None)

    _registrar_log('EDITAR', f'Editou o local {local.cidade} — {local.tipo_local}', local_id=local.id)
    db.session.commit()

    return jsonify({'ok': True, 'id': local.id})


@identidade_visual_bp.route('/api/excluir-local/<int:local_id>', methods=['POST'])
@requires_permission('identidade_visual.excluir')
def excluir_local(local_id):
    """Exclui um local. Restrito a usuários com acesso FULL ao módulo
    (permissão `identidade_visual.excluir` ou is_admin). A ação é auditada."""
    local = IdentidadeVisualLocal.query.get(local_id)
    if not local:
        return jsonify({'erro': 'Local não encontrado'}), 404

    # Remove arquivos físicos antes de apagar os registros filhos (cascade)
    for arquivo in local.arquivos.all():
        caminho = os.path.join(UPLOAD_FOLDER, arquivo.nome_servidor)
        if os.path.exists(caminho):
            try:
                os.remove(caminho)
            except OSError:
                current_app.logger.warning(
                    f'[IDENTIDADE_VISUAL] Falha ao remover arquivo {caminho}')

    descricao = f'Excluiu o local {_rotulo(local)} — {local.tipo_local} (#{local.id})'
    _registrar_log('EXCLUIR', descricao, local_id=local.id)

    db.session.delete(local)
    db.session.commit()
    current_app.logger.info(f'[IDENTIDADE_VISUAL] {descricao} por '
                            f'{getattr(current_user, "nome", "?")}')
    return jsonify({'ok': True})


@identidade_visual_bp.route('/exportar-excel')
@requires_permission('identidade_visual.visualizar')
def exportar_excel():
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    locais = IdentidadeVisualLocal.query.order_by(IdentidadeVisualLocal.cidade).all()

    # Quantidade de arquivos por local em UMA query — evita N+1 (status + count).
    contagem_arquivos = dict(
        db.session.query(
            IdentidadeVisualArquivo.local_id,
            func.count(IdentidadeVisualArquivo.id),
        ).group_by(IdentidadeVisualArquivo.local_id).all()
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Identidade Visual'

    headers = ['Cidade', 'Tipo Local', 'Endereço', 'Bairro', 'CEP', 'Placa',
               'Marca', 'Modelo', 'Cor', 'Status', 'Custo (R$)', 'Data/Hora', 'Qtd Arquivos']
    header_fill = PatternFill(start_color='0891B2', end_color='0891B2', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    for row_idx, l in enumerate(locais, 2):
        qtd = contagem_arquivos.get(l.id, 0)
        status = 'REALIZADO' if (l.data_acao is not None and qtd > 0) else 'PENDENTE'
        # Só toca nas colunas de veículo (deferred) quando o registro é veículo —
        # evita lazy-load/erro em linhas físicas antes do ALTER em produção.
        eh_veic = l.is_veiculo
        valores = [
            l.cidade or '',
            l.tipo_local,
            l.endereco or '',
            l.bairro or '',
            l.cep or '',
            (l.placa or '') if eh_veic else '',
            (l.marca or '') if eh_veic else '',
            (l.modelo or '') if eh_veic else '',
            (l.cor or '') if eh_veic else '',
            status,
            float(l.custo) if l.custo else None,
            l.data_acao.strftime('%d/%m/%Y %H:%M') if l.data_acao else '',
            qtd,
        ]
        for col, v in enumerate(valores, 1):
            cell = ws.cell(row=row_idx, column=col, value=v)
            cell.border = thin_border
            if col == 11 and v is not None:
                cell.number_format = '#,##0.00'

    col_widths = [22, 18, 40, 18, 12, 12, 16, 28, 14, 12, 15, 18, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'Relatorio_Fachadas_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx',
    )
