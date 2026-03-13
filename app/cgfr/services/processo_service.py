"""
Servico principal do modulo CGFR.
Logica de negocios para dashboard, classificacao e relatorios.
Espelha a implementacao do sistema original (website/app/services/processo_service.py).
"""
import io
import logging
from datetime import date, datetime
from decimal import Decimal

from sqlalchemy import or_, func

from app.extensions import db
from app.cgfr.models import CgfrProcessoEnviado, Acao
from app.cgfr.repositories.processo_local_repo import ProcessoLocalRepository
from app.models.nat_despesa import NatDespesa
from app.models.class_fonte import ClassFonte

logger = logging.getLogger(__name__)


class ProcessoService:
    """Servico de logica de negocios para processos CGFR."""

    # =========================================================================
    # Dashboard data (KPIs)
    # =========================================================================

    @staticmethod
    def get_dashboard_stats():
        """Retorna KPIs para o dashboard — espelha o original."""
        try:
            contagem = ProcessoLocalRepository.contar_por_status()
            total = contagem['total']
            classificados = contagem['classificados']
            pendentes = contagem['pendentes']

            row = db.session.query(
                func.coalesce(func.sum(CgfrProcessoEnviado.valor_solicitado), 0),
                func.coalesce(func.sum(CgfrProcessoEnviado.valor_aprovado), 0),
            ).first()
            valor_total_solicitado = float(row[0]) if row else 0.0
            valor_total_acordado = float(row[1]) if row else 0.0

            return {
                'total': total,
                'classificados': classificados,
                'pendentes_classificacao': pendentes,
                'valor_total_solicitado': valor_total_solicitado,
                'valor_total_acordado': valor_total_acordado,
                'valor_total_solicitado_fmt': _format_brl(valor_total_solicitado) or 'R$ 0,00',
                'valor_total_acordado_fmt': _format_brl(valor_total_acordado) or 'R$ 0,00',
            }
        except Exception as e:
            logger.error(f"Erro ao buscar estatisticas: {e}", exc_info=True)
            return {
                'total': 0, 'classificados': 0, 'pendentes_classificacao': 0,
                'valor_total_solicitado': 0, 'valor_total_acordado': 0,
                'valor_total_solicitado_fmt': 'R$ 0,00',
                'valor_total_acordado_fmt': 'R$ 0,00',
            }

    # =========================================================================
    # Listagem para DataTable CLIENT-SIDE (POST /api/data) — espelha original
    # =========================================================================

    @staticmethod
    def get_processos_paginados(search='', status_filter='all',
                                show_only_edited=False):
        """Retorna TODOS os processos formatados para DataTable client-side.
        Espelha website/app/services/processo_service.py::get_processos_paginados.
        """
        try:
            query = CgfrProcessoEnviado.query

            if search:
                term = f'%{search}%'
                query = query.filter(
                    or_(
                        CgfrProcessoEnviado.processo_formatado.ilike(term),
                        CgfrProcessoEnviado.especificacao.ilike(term),
                        CgfrProcessoEnviado.tipo_processo.ilike(term),
                        CgfrProcessoEnviado.fornecedor.ilike(term),
                    )
                )

            if status_filter == 'classified':
                query = query.filter(
                    CgfrProcessoEnviado.natureza_despesa_id.isnot(None),
                    CgfrProcessoEnviado.fonte_id.isnot(None),
                    CgfrProcessoEnviado.acao_id.isnot(None),
                )
            elif status_filter == 'pending':
                query = query.filter(
                    or_(
                        CgfrProcessoEnviado.natureza_despesa_id.is_(None),
                        CgfrProcessoEnviado.fonte_id.is_(None),
                        CgfrProcessoEnviado.acao_id.is_(None),
                    )
                )

            if show_only_edited:
                query = query.filter(
                    CgfrProcessoEnviado.natureza_despesa_id.isnot(None),
                    CgfrProcessoEnviado.fonte_id.isnot(None),
                    CgfrProcessoEnviado.acao_id.isnot(None),
                )

            # Ordena: registros com dados SEI primeiro, depois por data_inclusao desc
            processos = query.order_by(
                db.case(
                    (CgfrProcessoEnviado.objeto_do_pedido.isnot(None), 0),
                    else_=1,
                ),
                CgfrProcessoEnviado.data_inclusao.desc(),
            ).all()

            acao_ids = {p.acao_id for p in processos if p.acao_id}
            acao_map = {}
            if acao_ids:
                acoes = Acao.query.filter(Acao.id.in_(acao_ids)).all()
                acao_map = {a.id: a for a in acoes}

            records = [_format_record(p, acao_map) for p in processos]

            total = len(records)
            kpi_solicitado = sum(_to_float(r.get('valor_solicitado')) for r in records)
            kpi_aprovado = sum(_to_float(r.get('valor_aprovado')) for r in records)
            kpi_pendentes = sum(1 for r in records if not r['is_classified'])

            return {
                'records': records,
                'total': total,
                'kpi_solicitado': kpi_solicitado,
                'kpi_aprovado': kpi_aprovado,
                'kpi_solicitado_fmt': _format_brl(kpi_solicitado) or 'R$ 0,00',
                'kpi_aprovado_fmt': _format_brl(kpi_aprovado) or 'R$ 0,00',
                'kpi_pendentes': kpi_pendentes,
            }

        except Exception as e:
            logger.error(f"Erro ao buscar processos paginados: {e}", exc_info=True)
            return {
                'records': [], 'total': 0,
                'kpi_solicitado': 0, 'kpi_aprovado': 0,
                'kpi_solicitado_fmt': 'R$ 0,00', 'kpi_aprovado_fmt': 'R$ 0,00',
                'kpi_pendentes': 0,
            }

    # =========================================================================
    # Record completo para o modal de edicao
    # =========================================================================

    @staticmethod
    def get_record_completo(protocolo):
        """Busca dados completos de um processo. Retorna dict flat."""
        processo = CgfrProcessoEnviado.query.filter_by(
            processo_formatado=protocolo
        ).first()

        if not processo:
            return None

        acao_map = {}
        if processo.acao_id:
            a = Acao.query.get(processo.acao_id)
            if a:
                acao_map = {a.id: a}

        return _format_record(processo, acao_map)

    # =========================================================================
    # Listagem para DataTable SERVER-SIDE (legado, mantido para compat)
    # =========================================================================

    @staticmethod
    def listar_para_datatable(draw, start, length, filtros=None, search=None):
        """Retorna dados formatados para DataTable server-side."""
        query = ProcessoLocalRepository.listar_com_filtros(filtros, search)
        records_total = CgfrProcessoEnviado.query.count()
        records_filtered = query.count()

        processos = query.offset(start).limit(length).all()

        acao_ids = {p.acao_id for p in processos if p.acao_id}
        acao_map = {}
        if acao_ids:
            acoes = Acao.query.filter(Acao.id.in_(acao_ids)).all()
            acao_map = {a.id: a for a in acoes}

        data = [_format_record(p, acao_map) for p in processos]

        return {
            'draw': draw,
            'recordsTotal': records_total,
            'recordsFiltered': records_filtered,
            'data': data,
        }

    # =========================================================================
    # Salvar classificacao
    # =========================================================================

    @staticmethod
    def classificar_processo(protocolo, dados, usuario_id=None):
        return ProcessoLocalRepository.classificar(protocolo, dados, usuario_id)

    # =========================================================================
    # Opcoes para os selects (Natureza, Fonte, Acao)
    # =========================================================================

    @staticmethod
    def get_filter_options():
        """Retorna opcoes para os dropdowns de filtro (Select2)."""
        naturezas = NatDespesa.query.order_by(NatDespesa.titulo).all()
        fontes = ClassFonte.query.order_by(ClassFonte.descricao).all()
        acoes = Acao.query.order_by(Acao.titulo).all()

        return {
            'naturezas': [{'id': n.id, 'text': f'{n.codigo} - {n.titulo}'} for n in naturezas],
            'fontes': [{'id': f.id, 'text': f'{f.codigo} - {f.descricao}'} for f in fontes],
            'acoes': [{'id': a.id, 'text': f'{a.codigo} - {a.titulo}'} for a in acoes],
            'natdespesas': [{'codigo': n.id, 'id_titulo': f'{n.codigo} - {n.titulo}'} for n in naturezas],
        }

    @staticmethod
    def get_select_options():
        return ProcessoService.get_filter_options()

    # =========================================================================
    # Detalhes
    # =========================================================================

    @staticmethod
    def get_detalhes(protocolo):
        return ProcessoLocalRepository.get_by_protocolo(protocolo)

    # =========================================================================
    # Exportar Excel
    # =========================================================================

    @staticmethod
    def exportar_excel(filtros=None, search=None):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        processos = ProcessoLocalRepository.get_all_for_export(filtros, search)

        acao_ids = {p.acao_id for p in processos if p.acao_id}
        acao_map = {}
        if acao_ids:
            acoes_db = Acao.query.filter(Acao.id.in_(acao_ids)).all()
            acao_map = {a.id: a for a in acoes_db}

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Processos CGFR'

        headers = [
            'Processo', 'Tipo', 'Especificacao', 'Natureza', 'Fonte',
            'Acao', 'Fornecedor', 'Objeto', 'Tipo Despesa',
            'Valor Solicitado', 'Valor Aprovado', 'Prioridade',
            'Status', 'Data Inclusao',
        ]

        header_fill = PatternFill(start_color='DC2626', end_color='DC2626', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
        )

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        for row_idx, p in enumerate(processos, 2):
            values = [
                p.processo_formatado,
                p.tipo_processo or '',
                p.especificacao or '',
                p.natureza_rel.titulo if p.natureza_rel else '',
                p.fonte_rel.descricao if p.fonte_rel else '',
                acao_map[p.acao_id].titulo if p.acao_id and p.acao_id in acao_map else '',
                p.fornecedor or '',
                p.objeto_do_pedido or '',
                p.tipo_despesa or '',
                float(p.valor_solicitado) if p.valor_solicitado else 0,
                float(p.valor_aprovado) if p.valor_aprovado else 0,
                p.nivel_prioridade or '',
                p.status_classificacao,
                p.data_inclusao.strftime('%d/%m/%Y') if p.data_inclusao else '',
            ]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.border = thin_border

        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 2, 40)

        ws.auto_filter.ref = ws.dimensions

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    # =========================================================================
    # Relatorio agrupado
    # =========================================================================

    @staticmethod
    def gerar_relatorio_agrupado(filtros=None, search=None):
        processos = ProcessoLocalRepository.get_all_for_export(filtros, search)

        grupos = {}
        total_solicitado = Decimal('0')
        total_aprovado = Decimal('0')

        for p in processos:
            nat_nome = p.natureza_rel.titulo if p.natureza_rel else 'Sem Classificacao'
            if nat_nome not in grupos:
                grupos[nat_nome] = {
                    'natureza': nat_nome,
                    'processos': [],
                    'subtotal_solicitado': Decimal('0'),
                    'subtotal_aprovado': Decimal('0'),
                }

            grupos[nat_nome]['processos'].append(p)
            if p.valor_solicitado:
                grupos[nat_nome]['subtotal_solicitado'] += p.valor_solicitado
                total_solicitado += p.valor_solicitado
            if p.valor_aprovado:
                grupos[nat_nome]['subtotal_aprovado'] += p.valor_aprovado
                total_aprovado += p.valor_aprovado

        return {
            'grupos': sorted(grupos.values(), key=lambda g: g['natureza']),
            'total_solicitado': total_solicitado,
            'total_aprovado': total_aprovado,
            'total_processos': len(processos),
        }


# =============================================================================
# Funcoes auxiliares (module-level)
# =============================================================================

def _format_record(p, acao_map=None):
    """Converte CgfrProcessoEnviado ORM → dict compativel com frontend JS.
    Espelha website/app/services/processo_service.py::_format_record.
    """
    acao_map = acao_map or {}
    d = p.to_dict()

    # Aliases para compatibilidade com DataTable columns do original
    d['protocolo_formatado'] = d.get('processo_formatado', '')
    d['dt_enviado_fmt'] = d.get('tramitado_sead_cgfr', '') or '-'
    d['valor_acordado_fmt'] = _format_brl_display(d.get('valor_aprovado'))
    d['valor_solicitado_fmt'] = _format_brl_display(d.get('valor_solicitado'))
    d['is_classified'] = p.classificado
    d['classificado'] = p.classificado

    # Natureza/Fonte/Acao como texto para exibicao na tabela
    d['natureza_despesa'] = p.natureza_rel.titulo if p.natureza_rel else None
    d['acao'] = acao_map[p.acao_id].titulo if p.acao_id and p.acao_id in acao_map else None
    d['fonte'] = p.fonte_rel.descricao if p.fonte_rel else None

    # IDs para o modal de edicao
    d['natureza_despesa_id'] = p.natureza_despesa_id
    d['fonte_id'] = p.fonte_id
    d['acao_id'] = p.acao_id

    d.setdefault('link_acesso', None)

    di = p.data_inclusao
    if di:
        d['data_inclusao_fmt'] = di.strftime('%d/%m/%Y %H:%M')
        d['is_novo'] = di > datetime(2026, 1, 1, 0, 0, 0)
    else:
        d['data_inclusao_fmt'] = '-'
        d['is_novo'] = False

    d['data_recebido_cgfr'] = d.get('data_recebido_cgfr', '') or '-'
    d['data_da_reuniao'] = d.get('data_da_reuniao') or '-'

    d['status_cgfr'] = _calcular_status_cgfr(d)
    d['status_cgfr_badge'] = _badge_status_cgfr(d['status_cgfr'])

    return d


def _calcular_status_cgfr(d):
    if d.get('devolvido_cgfr_sead') == 1:
        return 'Devolvido'
    if d.get('recebido_cgfr') == 1:
        return 'Recebido'
    if d.get('tramitado_sead_cgfr'):
        return 'Enviado'
    return 'Nao enviado'


def _badge_status_cgfr(status):
    badges = {
        'Devolvido': 'bg-warning text-dark',
        'Recebido': 'bg-success',
        'Enviado': 'bg-info',
        'Nao enviado': 'bg-secondary',
    }
    return badges.get(status, 'bg-secondary')


def _format_brl(valor):
    if valor is None:
        return ''
    try:
        v = float(valor)
        return f'R$ {v:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')
    except (ValueError, TypeError):
        return ''


def _format_brl_display(valor):
    if valor is None:
        return '-'
    try:
        v = float(valor)
        if v == 0:
            return '-'
        return f'R$ {v:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')
    except (ValueError, TypeError):
        return '-'


def _to_float(value):
    if value is None:
        return 0.0
    try:
        return float(value)
    except (ValueError, TypeError):
        return 0.0
