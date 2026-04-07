"""
Rotas de relatorios CGFR.
Espelha website/app/reports/routes.py — relatorio agrupado por natureza de despesa,
com filtros dinamicos por coluna, export Excel e versao PDF para impressao.
"""
import io
import logging
from datetime import datetime

from flask import render_template, request, send_file
from flask_login import login_required
from sqlalchemy import or_
from sqlalchemy.orm import joinedload

from app.cgfr.routes import cgfr_bp
from app.cgfr.models import CgfrProcessoEnviado, Acao
from app.cgfr.services.processo_service import ProcessoService, _format_record, _to_float
from app.models.nat_despesa import NatDespesa
from app.utils.permissions import requires_permission

logger = logging.getLogger(__name__)

# Colunas filtraveis (mesmas do sistema original)
FILTER_COLUMNS = [
    'processo_formatado', 'objeto_do_pedido', 'ultimo_andamento_sigla',
    'geracao_sigla', 'ultimo_andamento_descricao', 'natureza_despesa',
    'fornecedor', 'deliberacao', 'tipo_processo', 'data_da_reuniao',
]


def _nat_sort_key(name):
    """Extrai a descricao apos o codigo para ordenacao alfabetica."""
    parts = name.split('-', 1)
    return parts[1].strip().lower() if len(parts) > 1 else name.lower()


def _build_nat_cache():
    """Carrega todas as naturezas em dict {codigo: descricao}."""
    cache = {}
    try:
        for nat in NatDespesa.query.all():
            cache[nat.id] = f'{nat.codigo}-{nat.titulo}' if nat.titulo else str(nat.codigo)
    except Exception as e:
        logger.warning(f"Erro ao carregar cache de naturezas: {e}")
    return cache


def _get_processos_filtrados():
    """
    Logica compartilhada entre report_page, report_pdf e report_excel.
    Busca processos do MySQL com filtros e agrupa por natureza.
    Retorna: (naturezas_com, naturezas_sem, total_processos, valor_total, active_filters, filter_values)
    """
    search = request.args.get('search', '').strip()
    show_only_edited = request.args.get('showOnlyEdited', '0') == '1'

    # Coletar filtros por coluna (formato: filters[coluna][]=valor)
    active_filters = {}
    for col in FILTER_COLUMNS:
        values = request.args.getlist(f'filters[{col}][]')
        if values:
            active_filters[col] = values

    query = CgfrProcessoEnviado.query.filter(
        CgfrProcessoEnviado.processo_formatado.like(f'{CgfrProcessoEnviado.PREFIXO_SEAD}%')
    ).options(
        joinedload(CgfrProcessoEnviado.natureza_rel),
        joinedload(CgfrProcessoEnviado.fonte_rel),
    )

    # Busca textual
    if search:
        terms = search.split()
        for term in terms:
            t = f'%{term}%'
            query = query.filter(
                or_(
                    CgfrProcessoEnviado.processo_formatado.ilike(t),
                    CgfrProcessoEnviado.especificacao.ilike(t),
                    CgfrProcessoEnviado.tipo_processo.ilike(t),
                    CgfrProcessoEnviado.fornecedor.ilike(t),
                    CgfrProcessoEnviado.objeto_do_pedido.ilike(t),
                    CgfrProcessoEnviado.ultimo_andamento_sigla.ilike(t),
                    CgfrProcessoEnviado.geracao_sigla.ilike(t),
                    CgfrProcessoEnviado.ultimo_andamento_descricao.ilike(t),
                )
            )

    # Filtro: apenas editados/classificados
    if show_only_edited:
        query = query.filter(
            CgfrProcessoEnviado.natureza_despesa_id.isnot(None),
            CgfrProcessoEnviado.fonte_id.isnot(None),
            CgfrProcessoEnviado.acao_id.isnot(None),
        )

    # Filtro: possui reserva
    reserva_filter = request.args.get('reserva', '')
    if reserva_filter == 'sim':
        query = query.filter(CgfrProcessoEnviado.possui_reserva == 1)
    elif reserva_filter == 'nao':
        query = query.filter(
            or_(CgfrProcessoEnviado.possui_reserva == 0, CgfrProcessoEnviado.possui_reserva.is_(None))
        )

    # Filtro: nivel de prioridade
    prioridade_filter = request.args.get('prioridade', '')
    if prioridade_filter:
        query = query.filter(CgfrProcessoEnviado.nivel_prioridade == prioridade_filter)

    # ---------------------------------------------------------------
    # 1) Buscar TODOS os processos (sem filtros de coluna) para
    #    popular os dropdowns de filtro com valores completos
    # ---------------------------------------------------------------
    all_processos = query.order_by(CgfrProcessoEnviado.fornecedor).all()

    # Cache de acoes (usado tanto para filter_values quanto para dados)
    acao_ids = {p.acao_id for p in all_processos if p.acao_id}
    acao_map = {}
    if acao_ids:
        acoes = Acao.query.filter(Acao.id.in_(acao_ids)).all()
        acao_map = {a.id: a for a in acoes}

    # Cache de naturezas
    nat_cache = _build_nat_cache()

    # Formatar TODOS os registros e extrair valores para dropdowns
    all_formatted = []
    for p in all_processos:
        p_dict = _format_record(p, acao_map)
        if p.natureza_despesa_id and p.natureza_despesa_id in nat_cache:
            nat_descricao = nat_cache[p.natureza_despesa_id]
        else:
            nat_descricao = 'Não informado'
        p_dict['natureza_despesa'] = nat_descricao if nat_descricao != 'Não informado' else None
        p_dict['_nat_descricao'] = nat_descricao
        all_formatted.append(p_dict)

    # Extrair valores unicos de TODOS os registros (antes de filtrar)
    all_by_nat = {}
    for p_dict in all_formatted:
        nat = p_dict['_nat_descricao']
        if nat not in all_by_nat:
            all_by_nat[nat] = []
        all_by_nat[nat].append(p_dict)
    filter_values = _extract_filter_values(all_by_nat)

    # ---------------------------------------------------------------
    # 2) Aplicar filtros de coluna em memoria (rapido)
    # ---------------------------------------------------------------
    filtered = all_formatted
    for col, values in active_filters.items():
        val_set = set(values)
        filtered = [p for p in filtered if str(p.get(col, '') or '') in val_set]

    # Agrupar por natureza de despesa
    dados_por_natureza = {}
    total_processos = 0
    valor_total = 0.0

    for p_dict in filtered:
        nat_descricao = p_dict['_nat_descricao']
        if nat_descricao not in dados_por_natureza:
            dados_por_natureza[nat_descricao] = []
        dados_por_natureza[nat_descricao].append(p_dict)
        total_processos += 1
        valor_total += _to_float(p_dict.get('valor_solicitado'))

    # Separar em 'com natureza' e 'sem natureza'
    naturezas_com = dict(sorted(
        ((k, v) for k, v in dados_por_natureza.items() if k != 'Não informado'),
        key=lambda x: _nat_sort_key(x[0]),
    ))
    naturezas_sem = {k: v for k, v in dados_por_natureza.items() if k == 'Não informado'}

    return naturezas_com, naturezas_sem, total_processos, valor_total, active_filters, filter_values


def _extract_filter_values(dados_por_natureza):
    """Extrai valores unicos por coluna para popular os filtros dropdown no JS."""
    all_values = {col: set() for col in FILTER_COLUMNS}

    for processos in dados_por_natureza.values():
        for p in processos:
            for col in FILTER_COLUMNS:
                val = p.get(col)
                if val is not None and val != '' and val != '-':
                    all_values[col].add(str(val))

    return {col: sorted(vals) for col, vals in all_values.items()}


@cgfr_bp.route('/report/')
@login_required
@requires_permission('cgfr')
def report():
    """Pagina do relatorio com filtros e preview paginado — espelha o original."""
    naturezas_com, naturezas_sem, total_processos, valor_total, active_filters, filter_values = \
        _get_processos_filtrados()

    return render_template(
        'cgfr/reports/report.html',
        naturezas_com=naturezas_com,
        naturezas_sem=naturezas_sem,
        total_processos=total_processos,
        valor_total=valor_total,
        active_filters=active_filters,
        filter_values=filter_values,
        now_date=datetime.now().strftime('%d/%m/%Y %H:%M'),
    )


@cgfr_bp.route('/report/pdf')
@login_required
@requires_permission('cgfr')
def report_pdf():
    """Versao print-friendly do relatorio — abre em nova aba e dispara window.print()."""
    naturezas_com, naturezas_sem, total_processos, valor_total, active_filters, _ = \
        _get_processos_filtrados()

    return render_template(
        'cgfr/reports/report_pdf.html',
        naturezas_com=naturezas_com,
        naturezas_sem=naturezas_sem,
        total_processos=total_processos,
        valor_total=valor_total,
        active_filters=active_filters,
        now_date=datetime.now().strftime('%d/%m/%Y %H:%M'),
    )


@cgfr_bp.route('/report/excel')
@login_required
@requires_permission('cgfr')
def report_excel():
    """Exporta relatorio em formato Excel (.xlsx) com filtros aplicados."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    naturezas_com, naturezas_sem, total_processos, valor_total, active_filters, _ = \
        _get_processos_filtrados()

    wb = Workbook()
    ws = wb.active
    ws.title = 'Relatorio CGFR'

    # Estilos
    header_font = Font(bold=True, size=10, color='FFFFFF')
    header_fill = PatternFill(start_color='1E4D8C', end_color='1E4D8C', fill_type='solid')
    thin_border = Border(bottom=Side(style='thin', color='E2E8F0'))

    # Cabecalho
    col_headers = [
        'Natureza', 'Processo', 'Objeto do Pedido', 'Fornecedor',
        'Data Envio', 'Data Receb.', 'Data Reuniao',
        'Valor Solic.', 'Valor Aprov.', 'Deliberacao', 'Tipo Processo',
    ]
    for i, h in enumerate(col_headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(wrap_text=True, vertical='center')

    def _safe_str(val, default='-'):
        if val is None:
            return default
        s = str(val)
        return s if s else default

    def _to_float_safe(val):
        if val is None:
            return 0.0
        try:
            return float(val)
        except (ValueError, TypeError):
            return 0.0

    # Dados flat
    all_naturezas = list(naturezas_com.items()) + list(naturezas_sem.items())
    row = 2

    for natureza, processos in all_naturezas:
        for p in processos:
            ws.cell(row=row, column=1, value=natureza)
            ws.cell(row=row, column=2, value=_safe_str(p.get('protocolo_formatado'), ''))
            ws.cell(row=row, column=3, value=_safe_str(p.get('objeto_do_pedido'), ''))
            ws.cell(row=row, column=4, value=_safe_str(p.get('fornecedor'), ''))
            ws.cell(row=row, column=5, value=_safe_str(p.get('dt_enviado_fmt'), '-'))
            ws.cell(row=row, column=6, value=_safe_str(p.get('data_recebido_cgfr'), '-'))
            ws.cell(row=row, column=7, value=_safe_str(p.get('data_da_reuniao'), '-'))

            vs = _to_float_safe(p.get('valor_solicitado'))
            va = _to_float_safe(p.get('valor_aprovado'))
            c8 = ws.cell(row=row, column=8, value=vs if vs else None)
            c9 = ws.cell(row=row, column=9, value=va if va else None)
            if vs:
                c8.number_format = '#,##0.00'
            if va:
                c9.number_format = '#,##0.00'

            ws.cell(row=row, column=10, value=_safe_str(p.get('deliberacao'), 'Pendente'))
            ws.cell(row=row, column=11, value=_safe_str(p.get('tipo_processo'), ''))

            for col_idx in range(1, 12):
                ws.cell(row=row, column=col_idx).border = thin_border
            row += 1

    # Larguras das colunas
    col_widths = [30, 18, 38, 22, 15, 15, 15, 17, 17, 15, 15]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Auto-filtro
    if row > 2:
        ws.auto_filter.ref = f'A1:K{row - 1}'

    # Salvar e enviar
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"relatorio_cgfr_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename,
    )
