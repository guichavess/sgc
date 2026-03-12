"""
Rotas de API - Endpoints JSON para cascading dropdowns, busca de itens,
vinculação de itens ao contrato e exportação de relatórios.
"""
from io import BytesIO
from flask import jsonify, request, send_file
from flask_login import login_required, current_user

from app.prestacoes_contratos.routes import prestacoes_contratos_bp
from app.extensions import db
from app.services.prestacao_contrato_service import PrestacaoContratoService
from app.utils.permissions import requires_permission


# ========================================================
# CATSERV - Cascading: Seção → Divisão → Grupo → Classe
# ========================================================

@prestacoes_contratos_bp.route('/api/catserv/secoes')
@login_required
def api_catserv_secoes():
    """Lista todas as seções CATSERV."""
    from app.models.catserv import CatservSecao
    secoes = CatservSecao.query.order_by(CatservSecao.nome).all()
    return jsonify({
        'success': True,
        'data': [{'id': s.codigo_secao, 'nome': s.nome} for s in secoes]
    })


@prestacoes_contratos_bp.route('/api/catserv/divisoes/<int:secao_id>')
@login_required
def api_catserv_divisoes(secao_id):
    """Lista divisões de uma seção."""
    from app.models.catserv import CatservDivisao
    divisoes = CatservDivisao.query.filter_by(
        codigo_secao=secao_id
    ).order_by(CatservDivisao.nome).all()
    return jsonify({
        'success': True,
        'data': [{'id': d.codigo_divisao, 'nome': d.nome} for d in divisoes]
    })


@prestacoes_contratos_bp.route('/api/catserv/grupos/<int:divisao_id>')
@login_required
def api_catserv_grupos(divisao_id):
    """Lista grupos de uma divisão."""
    from app.models.catserv import CatservGrupo
    grupos = CatservGrupo.query.filter_by(
        codigo_divisao=divisao_id
    ).order_by(CatservGrupo.nome).all()
    return jsonify({
        'success': True,
        'data': [{'id': g.codigo_grupo, 'nome': g.nome} for g in grupos]
    })


@prestacoes_contratos_bp.route('/api/catserv/classes/<int:grupo_id>')
@login_required
def api_catserv_classes(grupo_id):
    """Lista classes de um grupo."""
    from app.models.catserv import CatservClasse
    classes = CatservClasse.query.filter_by(
        codigo_grupo=grupo_id
    ).order_by(CatservClasse.nome).all()
    return jsonify({
        'success': True,
        'data': [{'id': c.codigo_classe, 'nome': c.nome} for c in classes]
    })


@prestacoes_contratos_bp.route('/api/catserv/servicos')
@login_required
def api_catserv_servicos():
    """Lista serviços filtrados por classe_id, com busca e paginação."""
    from app.models.catserv import CatservServico

    classe_id = request.args.get('classe_id', type=int)
    search = request.args.get('search', '').strip()
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 50, type=int)

    if not classe_id:
        return jsonify({'success': False, 'error': 'classe_id é obrigatório'}), 400

    query = CatservServico.query.filter_by(codigo_classe=classe_id)

    if search:
        query = query.filter(CatservServico.nome.ilike(f'%{search}%'))

    query = query.order_by(CatservServico.nome)
    total = query.count()
    servicos = query.offset((page - 1) * per_page).limit(per_page).all()

    return jsonify({
        'success': True,
        'data': [{'id': s.codigo_servico, 'nome': s.nome} for s in servicos],
        'total': total,
        'page': page,
        'per_page': per_page
    })


# ========================================================
# CATMAT - Cascading: Grupo → Classe
# ========================================================

@prestacoes_contratos_bp.route('/api/catmat/grupos')
@login_required
def api_catmat_grupos():
    """Lista todos os grupos CATMAT."""
    from app.models.catmat import CatmatGrupo
    grupos = CatmatGrupo.query.order_by(CatmatGrupo.nome).all()
    return jsonify({
        'success': True,
        'data': [{'id': g.id, 'codigo': g.codigo, 'nome': g.nome} for g in grupos]
    })


@prestacoes_contratos_bp.route('/api/catmat/classes/<int:grupo_codigo>')
@login_required
def api_catmat_classes(grupo_codigo):
    """Lista classes de um grupo CATMAT (filtrado por codigo do grupo)."""
    from app.models.catmat import CatmatClasse
    classes = CatmatClasse.query.filter_by(
        codigo_grupo=grupo_codigo
    ).order_by(CatmatClasse.nome).all()
    return jsonify({
        'success': True,
        'data': [{'id': c.id, 'codigo': c.codigo, 'nome': c.nome} for c in classes]
    })


@prestacoes_contratos_bp.route('/api/catmat/pdms/<int:classe_codigo>')
@login_required
def api_catmat_pdms(classe_codigo):
    """Lista PDMs de uma classe CATMAT (filtrado por codigo da classe)."""
    from app.models.catmat import CatmatPdm
    pdms = CatmatPdm.query.filter_by(
        codigo_classe=classe_codigo
    ).order_by(CatmatPdm.nome).all()
    return jsonify({
        'success': True,
        'data': [{'id': p.id, 'codigo': p.codigo, 'nome': p.nome} for p in pdms]
    })


@prestacoes_contratos_bp.route('/api/catmat/itens')
@login_required
def api_catmat_itens():
    """Lista itens CATMAT filtrados por pdm_id ou classe_id, com busca e paginação.

    Se pdm_id fornecido: filtra itens direto pelo PDM (tipificação até PDM).
    Senão: fallback por classe_id → PDMs → itens.
    """
    from app.models.catmat import CatmatItem, CatmatPdm

    pdm_id = request.args.get('pdm_id', type=int)
    classe_id = request.args.get('classe_id', type=int)
    search = request.args.get('search', '').strip()
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 50, type=int)

    if pdm_id:
        # Filtro direto pelo PDM tipificado
        pdm = db.session.get(CatmatPdm, pdm_id)
        if not pdm:
            return jsonify({'success': False, 'error': 'PDM não encontrado'}), 404
        query = CatmatItem.query.filter(CatmatItem.codigo_pdm == pdm.codigo)
    elif classe_id:
        # Fallback: filtro por classe (todos os PDMs da classe)
        from app.models.catmat import CatmatClasse
        classe = db.session.get(CatmatClasse, classe_id)
        if not classe:
            return jsonify({'success': False, 'error': 'Classe não encontrada'}), 404
        pdm_codigos = db.session.query(CatmatPdm.codigo).filter(
            CatmatPdm.codigo_classe == classe.codigo
        ).subquery()
        query = CatmatItem.query.filter(CatmatItem.codigo_pdm.in_(pdm_codigos))
    else:
        return jsonify({'success': False, 'error': 'pdm_id ou classe_id é obrigatório'}), 400

    if search:
        query = query.filter(CatmatItem.descricao.ilike(f'%{search}%'))

    query = query.order_by(CatmatItem.descricao)
    total = query.count()
    itens = query.offset((page - 1) * per_page).limit(per_page).all()

    return jsonify({
        'success': True,
        'data': [{'id': i.id, 'codigo': i.codigo, 'descricao': i.descricao} for i in itens],
        'total': total,
        'page': page,
        'per_page': per_page
    })


# ========================================================
# VINCULAÇÃO DE ITENS AO CONTRATO
# ========================================================

@prestacoes_contratos_bp.route('/api/contratos/<codigo>/itens-vinculados')
@login_required
def api_listar_vinculados(codigo):
    """Lista itens vinculados a um contrato."""
    try:
        tipo = request.args.get('tipo')  # Opcional: 'S' ou 'M'
        itens = PrestacaoContratoService.listar_itens_vinculados(codigo, tipo=tipo)
        return jsonify({
            'success': True,
            'data': [
                {
                    'vinculacao_id': item['vinculacao_id'],
                    'tipo': item['tipo'],
                    'item_id': item.get('item_id'),
                    'codigo': item.get('codigo'),
                    'descricao': item.get('descricao'),
                    'associacao': item.get('associacao'),
                    'item_contrato_descricao': item.get('item_contrato_descricao'),
                    'data_vinculacao': item['data_vinculacao'].strftime('%d/%m/%Y %H:%M') if item.get('data_vinculacao') else None
                }
                for item in itens
            ],
            'total': len(itens)
        })
    except Exception as e:
        return jsonify({'success': False, 'message': f'Erro: {str(e)}'}), 500


@prestacoes_contratos_bp.route('/api/itens-contrato/buscar')
@login_required
def api_itens_contrato_buscar():
    """Busca itens da tabela itens_contrato com autocomplete.
    Filtra pela tipificação do contrato (catserv_classe_id / catmat_pdm_id).
    Retorna dados do item + associação CATMAT/CATSERV (de-para) quando existir."""
    from app.models.item_contrato import ItemContrato
    from app.models.catmat import CatmatItem, CatmatPdm
    from app.models.catserv import CatservServico
    from sqlalchemy import or_

    search = request.args.get('search', '').strip()
    per_page = request.args.get('per_page', 20, type=int)

    # Parâmetros de tipificação do contrato (para filtragem)
    catserv_classe_id = request.args.get('catserv_classe_id', type=int)
    catmat_pdm_id = request.args.get('catmat_pdm_id', type=int)
    tipo_contrato = request.args.get('tipo_contrato', '').strip() or None

    if len(search) < 2:
        return jsonify({'success': False, 'error': 'Digite pelo menos 2 caracteres'}), 400

    # Busca pelo texto
    query = ItemContrato.query.filter(
        ItemContrato.descricao.ilike(f'%{search}%')
    )

    # Aplicar filtro de tipificação (mesma lógica do service)
    ids_servicos_validos = set()
    ids_catmat_validos = set()

    if tipo_contrato in ('SERVICO', 'MISTO') and catserv_classe_id:
        servicos = CatservServico.query.filter_by(
            codigo_classe=catserv_classe_id
        ).all()
        ids_servicos_validos = {s.codigo_servico for s in servicos}

    if tipo_contrato in ('MATERIAL', 'MISTO') and catmat_pdm_id:
        # catmat_pdm_id no contrato é o id (autoincrement) do CatmatPdm.
        # catmat_item_id em itens_contrato armazena catmat_itens.id.
        # Buscar todos os catmat_itens que pertencem a esse PDM.
        pdm = db.session.get(CatmatPdm, catmat_pdm_id)
        if pdm:
            itens_catmat = CatmatItem.query.filter_by(
                codigo_pdm=pdm.codigo
            ).all()
            ids_catmat_validos = {i.id for i in itens_catmat}

    tem_filtro = bool(ids_servicos_validos or ids_catmat_validos)

    if tem_filtro:
        condicoes = []
        if ids_servicos_validos:
            condicoes.append(
                ItemContrato.catserv_servico_id.in_(ids_servicos_validos)
            )
        if ids_catmat_validos:
            condicoes.append(
                ItemContrato.catmat_item_id.in_(ids_catmat_validos)
            )
        if condicoes:
            query = query.filter(or_(*condicoes))
        else:
            # Nenhum ID válido — não retorna nada
            query = query.filter(db.literal(False))
    else:
        # Sem tipificação — só mostra itens com de-para preenchido
        query = query.filter(
            or_(
                ItemContrato.catserv_servico_id.isnot(None),
                ItemContrato.catmat_item_id.isnot(None)
            )
        )

    query = query.order_by(ItemContrato.descricao).limit(per_page)
    itens = query.all()

    resultado = []
    for item in itens:
        associacao = None

        # Verificar de-para CATSERV
        if item.catserv_servico_id:
            srv = db.session.get(CatservServico, item.catserv_servico_id)
            if srv:
                associacao = f'{srv.codigo_servico} - {srv.nome}'

        # Verificar de-para CATMAT (catmat_item_id armazena catmat_itens.id)
        if not associacao and item.catmat_item_id:
            mat = db.session.get(CatmatItem, item.catmat_item_id)
            if mat:
                associacao = f'{mat.codigo} - {mat.descricao}'

        resultado.append({
            'id': item.id,
            'descricao': item.descricao,
            'tipo_item': item.tipo_item,
            'catserv_servico_id': item.catserv_servico_id,
            'catmat_item_id': item.catmat_item_id,
            'associacao': associacao
        })

    return jsonify({'success': True, 'data': resultado})


@prestacoes_contratos_bp.route('/api/contratos/<codigo>/vincular-item', methods=['POST'])
@login_required
@requires_permission('prestacoes_contratos.criar')
def api_vincular_item(codigo):
    """Vincula um item ao contrato (via item_contrato_id)."""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'message': 'Dados inválidos.'}), 400

        item_contrato_id = data.get('item_contrato_id')

        if not item_contrato_id:
            return jsonify({'success': False, 'message': 'item_contrato_id é obrigatório.'}), 400

        success, message, vinculacao_id = PrestacaoContratoService.vincular_item(
            codigo_contrato=codigo,
            item_contrato_id=item_contrato_id,
            usuario_id=current_user.id
        )

        if success:
            return jsonify({
                'success': True,
                'message': message,
                'vinculacao_id': vinculacao_id
            })
        else:
            return jsonify({'success': False, 'message': message}), 400

    except Exception as e:
        return jsonify({'success': False, 'message': f'Erro: {str(e)}'}), 500


@prestacoes_contratos_bp.route('/api/contratos/<codigo>/vincular-itens-batch', methods=['POST'])
@login_required
@requires_permission('prestacoes_contratos.criar')
def api_vincular_itens_batch(codigo):
    """Vincula múltiplos itens ao contrato de uma vez."""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'message': 'Dados inválidos.'}), 400

        item_contrato_ids = data.get('item_contrato_ids', [])

        if not item_contrato_ids or not isinstance(item_contrato_ids, list):
            return jsonify({'success': False, 'message': 'Selecione pelo menos um item.'}), 400

        vinculados = 0
        erros = []

        for item_id in item_contrato_ids:
            success, message, _ = PrestacaoContratoService.vincular_item(
                codigo_contrato=codigo,
                item_contrato_id=item_id,
                usuario_id=current_user.id
            )
            if success:
                vinculados += 1
            else:
                erros.append(message)

        if vinculados > 0:
            msg = f'{vinculados} item(ns) vinculado(s) com sucesso.'
            if erros:
                msg += f' {len(erros)} ignorado(s) (já vinculados ou inválidos).'
            return jsonify({'success': True, 'message': msg, 'vinculados': vinculados})
        else:
            return jsonify({
                'success': False,
                'message': erros[0] if erros else 'Nenhum item vinculado.'
            }), 400

    except Exception as e:
        return jsonify({'success': False, 'message': f'Erro: {str(e)}'}), 500


@prestacoes_contratos_bp.route('/api/contratos/<codigo>/vincular-catalogo', methods=['POST'])
@login_required
@requires_permission('prestacoes_contratos.criar')
def api_vincular_catalogo(codigo):
    """Vincula itens do catálogo (serviços CATSERV ou PDMs CATMAT) ao contrato."""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'message': 'Dados inválidos.'}), 400

        itens = data.get('itens', [])
        if not itens or not isinstance(itens, list):
            return jsonify({'success': False, 'message': 'Selecione pelo menos um item.'}), 400

        vinculados = 0
        erros = []

        for item in itens:
            tipo = item.get('tipo')
            catalogo_id = item.get('id')
            if not tipo or not catalogo_id:
                erros.append('Item inválido.')
                continue

            success, message, _ = PrestacaoContratoService.vincular_catalogo_item(
                codigo_contrato=codigo,
                tipo=tipo,
                catalogo_id=int(catalogo_id),
                usuario_id=current_user.id
            )
            if success:
                vinculados += 1
            else:
                erros.append(message)

        if vinculados > 0:
            msg = f'{vinculados} item(ns) vinculado(s) com sucesso.'
            if erros:
                msg += f' {len(erros)} ignorado(s) (já vinculados ou inválidos).'
            return jsonify({'success': True, 'message': msg, 'vinculados': vinculados})
        else:
            return jsonify({
                'success': False,
                'message': erros[0] if erros else 'Nenhum item vinculado.'
            }), 400

    except Exception as e:
        return jsonify({'success': False, 'message': f'Erro: {str(e)}'}), 500


@prestacoes_contratos_bp.route('/api/itens-vinculados/<int:vinculacao_id>', methods=['DELETE'])
@login_required
@requires_permission('prestacoes_contratos.excluir')
def api_desvincular_item(vinculacao_id):
    """Remove vinculação de um item do contrato."""
    try:
        success, message, warning = PrestacaoContratoService.desvincular_item(vinculacao_id)

        if success:
            return jsonify({
                'success': True,
                'message': message,
                'warning': warning
            })
        else:
            return jsonify({'success': False, 'message': message}), 400

    except Exception as e:
        return jsonify({'success': False, 'message': f'Erro: {str(e)}'}), 500


# ========================================================
# FILTROS ENCADEADOS: Natureza → Tipo Patrimonial → SubItem
# ========================================================

@prestacoes_contratos_bp.route('/api/filtro/tipos-patrimoniais')
@login_required
def api_filtro_tipos_patrimoniais():
    """Retorna Tipos Patrimoniais filtrados por natureza (para cascading dropdown)."""
    natureza = request.args.get('natureza', '').strip()
    dados = PrestacaoContratoService.listar_tipos_patrimoniais_por_natureza(
        natureza_codigo=natureza or None
    )
    return jsonify({'success': True, 'data': dados})


@prestacoes_contratos_bp.route('/api/filtro/subitens')
@login_required
def api_filtro_subitens():
    """Retorna SubItens filtrados por natureza e tipo patrimonial (para cascading dropdown)."""
    natureza = request.args.get('natureza', '').strip()
    tipo_patrimonial = request.args.get('tipo_patrimonial', '').strip()
    dados = PrestacaoContratoService.listar_subitens_por_natureza(
        natureza_codigo=natureza or None,
        tipo_patrimonial_codigo=tipo_patrimonial or None
    )
    return jsonify({'success': True, 'data': dados})


# ========================================================
# EXPORTAÇÃO EXCEL — Relatório Financeiro do Contrato
# ========================================================

@prestacoes_contratos_bp.route('/api/contratos/<codigo>/relatorio-financeiro')
@login_required
def api_relatorio_financeiro(codigo):
    """Gera Excel com 4 abas (Empenhos, Liquidações, PD, Pagamentos) para o contrato."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    empenhos = PrestacaoContratoService.listar_empenhos(codigo)
    liquidacoes = PrestacaoContratoService.listar_liquidacoes(codigo)
    pds = PrestacaoContratoService.listar_pds(codigo)
    pagamentos = PrestacaoContratoService.listar_pagamentos_contrato(codigo)

    wb = Workbook()

    # ── Estilos ──
    header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='343990', end_color='343990', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0')
    )
    money_fmt = '#,##0.00'

    def style_header(ws, headers):
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

    def auto_width(ws, min_width=10, max_width=45):
        for col in ws.columns:
            col_letter = get_column_letter(col[0].column)
            lengths = []
            for cell in col:
                if cell.value:
                    lengths.append(len(str(cell.value)))
            width = max(lengths) if lengths else min_width
            ws.column_dimensions[col_letter].width = min(max(width + 2, min_width), max_width)

    def fmt_date(dt):
        if dt:
            return dt.strftime('%d/%m/%Y')
        return ''

    def get_ano(dt):
        if dt:
            return dt.year
        return ''

    # ── Aba 1: Empenhos ──
    ws_emp = wb.active
    ws_emp.title = 'Empenhos'
    emp_headers = ['Ano', 'Código', 'Credor', 'Data Emissão', 'Status', 'Valor']
    style_header(ws_emp, emp_headers)

    for row_idx, emp in enumerate(empenhos, 2):
        ws_emp.cell(row=row_idx, column=1, value=get_ano(emp.dataEmissao)).border = thin_border
        ws_emp.cell(row=row_idx, column=2, value=emp.codigo or '').border = thin_border
        ws_emp.cell(row=row_idx, column=3, value=emp.nomeCredor or '').border = thin_border
        ws_emp.cell(row=row_idx, column=4, value=fmt_date(emp.dataEmissao)).border = thin_border
        ws_emp.cell(row=row_idx, column=5, value=emp.statusDocumento or '').border = thin_border
        c = ws_emp.cell(row=row_idx, column=6, value=emp.valor or 0)
        c.number_format = money_fmt
        c.border = thin_border

    auto_width(ws_emp)

    # ── Aba 2: Liquidações ──
    ws_liq = wb.create_sheet('Liquidações')
    liq_headers = ['Ano', 'Código', 'Credor', 'Data Emissão', 'NE Vinculada', 'Valor']
    style_header(ws_liq, liq_headers)

    for row_idx, liq in enumerate(liquidacoes, 2):
        ws_liq.cell(row=row_idx, column=1, value=get_ano(liq.dataEmissao)).border = thin_border
        ws_liq.cell(row=row_idx, column=2, value=liq.codigo or '').border = thin_border
        ws_liq.cell(row=row_idx, column=3, value=liq.nomeCredor or '').border = thin_border
        ws_liq.cell(row=row_idx, column=4, value=fmt_date(liq.dataEmissao)).border = thin_border
        ws_liq.cell(row=row_idx, column=5, value=liq.codigoEmpenhoVinculado or '').border = thin_border
        c = ws_liq.cell(row=row_idx, column=6, value=liq.valor or 0)
        c.number_format = money_fmt
        c.border = thin_border

    auto_width(ws_liq)

    # ── Aba 3: PD (Programação de Desembolso) ──
    ws_pd = wb.create_sheet('PD')
    pd_headers = ['Ano', 'Código', 'Credor', 'Data Emissão', 'NE Vinculada', 'Status', 'Valor']
    style_header(ws_pd, pd_headers)

    for row_idx, pd_item in enumerate(pds, 2):
        ws_pd.cell(row=row_idx, column=1, value=get_ano(pd_item.dataEmissao)).border = thin_border
        ws_pd.cell(row=row_idx, column=2, value=pd_item.codigo or '').border = thin_border
        ws_pd.cell(row=row_idx, column=3, value=pd_item.nomeCredor or '').border = thin_border
        ws_pd.cell(row=row_idx, column=4, value=fmt_date(pd_item.dataEmissao)).border = thin_border
        ws_pd.cell(row=row_idx, column=5, value=pd_item.codigoNE or '').border = thin_border
        ws_pd.cell(row=row_idx, column=6, value=pd_item.statusExecucao or '').border = thin_border
        c = ws_pd.cell(row=row_idx, column=7, value=pd_item.valor or 0)
        c.number_format = money_fmt
        c.border = thin_border

    auto_width(ws_pd)

    # ── Aba 4: Pagamentos ──
    ws_pag = wb.create_sheet('Pagamentos')
    pag_headers = ['Ano', 'Código', 'Credor', 'Data Emissão', 'Data Pagamento', 'Tipo OB', 'Valor']
    style_header(ws_pag, pag_headers)

    for row_idx, pag in enumerate(pagamentos, 2):
        ws_pag.cell(row=row_idx, column=1, value=get_ano(pag.dataEmissao)).border = thin_border
        ws_pag.cell(row=row_idx, column=2, value=pag.codigo or '').border = thin_border
        ws_pag.cell(row=row_idx, column=3, value=pag.nomeCredor or '').border = thin_border
        ws_pag.cell(row=row_idx, column=4, value=fmt_date(pag.dataEmissao)).border = thin_border
        ws_pag.cell(row=row_idx, column=5, value=fmt_date(pag.dataPagamento)).border = thin_border
        ws_pag.cell(row=row_idx, column=6, value=pag.tipoOB or '').border = thin_border
        c = ws_pag.cell(row=row_idx, column=7, value=pag.valor or 0)
        c.number_format = money_fmt
        c.border = thin_border

    auto_width(ws_pag)

    # ── Gerar arquivo em memória ──
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f'relatorio_financeiro_{codigo}.xlsx'
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


# ========================================================
# RELATÓRIO EXCEL — LISTA GERAL DE CONTRATOS
# ========================================================

@prestacoes_contratos_bp.route('/api/relatorio-contratos')
@login_required
def api_relatorio_contratos():
    """Gera Excel com a lista geral de contratos (mesmos dados da tabela principal + SubItens)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from app.models.contrato import Contrato
    from app.models.empenho_item import EmpenhoItem, ClassSubItemDespesa, ClassTipoPatrimonial
    from app.models.class_fonte import ClassFonte
    from sqlalchemy import func, literal_column

    # ── Buscar TODOS os contratos (sem paginação) ──
    contratos = Contrato.query.order_by(Contrato.codigo.desc()).all()

    # ── Buscar classificadores em batch (mesmo padrão do dashboard) ──
    codigos = [c.codigo for c in contratos]
    mapa_naturezas = PrestacaoContratoService.buscar_naturezas_por_contratos(codigos)
    mapa_classif = PrestacaoContratoService.buscar_classificadores_por_contratos(codigos)

    # ── Buscar SubItens por contrato (via empenho_itens) ──
    EXCLUDE_NATUREZA = {'339092', '449092'}
    codigos_int = []
    for c in codigos:
        try:
            codigos_int.append(int(c))
        except (ValueError, TypeError):
            pass

    mapa_subitens = {}
    if codigos_int:
        rows_sub = db.session.query(
            EmpenhoItem.CodContrato,
            literal_column(
                "GROUP_CONCAT(DISTINCT empenho_itens.SubItemDespesa SEPARATOR '||')"
            ).label('subitens_raw')
        ).filter(
            EmpenhoItem.CodContrato.in_(codigos_int),
            EmpenhoItem.Natureza.notin_(EXCLUDE_NATUREZA)
        ).group_by(EmpenhoItem.CodContrato).all()

        # Coletar todos os códigos de subitem para resolver nomes
        todos_cod_sub = set()
        for r in rows_sub:
            if r.subitens_raw:
                for val in r.subitens_raw.split('||'):
                    todos_cod_sub.add(val.strip())

        # Resolver nomes
        mapa_subitem_nome = {}
        if todos_cod_sub:
            for s in ClassSubItemDespesa.query.all():
                cod1 = str(s.valoresClassificador1 or '').strip()
                cod2 = str(s.valoresClassificador2 or '').strip()
                cod_completo = f"{cod1}.{cod2}" if cod2 else cod1
                mapa_subitem_nome[cod_completo] = s.nomeClassificador

        for r in rows_sub:
            cod_str = str(r.CodContrato)
            subitens_list = []
            if r.subitens_raw:
                codigos_vistos = set()
                for val in r.subitens_raw.split('||'):
                    val = val.strip()
                    if val and val not in codigos_vistos:
                        codigos_vistos.add(val)
                        nome = mapa_subitem_nome.get(val, val)
                        subitens_list.append(f"{val} - {nome}")
            mapa_subitens[cod_str] = subitens_list

    # ── Montar Excel ──
    wb = Workbook()
    ws = wb.active
    ws.title = 'Contratos'

    # Estilos
    header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='343990', end_color='343990', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0')
    )
    wrap_align = Alignment(vertical='top', wrap_text=True)
    money_fmt = '#,##0.00'

    headers = [
        'Código', 'Número Original', 'Contratado', 'Objeto', 'Situação',
        'Natureza', 'Tipo de Despesa', 'Tipo Execução', 'Centro de Custo',
        'Valor', 'SubItens de Despesa'
    ]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for row_idx, contrato in enumerate(contratos, 2):
        cod = contrato.codigo

        # Naturezas
        nats = mapa_naturezas.get(cod, [])
        nat_txt = '\n'.join(f"{n['codigo']} - {n['titulo']}" for n in nats) if nats else ''

        # Tipo de despesa (TipoPatrimonial)
        classif = mapa_classif.get(cod, {})
        tps = classif.get('tipos_patrimoniais', [])
        tp_txt = '\n'.join(f"{t['codigo']} - {t['nome']}" for t in tps) if tps else ''

        # SubItens
        subs = mapa_subitens.get(cod, [])
        sub_txt = '\n'.join(subs) if subs else ''

        # Tipo Execução e Centro de Custo
        tipo_exec = contrato.tipo_execucao.descricao if contrato.tipo_execucao else ''
        centro = contrato.centro_de_custo.descricao if contrato.centro_de_custo else ''

        ws.cell(row=row_idx, column=1, value=cod).border = thin_border
        ws.cell(row=row_idx, column=2, value=contrato.numeroOriginal or '').border = thin_border
        ws.cell(row=row_idx, column=3, value=contrato.nomeContratado or '').border = thin_border

        c_obj = ws.cell(row=row_idx, column=4, value=contrato.objeto or '')
        c_obj.border = thin_border
        c_obj.alignment = wrap_align

        ws.cell(row=row_idx, column=5, value=contrato.situacao or '').border = thin_border

        c_nat = ws.cell(row=row_idx, column=6, value=nat_txt)
        c_nat.border = thin_border
        c_nat.alignment = wrap_align

        c_tp = ws.cell(row=row_idx, column=7, value=tp_txt)
        c_tp.border = thin_border
        c_tp.alignment = wrap_align

        ws.cell(row=row_idx, column=8, value=tipo_exec).border = thin_border
        ws.cell(row=row_idx, column=9, value=centro).border = thin_border

        c_val = ws.cell(row=row_idx, column=10, value=contrato.valor or 0)
        c_val.number_format = money_fmt
        c_val.border = thin_border

        c_sub = ws.cell(row=row_idx, column=11, value=sub_txt)
        c_sub.border = thin_border
        c_sub.alignment = wrap_align

    # Auto-ajustar largura das colunas
    col_widths = {1: 12, 2: 14, 3: 35, 4: 50, 5: 14, 6: 30, 7: 30, 8: 16, 9: 20, 10: 16, 11: 45}
    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Filtro automático
    ws.auto_filter.ref = f"A1:K{len(contratos) + 1}"

    # ── Gerar arquivo ──
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    from datetime import date
    filename = f'relatorio_contratos_{date.today().strftime("%Y%m%d")}.xlsx'
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


# ========================================================
# PLANEJAMENTO ORÇAMENTÁRIO - Excel por contrato
# ========================================================

@prestacoes_contratos_bp.route('/contratos/<int:codigo>/planejamento/excel')
@login_required
def api_planejamento_excel(codigo):
    """Gera Excel com planejamento orçamentário do contrato."""
    from datetime import date
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from app.models.planejamento_orcamentario import PlanejamentoOrcamentario

    dados = PlanejamentoOrcamentario.query.filter_by(
        cod_contrato=str(codigo)
    ).order_by(PlanejamentoOrcamentario.competencia).all()

    # Buscar liquidado por mês para este contrato
    from app.extensions import db
    from sqlalchemy import text as sa_text
    from decimal import Decimal
    liq_map = {}
    cod_numerico = str(codigo).replace('.', '').replace('/', '')
    try:
        cod_int = int(cod_numerico)
        sql = sa_text("""
            SELECT MONTH(dataEmissao) AS mes, YEAR(dataEmissao) AS ano,
                   SUM(CASE WHEN tipoAlteracao = 'ANULACAO' THEN -valor ELSE valor END) AS total
            FROM liquidacao
            WHERE statusDocumento = 'CONTABILIZADO'
              AND codigoUG = '210101'
              AND codContrato = :cod
            GROUP BY YEAR(dataEmissao), MONTH(dataEmissao)
        """)
        for row in db.session.execute(sql, {'cod': cod_int}).fetchall():
            chave = f'{int(row[0]):02d}/{int(row[1])}'
            liq_map[chave] = float(row[2]) if row[2] else 0.0
    except (ValueError, TypeError):
        pass

    wb = Workbook()
    ws = wb.active
    ws.title = 'Planejamento'

    header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='343990', end_color='343990', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0')
    )
    money_fmt = '#,##0.00'

    headers = ['Competência', 'Valor Planejado', 'Valor Pago', 'Planejamento Inicial', 'Repactuação/Prorrogação']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    for row_idx, p in enumerate(dados, 2):
        ws.cell(row=row_idx, column=1, value=p.competencia).border = thin_border
        c = ws.cell(row=row_idx, column=2, value=float(p.valor or 0))
        c.number_format = money_fmt
        c.border = thin_border
        liq_val = liq_map.get(p.competencia, 0.0)
        c = ws.cell(row=row_idx, column=3, value=liq_val)
        c.number_format = money_fmt
        c.border = thin_border
        ws.cell(row=row_idx, column=4, value='Sim' if p.planejamento_inicial else 'Não').border = thin_border
        ws.cell(row=row_idx, column=5, value='Sim' if p.repactuacao_prorrogacao else 'Não').border = thin_border

    # Total
    total_row = len(dados) + 2
    c = ws.cell(row=total_row, column=1, value='Total')
    c.font = Font(bold=True)
    c.border = thin_border
    total_val = sum(float(p.valor or 0) for p in dados)
    c = ws.cell(row=total_row, column=2, value=total_val)
    c.number_format = money_fmt
    c.font = Font(bold=True, color='2E7D32')
    c.border = thin_border
    total_liq = sum(liq_map.get(p.competencia, 0.0) for p in dados)
    c = ws.cell(row=total_row, column=3, value=total_liq)
    c.number_format = money_fmt
    c.font = Font(bold=True, color='1565C0')
    c.border = thin_border

    for col_idx in range(1, 6):
        ws.column_dimensions[get_column_letter(col_idx)].width = 22

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f'planejamento_{codigo}_{date.today().strftime("%Y%m%d")}.xlsx'
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )
